import foxinfo_share_utility.share_api as share_api
import foxinfo_share_utility.share_ui as share_ui
import foxinfo_share_utility.share_icon as share_icon
from DownloadData import Download
import requests
from bs4 import BeautifulSoup
import json
import os
import shutil
import sys
import datetime
import time
import math
import copy
import re
import logging
from collections import deque
from logging.handlers import TimedRotatingFileHandler
from QtStockPriceMainWindow import Ui_MainWindow  # 導入轉換後的 UI 類
from QtStockCapitalIncreaseEditDialog import Ui_Dialog as Ui_StockCapitalIncreaseDialog
from QtStockTradingEditDialog import Ui_Dialog as Ui_StockTradingDialog
from QtStockRegularTradingEditDialog import Ui_Dialog as Ui_StockRegularTradingDialog
from QtStockDividendEditDialog import Ui_Dialog as Ui_StockDividendDialog
from QtStockCapitalReductionEditDialog import Ui_Dialog as Ui_StockCapitalReductionDialog
from QtStockSplitEditDialog import Ui_Dialog as Ui_StockSplitDialog
from QtStockDividendPositionDateEditDialog import Ui_Dialog as Ui_StockDividendPositionDateEditDialog
from QtCashTransferEditDialog import Ui_Dialog as Ui_CashTransferDialog
from QtDuplicateOptionDialog import Ui_Dialog as Ui_DuplicateOptionDialog
from QtStockDividendTransferFeeEditSpinboxDialog import Ui_Dialog as Ui_StockDividendTransferFeeEditSpinboxDialog
from QtAboutDialog import Ui_Dialog as Ui_AboutDialog
from QtAccountSettingEditDialog import Ui_Dialog as Ui_AccountSettingEditDialog
from QtShowItemEditDialog import Ui_Dialog as Ui_ShowItemEditDialog
from PySide6.QtWidgets import QApplication, QMainWindow, QDialog, QButtonGroup, QMessageBox, QHeaderView, QVBoxLayout, QHBoxLayout, \
                              QLabel, QLineEdit, QTabBar, QWidget, QTableView, QComboBox, QPushButton, QSizePolicy, QSpacerItem, QCheckBox, \
                              QProgressBar, QTabWidget, QMenu, QListView, QAbstractItemView, QListWidget, QListWidgetItem
from PySide6.QtGui import QStandardItemModel, QStandardItem, QBrush
from PySide6.QtCore import Qt, QModelIndex, QSignalBlocker, QSize, QThread, QObject, Signal, QSettings, QPoint, QItemSelection, QItemSelectionModel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from enum import Enum, IntEnum, auto
from decimal import Decimal, ROUND_HALF_UP, ROUND_CEILING
from scipy.optimize import newton
from typing import Union

g_user_dir = os.path.expanduser("~")                     
g_exe_dir = os.path.dirname(__file__)                    
g_exe2_dir = os.path.dirname( sys.executable )           
g_abs_dir = os.path.dirname( os.path.abspath(__file__) ) 
print( "g_user_dir :" + g_user_dir ) #開發模式跟打包模式下都是C:\Users\foxki
print( "g_exe_dir :" + g_exe_dir )   #開發模式下是D:\_2.code\PythonStockPrice                              #打包模式後是 D:\_2.code\PythonStockPrice\dist\StockPriceMainWindow\_internal 
print( "g_exe2_dir :" + g_exe2_dir ) #開發模式下是C:\Users\foxki\AppData\Local\Programs\Python\Python312   #打包模式後是:D:\_2.code\PythonStockPrice\dist\StockPriceMainWindow
print( "g_abs_dir :" + g_abs_dir )   #開發模式下是D:\_2.code\PythonStockPrice                              #打包模式後是 D:\_2.code\PythonStockPrice\dist\StockPriceMainWindow\_internal 

reg_settings = QSettings( "FoxInfo", "StockInventory" )

if getattr( sys, 'frozen', False ):
    # PyInstaller 打包後執行時

    # region 設定 錯誤資訊 logging
    # 設定日誌
    logger = logging.getLogger()
    logger.setLevel(logging.ERROR)

    current_date = datetime.datetime.today().strftime("%Y-%m-%d")
    log_filename = f"log_{current_date}.txt"
    # 日誌檔案處理器
    file_handler = TimedRotatingFileHandler(
        filename = log_filename,  # 基本文件名
        when = "midnight",  # 依據每天午夜分割日誌
        interval = 1,  # 每 1 天生成一個新檔案
        backupCount = 7,  # 保留最近 7 天的日誌檔案
        encoding = "utf-8"  # 確保支援 UTF-8 編碼
    )
    file_handler.setLevel(logging.ERROR)
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(file_formatter)

    # 終端處理器
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.ERROR)
    console_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    console_handler.setFormatter(console_formatter)

    # 添加處理器到 logger
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    # 自定義未處理例外的鉤子
    def handle_exception(exc_type, exc_value, exc_traceback):
        if issubclass(exc_type, KeyboardInterrupt):
            # 如果是鍵盤中斷，保持預設行為
            sys.__excepthook__(exc_type, exc_value, exc_traceback)
            return
        # 記錄例外
        logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback))
    # 將自定義的鉤子設定為全局的未處理例外處理器
    sys.excepthook = handle_exception
    # endregion

    g_exe_root_dir = os.path.dirname(__file__) #使用--onefile打包 C:\Users\foxki\AppData\Local\Temp\_MEI60962 否則就是 D:\_2.code\PythonStockPrice\dist\StockPriceMainWindow\_internal
    g_data_dir = reg_settings.value( "TemporaryFolderPath", os.path.join( g_user_dir, "AppData", "Local", "FoxInfo" ) ) #C:\Users\foxki\AppData\Local\FoxInfo 
else:
    # VSCode執行 Python 腳本時
    g_exe_root_dir = os.path.dirname( os.path.abspath(__file__) )
    g_data_dir = reg_settings.value( "TemporaryFolderPathDebug", g_exe_root_dir ) #D:\_2.code\PythonStockPrice

styles_css_path = os.path.join( g_exe_root_dir, 'resources\\styles.css' ) 

#region enum相關
class TradingType( IntEnum ):
    TEMPLATE = 0
    SELL = 1
    BUY = 2
    REGULAR_BUY = 3
    CAPITAL_INCREASE = 4
    DIVIDEND = 5
    CAPITAL_REDUCTION = 6
    SPLIT = 7
    MERGE = 8

class TradingPriceType( Enum ):
    PER_SHARE = 0
    TOTAL = 1

class TradingFeeType( Enum ):
    VARIABLE = 0
    CONSTANT = 1

class DividendValueType( Enum ):
    PER_SHARE = 0
    TOTAL = 1

class CapitalReductionType( Enum ):
    CASH_RETURN = 0
    DEFICIT = 1

class AutoDividendType( Enum ):
    AUTO = 0
    MANUAL = 1
    MANUAL_WITH_AUTO = 2

class TradingData( Enum ):
    TRADING_DATE = 0
    TRADING_TYPE = auto() # 0:賣出, 1:買進, 2:股利, 3:減資
    TRADING_PRICE_TYPE = auto() # 0:每股, 1:總金額
    PER_SHARE_TRADING_PRICE = auto()
    TOTAL_TRADING_PRICE = auto()
    TRADING_QUANTITY = auto()
    TRADING_FEE_DISCOUNT = auto()
    REGULAR_BUY_TRADING_FEE_TYPE = auto() # 0:固定, 1:變動
    REGULAR_BUY_TRADING_FEE_MINIMUM = auto()
    REGULAR_BUY_TRADING_FEE_CONSTANT = auto()
    DIVIDEND_VALUE_TYPE = auto() # 0:每股, 1:總額
    STOCK_DIVIDEND = auto()
    CASH_DIVIDEND = auto()
    CUSTOM_EXTRA_INSURANCE_FEE = auto() #-1表示不使用自訂的補充保費
    CAPITAL_REDUCTION_PER_SHARE = auto()
    CAPITAL_REDUCTION_TYPE = auto()
    USE_AUTO_DIVIDEND_DATA = auto()
    DAYING_TRADING = auto()
    SORTED_INDEX_NON_SAVE = auto() #不會記錄
    TRADING_VALUE_NON_SAVE = auto() #不會記錄
    TRADING_FEE_NON_SAVE = auto() #不會記錄
    TRADING_TAX_NON_SAVE = auto() #不會記錄
    TRADING_COST_NON_SAVE = auto() #不會記錄
    STOCK_DIVIDEND_GAIN_NON_SAVE = auto() #不會記錄
    STOCK_DIVIDEND_POSITION_DATE_NON_SAVE = auto() #不會記錄 股票股利入帳日(不是除權日)
    CASH_DIVIDEND_GAIN_NON_SAVE = auto() #不會記錄
    EXTRA_INSURANCE_FEE_NON_SAVE = auto() #不會記錄
    ACCUMULATED_COST_NON_SAVE = auto() #不會記錄
    ACCUMULATED_COST_WITHOUT_CONSIDERING_DIVIDEND_NON_SAVE = auto() #不會記錄
    ACCUMULATED_QUANTITY_NON_SAVE = auto() #不會記錄
    ACCUMULATED_AVERAGE_COST_NON_SAVE = auto() #不會記錄
    ACCUMULATED_AVERAGE_COST_WITHOUT_CONSIDERING_DIVIDEND_NON_SAVE = auto() #不會記錄
    IS_AUTO_DIVIDEND_DATA_NON_SAVE = auto() #不會記錄
    ACCUMULATED_STOCK_DIVIDEND_GAIN_NON_SAVE = auto() #不會記錄
    ACCUMULATED_CASH_DIVIDEND_GAIN_NON_SAVE = auto() #不會記錄
    IS_REALLY_DAYING_TRADING_NON_SAVE = auto() #不會記錄
    SELLING_PROFIT_NON_SAVE = auto() #不會記錄 本次出售損益
    CURRENT_BUYING_COST_NON_SAVE = auto() #不會記錄 現股總買進成本
    
class TradingCost( Enum ):
    TRADING_VALUE = 0
    TRADING_FEE = 1
    TRADING_TAX = 2
    TRADING_TOTAL_COST = 3

class TransferType( Enum ):
    TRANSFER_IN = 0
    TRANSFER_OUT = 1

class TransferData( Enum ):
    TRANSFER_DATE = 0
    TRANSFER_TYPE = 1 # 0:入金, 1:出金
    TRANSFER_VALUE = 2
    SORTED_INDEX_NON_SAVE = 3 #不會記錄
    TOTAL_VALUE_NON_SAVE = 4 #不會記錄

class StockInfoType( Enum ):
    LATEST_PRICE = 1 #收盤價
    QUANTITY = 2 #庫存股數
    LATEST_MARKET_VALUE = 3 #現值
    CURRENT_COST = 4 #現股成本
    UNREALIZED_PROFIT = 5 #未實現損益
    UNREALIZED_PROFIT_RATIO = 6 #未實現報酬率
    REALIZED_PROFIT = 7 #已實現損益
    BREAK_EVEN_PRICE = 8 #損益平衡價
    ACCUMULATED_COST = 9 #累計成本
    ACCUMULATED_AVERAGE_COST = 10 #累計平均成本
    ACCUMULATED_PROFIT = 11 #累計損益
    ACCUMULATED_TRADING_FEE = 12 #累計手續費
    ACCUMULATED_TAX = 13 #累計交易稅
    ACCUMULATED_DIVIDEND_INCOME = 14 #累計股利所得
    XIRR_VALUE = 15 #平均年化報酬率
    HOLDING_MARKET_RATIO = 16 #持股淨值比

class DecimalRoundType( Enum ):
    ROUND_DOWN = 0 #無條件捨去
    ROUND_OFF = 1 #四捨五入
    ROUND_UP = 2 #無條件進位

class DiscountTimeType( Enum ):
    IMMEDIATE = 0 #即時折扣
    DEFERRED_REFUND = 1 #事後退佣

g_dict_stock_info = {
    StockInfoType.LATEST_PRICE: "收盤價",
    StockInfoType.QUANTITY: "庫存股數",
    StockInfoType.LATEST_MARKET_VALUE: "市值",
    StockInfoType.CURRENT_COST: "現股成本",
    StockInfoType.UNREALIZED_PROFIT: "未實現損益",
    StockInfoType.UNREALIZED_PROFIT_RATIO: "未實現報酬率",
    StockInfoType.REALIZED_PROFIT: "已實現損益",
    StockInfoType.BREAK_EVEN_PRICE: "損益平衡價",
    StockInfoType.ACCUMULATED_COST: "累計成本",
    StockInfoType.ACCUMULATED_AVERAGE_COST: "累計平均成本",
    StockInfoType.ACCUMULATED_PROFIT: "累計損益",
    StockInfoType.ACCUMULATED_TRADING_FEE: "累計手續費",
    StockInfoType.ACCUMULATED_TAX: "累計交易稅",
    StockInfoType.ACCUMULATED_DIVIDEND_INCOME: "累計股利所得",
    StockInfoType.XIRR_VALUE: "平均年化報酬率",
    StockInfoType.HOLDING_MARKET_RATIO: "持股淨值比"
}
#endregion

#region 各式Dialog
class EditTabTitleDialog( QDialog ):
    """一個小對話框用於編輯 Tab 標題"""
    def __init__( self, current_title, parent = None ):
        super().__init__( parent )
        self.setWindowTitle( "修改名稱" )
        self.resize( 300, 100 )

        self.layout = QVBoxLayout( self )

        self.label = QLabel("輸入名稱:")
        self.layout.addWidget( self.label )

        self.line_edit = QLineEdit( self )
        self.line_edit.setText( current_title )
        self.line_edit.selectAll()
        self.layout.addWidget( self.line_edit )

        self.horizontalLayout_4 = QHBoxLayout()
        self.qtOkPushButton = QPushButton( self )
        self.qtOkPushButton.setText( "確認" )
        self.qtCancelPushButton = QPushButton( self )
        self.qtCancelPushButton.setText( "取消" )

        self.horizontalLayout_4.addWidget(self.qtOkPushButton)
        self.horizontalLayout_4.addWidget(self.qtCancelPushButton)

        self.qtOkPushButton.clicked.connect( self.accept )
        self.qtCancelPushButton.clicked.connect( self.reject )

        self.layout.addLayout( self.horizontalLayout_4 )

    def get_new_title(self):
        """返回用戶輸入的標題"""
        return self.line_edit.text()

class ImportDataDuplicateOptionDialog( QDialog ):
    def __init__( self, parent = None ):
        super().__init__( parent )

        self.ui = Ui_DuplicateOptionDialog()
        self.ui.setupUi( self )
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )
        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.b_overwrite = False

    def accept_data( self ):
        self.b_overwrite = self.ui.qtOverWriteRadioButton.isChecked()
        self.accept()
    
    def cancel( self ):
        self.reject()

class StockDividendTransferFeeEditSpinboxDialog( QDialog ):
    def __init__( self, n_transfer_fee, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockDividendTransferFeeEditSpinboxDialog()
        self.ui.setupUi( self )
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )
        self.ui.qtDividendTransferFeeSpinBox.setValue( n_transfer_fee )
        self.n_new_transfer_fee = n_transfer_fee

        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        
    def accept_data( self ):
        self.n_new_transfer_fee = self.ui.qtDividendTransferFeeSpinBox.value()
        self.accept()

    def cancel( self ):
        self.reject()

class StockTradingEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, e_decimal_round_type : DecimalRoundType, b_etf, b_discount, f_discount_value, n_minimum_common_trading_fee, n_minimum_odd_trading_fee, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockTradingDialog()
        self.ui.setupUi( self )
        
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtDateEdit.dateChanged.connect( lambda: Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel ) )
        self.ui.qtDiscountCheckBox.setChecked( b_discount )
        self.ui.qtDiscountRateDoubleSpinBox.setValue( f_discount_value * 10 )
        self.ui.qtDiscountRateDoubleSpinBox.setEnabled( b_discount )

        self.ui.qtDiscountCheckBox.stateChanged.connect( self.on_discount_check_box_state_changed )
        self.ui.qtDiscountRateDoubleSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtBuyRadioButton.toggled.connect( self.compute_cost )
        self.ui.qtSellRadioButton.toggled.connect( self.compute_cost )
        self.ui.qtCommonTradeRadioButton.toggled.connect( self.on_trading_type_changed )
        self.ui.qtOddTradeRadioButton.toggled.connect( self.on_trading_type_changed )
        self.ui.qtPriceDoubleSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtCommonTradeCountSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtOddTradeCountSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.b_etf = b_etf
        self.n_minimum_common_trading_fee = n_minimum_common_trading_fee
        self.n_minimum_odd_trading_fee = n_minimum_odd_trading_fee
        self.str_stock_name = str_stock_name
        self.e_decimal_round_type = e_decimal_round_type
        # self.load_stylesheet("style.css")
        self.dict_trading_data = {}
        self.compute_cost()
        Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel )

    def load_stylesheet( self, file_path ):
        try:
            with open(file_path, "r", encoding="utf-8") as file:  # 指定 UTF-8 編碼
                stylesheet = file.read()
                self.setStyleSheet(stylesheet)
        except FileNotFoundError:
            print(f"CSS 檔案 {file_path} 找不到")
        except Exception as e:
            print(f"讀取 CSS 檔案時發生錯誤: {e}")

    def on_trading_type_changed( self ):
        if self.ui.qtCommonTradeRadioButton.isChecked():
            self.ui.qtCommonTradeCountSpinBox.setEnabled( True )
            self.ui.qtOddTradeCountSpinBox.setEnabled( False )
        else:
            self.ui.qtCommonTradeCountSpinBox.setEnabled( False )
            self.ui.qtOddTradeCountSpinBox.setEnabled( True )

        self.compute_cost()

    def on_discount_check_box_state_changed( self, state ):
        if state == 2:
            self.ui.qtDiscountRateDoubleSpinBox.setEnabled( True )
        else:
            self.ui.qtDiscountRateDoubleSpinBox.setEnabled( False )

        self.compute_cost()

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_trading_type( self, e_trading_type ):
        with ( QSignalBlocker( self.ui.qtBuyRadioButton ),
               QSignalBlocker( self.ui.qtSellRadioButton ) ):
            if e_trading_type == TradingType.BUY:
                self.ui.qtBuyRadioButton.setChecked( True )
            else:
                self.ui.qtSellRadioButton.setChecked( True )

    def setup_trading_count( self, f_count ):
        with ( QSignalBlocker( self.ui.qtCommonTradeRadioButton ),
               QSignalBlocker( self.ui.qtOddTradeRadioButton ),
               QSignalBlocker( self.ui.qtCommonTradeCountSpinBox ),
               QSignalBlocker( self.ui.qtOddTradeCountSpinBox ) ):
            if f_count % 1000 == 0:
                self.ui.qtCommonTradeRadioButton.setChecked( True )
                self.ui.qtCommonTradeCountSpinBox.setValue( f_count / 1000 )
                self.ui.qtCommonTradeCountSpinBox.setEnabled( True )
                self.ui.qtOddTradeCountSpinBox.setValue( 0 )
                self.ui.qtOddTradeCountSpinBox.setEnabled( False )
            else:
                self.ui.qtOddTradeRadioButton.setChecked( True )
                self.ui.qtCommonTradeCountSpinBox.setValue( 0 )
                self.ui.qtCommonTradeCountSpinBox.setEnabled( False )
                self.ui.qtOddTradeCountSpinBox.setValue( f_count )
                self.ui.qtOddTradeCountSpinBox.setEnabled( True )

    def setup_trading_price( self, f_price ):
        with ( QSignalBlocker( self.ui.qtPriceDoubleSpinBox ) ):
            self.ui.qtPriceDoubleSpinBox.setValue( f_price )

    def setup_trading_discount( self, f_discount_value ):
        with ( QSignalBlocker( self.ui.qtDiscountCheckBox ),
               QSignalBlocker( self.ui.qtDiscountRateDoubleSpinBox ) ):
            if f_discount_value != 1:
                self.ui.qtDiscountCheckBox.setChecked( True )
                self.ui.qtDiscountRateDoubleSpinBox.setValue( f_discount_value * 10 )
                self.ui.qtDiscountRateDoubleSpinBox.setEnabled( True )
            else:
                self.ui.qtDiscountCheckBox.setChecked( False )
                self.ui.qtDiscountRateDoubleSpinBox.setValue( 6 )
                self.ui.qtDiscountRateDoubleSpinBox.setEnabled( False )
    
    def setup_daying_trading( self, b_daying_trading ):
        with ( QSignalBlocker( self.ui.qtDayingTradingCheckBox ) ):
            if b_daying_trading:
                self.ui.qtDayingTradingCheckBox.setChecked( True )
            else:
                self.ui.qtDayingTradingCheckBox.setChecked( False )

    def get_trading_type( self ):
        if self.ui.qtBuyRadioButton.isChecked():
            return TradingType.BUY
        else:
            return TradingType.SELL

    def get_trading_count( self ):
        if self.ui.qtCommonTradeRadioButton.isChecked():
            return self.ui.qtCommonTradeCountSpinBox.value() * 1000
        else:
            return self.ui.qtOddTradeCountSpinBox.value()
        
    def get_trading_fee_discount( self ):
        if self.ui.qtDiscountCheckBox.isChecked():
            return self.ui.qtDiscountRateDoubleSpinBox.value() / 10
        else:
            return 1

    def compute_cost( self ):
        e_trading_type = self.get_trading_type()
        f_trading_price = self.ui.qtPriceDoubleSpinBox.value()
        n_trading_count = self.get_trading_count()
        f_trading_fee_discount = self.get_trading_fee_discount() 
        b_bond = True if '債' in self.str_stock_name else False
        dict_result = Utility.compute_cost( e_trading_type, self.e_decimal_round_type, f_trading_price, n_trading_count, f_trading_fee_discount, self.n_minimum_common_trading_fee, self.n_minimum_odd_trading_fee, self.b_etf, False, b_bond )

        if e_trading_type == TradingType.BUY:
            self.ui.qtTradingValueLineEdit.setText( format( dict_result[ TradingCost.TRADING_VALUE ], ',' ) )
            self.ui.qtTotalCostLineEdit.setText( format( dict_result[ TradingCost.TRADING_TOTAL_COST ], ',' ) )
        elif e_trading_type == TradingType.SELL:
            self.ui.qtTradingValueLineEdit.setText( format( -dict_result[ TradingCost.TRADING_VALUE ], ',' ) )
            self.ui.qtTotalCostLineEdit.setText( format( -dict_result[ TradingCost.TRADING_TOTAL_COST ], ',' ) )
        self.ui.qtFeeLineEdit.setText( format( dict_result[ TradingCost.TRADING_FEE ], ',' ) )
        self.ui.qtTaxLineEdit.setText( format( dict_result[ TradingCost.TRADING_TAX ], ',' ) )

    def accept_data( self ):
        if float( self.ui.qtTotalCostLineEdit.text().replace( ',', '' ) ) != 0:
            
            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    self.get_trading_type(),                            #交易種類
                                                                    TradingPriceType.PER_SHARE,                         #交易價格種類
                                                                    self.ui.qtPriceDoubleSpinBox.value(),               #每股交易價格
                                                                    0,                                                  #總交易價格
                                                                    self.get_trading_count(),                           #交易股數
                                                                    TradingFeeType.VARIABLE,                            #手續費種類
                                                                    self.get_trading_fee_discount(),                    #手續費折扣
                                                                    0,                                                  #手續費最低金額
                                                                    0,                                                  #手續費固定金額
                                                                    DividendValueType.PER_SHARE,                        #股利金額種類
                                                                    0,                                                  #股票股利
                                                                    0,                                                  #現金股利
                                                                    -1,                                                 #自訂的補充保費
                                                                    0,                                                  #每股減資金額
                                                                    CapitalReductionType.CASH_RETURN,                   #減資種類
                                                                    self.ui.qtDayingTradingCheckBox.isChecked() )       #是否為當沖交易                                          
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

class StockRegularTradingEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, e_decimal_round_type : DecimalRoundType, e_trading_price_type : TradingPriceType, e_trading_fee_type : TradingFeeType, b_discount, f_discount_value, n_trading_fee_minimum, n_trading_fee_constant, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockRegularTradingDialog()
        self.ui.setupUi( self )
        
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )
        
        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        self.e_decimal_round_type = e_decimal_round_type
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtDateEdit.dateChanged.connect( lambda: Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel ) )

        self.setup_trading_price_type( e_trading_price_type )
        self.setup_trading_fee_type( e_trading_fee_type )
        self.ui.qtDiscountRateDoubleSpinBox.setEnabled( b_discount )
        self.ui.qtDiscountCheckBox.setChecked( b_discount )
        self.ui.qtDiscountRateDoubleSpinBox.setValue( f_discount_value * 10 )

        self.setup_trading_fee_minimum( n_trading_fee_minimum )
        self.setup_trading_fee_constant( n_trading_fee_constant )

        button_group_1 = QButtonGroup( self )
        button_group_1.addButton( self.ui.qtVariableFeeRadioButton )
        button_group_1.addButton( self.ui.qtConstantFeeRadioButton )

        self.ui.qtOddTradeCountSpinBox.valueChanged.connect( self.compute_cost )

        self.ui.qtPerSharePriceRadioButton.toggled.connect( self.update_ui_and_compute_cost )
        self.ui.qtTotalPriceRadioButton.toggled.connect( self.update_ui_and_compute_cost )
        self.ui.qtPerSharePriceDoubleSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtTotalPriceSpinBox.valueChanged.connect( self.compute_cost )

        self.ui.qtVariableFeeRadioButton.toggled.connect( self.update_ui_and_compute_cost )
        self.ui.qtConstantFeeRadioButton.toggled.connect( self.update_ui_and_compute_cost )
        self.ui.qtDiscountCheckBox.stateChanged.connect( self.update_ui_and_compute_cost )
        self.ui.qtDiscountRateDoubleSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtTradingFeeMinimumSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtTradingFeeConstantSpinBox.valueChanged.connect( self.compute_cost )

        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        # self.load_stylesheet("style.css")
        self.dict_trading_data = {}
        self.update_ui_and_compute_cost()
        Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel )

    def load_stylesheet( self, file_path ):
        try:
            with open(file_path, "r", encoding="utf-8") as file:  # 指定 UTF-8 編碼
                stylesheet = file.read()
                self.setStyleSheet(stylesheet)
        except FileNotFoundError:
            print(f"CSS 檔案 {file_path} 找不到")
        except Exception as e:
            print(f"讀取 CSS 檔案時發生錯誤: {e}")

    def update_ui_and_compute_cost( self ):
        self.update_ui()
        self.compute_cost()

    def update_ui( self ):
        with ( QSignalBlocker( self.ui.qtPerSharePriceRadioButton ),
               QSignalBlocker( self.ui.qtTotalPriceRadioButton ),
               QSignalBlocker( self.ui.qtPerSharePriceDoubleSpinBox ),
               QSignalBlocker( self.ui.qtTotalPriceSpinBox ),
               QSignalBlocker( self.ui.qtVariableFeeRadioButton ),
               QSignalBlocker( self.ui.qtConstantFeeRadioButton ),
               QSignalBlocker( self.ui.qtDiscountCheckBox ),
               QSignalBlocker( self.ui.qtDiscountRateDoubleSpinBox ),
               QSignalBlocker( self.ui.qtTradingFeeMinimumSpinBox ),
               QSignalBlocker( self.ui.qtTradingFeeConstantSpinBox ) ):

            if self.ui.qtPerSharePriceRadioButton.isChecked():
                self.ui.qtPerSharePriceDoubleSpinBox.setEnabled( True )
                self.ui.qtTotalPriceSpinBox.setEnabled( False )
            else:
                self.ui.qtPerSharePriceDoubleSpinBox.setEnabled( False )
                self.ui.qtTotalPriceSpinBox.setEnabled( True )

            if self.ui.qtVariableFeeRadioButton.isChecked():
                self.ui.qtDiscountCheckBox.setEnabled( True )

                if self.ui.qtDiscountCheckBox.isChecked():
                    self.ui.qtDiscountRateDoubleSpinBox.setEnabled( True )
                    self.ui.label_10.setEnabled( True )
                else:
                    self.ui.qtDiscountRateDoubleSpinBox.setEnabled( False )
                    self.ui.label_10.setEnabled( False )

                self.ui.label_8.setEnabled( True )
                self.ui.qtTradingFeeMinimumSpinBox.setEnabled( True )
                self.ui.label_11.setEnabled( True )

                self.ui.qtTradingFeeConstantSpinBox.setEnabled( False )
                self.ui.label_4.setEnabled( False )
            else:
                self.ui.qtDiscountCheckBox.setEnabled( False )
                self.ui.qtDiscountRateDoubleSpinBox.setEnabled( False )
                self.ui.label_10.setEnabled( False )
                self.ui.label_8.setEnabled( False )
                self.ui.qtTradingFeeMinimumSpinBox.setEnabled( False )
                self.ui.label_11.setEnabled( False )
                self.ui.qtTradingFeeConstantSpinBox.setEnabled( True )
                self.ui.label_4.setEnabled( True )

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_trading_count( self, n_count ):
        with ( QSignalBlocker( self.ui.qtOddTradeCountSpinBox ) ):
            self.ui.qtOddTradeCountSpinBox.setValue( n_count )

    def setup_trading_price_type( self, e_trading_price_type: TradingPriceType ):
        with ( QSignalBlocker( self.ui.qtPerSharePriceRadioButton ),
               QSignalBlocker( self.ui.qtTotalPriceRadioButton ) ):
            if e_trading_price_type == TradingPriceType.PER_SHARE:
                self.ui.qtPerSharePriceRadioButton.setChecked( True )
            else:
                self.ui.qtTotalPriceRadioButton.setChecked( True )

            self.update_ui()

    def setup_per_share_trading_price( self, f_price ):
        with ( QSignalBlocker( self.ui.qtPerSharePriceDoubleSpinBox ) ):
            self.ui.qtPerSharePriceDoubleSpinBox.setValue( f_price )

    def setup_total_trading_price( self, n_total_price ):
        with ( QSignalBlocker( self.ui.qtTotalPriceSpinBox ) ):
            self.ui.qtTotalPriceSpinBox.setValue( n_total_price )

    def setup_trading_fee_type( self, e_trading_fee_type : TradingFeeType ):
        with ( QSignalBlocker( self.ui.qtVariableFeeRadioButton ),
               QSignalBlocker( self.ui.qtConstantFeeRadioButton ) ):
            if e_trading_fee_type == TradingFeeType.VARIABLE:
                self.ui.qtVariableFeeRadioButton.setChecked( True )
            else:
                self.ui.qtConstantFeeRadioButton.setChecked( True )

            self.update_ui()

    def setup_trading_discount( self, f_discount_value ):
        with ( QSignalBlocker( self.ui.qtDiscountCheckBox ),
               QSignalBlocker( self.ui.qtDiscountRateDoubleSpinBox ) ):
            if f_discount_value != 1:
                self.ui.qtDiscountCheckBox.setChecked( True )
                self.ui.qtDiscountRateDoubleSpinBox.setValue( f_discount_value * 10 )
                self.ui.qtDiscountRateDoubleSpinBox.setEnabled( True )
            else:
                self.ui.qtDiscountCheckBox.setChecked( False )
                self.ui.qtDiscountRateDoubleSpinBox.setValue( 6 )
                self.ui.qtDiscountRateDoubleSpinBox.setEnabled( False )

            self.update_ui()

    def setup_trading_fee_minimum( self, n_trading_fee_minimum ):
        with ( QSignalBlocker( self.ui.qtTradingFeeMinimumSpinBox ) ):
            self.ui.qtTradingFeeMinimumSpinBox.setValue( n_trading_fee_minimum )

    def setup_trading_fee_constant( self, n_trading_fee_constant ):
        with ( QSignalBlocker( self.ui.qtTradingFeeConstantSpinBox ) ):
            self.ui.qtTradingFeeConstantSpinBox.setValue( n_trading_fee_constant )

    def get_trading_price_type( self ):
        if self.ui.qtPerSharePriceRadioButton.isChecked():
            return TradingPriceType.PER_SHARE
        else:
            return TradingPriceType.TOTAL

    def get_trading_fee_type( self ):
        if self.ui.qtVariableFeeRadioButton.isChecked():
            return TradingFeeType.VARIABLE
        else:
            return TradingFeeType.CONSTANT
        
    def get_trading_fee_discount( self ):
        if self.ui.qtDiscountCheckBox.isChecked():
            return self.ui.qtDiscountRateDoubleSpinBox.value() / 10
        else:
            return 1
        
    def compute_cost( self ):
        if self.ui.qtPerSharePriceRadioButton.isChecked():
            n_trading_count = int( self.ui.qtOddTradeCountSpinBox.value() ) 
            f_trading_price = Decimal( str( self.ui.qtPerSharePriceDoubleSpinBox.value() ) )
            n_trading_value = int( f_trading_price * n_trading_count )
        else:
            n_trading_value = int( self.ui.qtTotalPriceSpinBox.value() )
        
        e_trading_fee_type = self.get_trading_fee_type()
        
        if n_trading_value == 0:
            n_trading_fee = 0
        elif e_trading_fee_type == TradingFeeType.VARIABLE:
            f_trading_fee_discount = Decimal( str( self.get_trading_fee_discount() ) )
            n_trading_fee_minimum = int( self.ui.qtTradingFeeMinimumSpinBox.value() ) 
            n_trading_fee =  Utility.round_to( n_trading_value * Decimal( '0.001425' ) * f_trading_fee_discount, self.e_decimal_round_type )
            n_trading_fee = max( n_trading_fee_minimum, n_trading_fee )
        else:
            n_trading_fee = int( self.ui.qtTradingFeeConstantSpinBox.value() )

        n_trading_total_cost = n_trading_value + n_trading_fee

        self.ui.qtTradingValueLineEdit.setText( format( n_trading_value, ',' ) )
        self.ui.qtFeeLineEdit.setText( format( n_trading_fee, ',' ) )
        self.ui.qtTotalCostLineEdit.setText( format( n_trading_total_cost, ',' ) )

    def accept_data( self ):
        if float( self.ui.qtTotalCostLineEdit.text().replace( ',', '' ) ) != 0:
            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    TradingType.REGULAR_BUY,                            #交易種類
                                                                    self.get_trading_price_type(),                      #交易價格種類
                                                                    self.ui.qtPerSharePriceDoubleSpinBox.value(),       #每股交易價格
                                                                    self.ui.qtTotalPriceSpinBox.value(),                #總交易價格
                                                                    self.ui.qtOddTradeCountSpinBox.value(),             #交易股數
                                                                    self.get_trading_fee_type(),                        #手續費種類
                                                                    self.get_trading_fee_discount(),                    #手續費折扣
                                                                    self.ui.qtTradingFeeMinimumSpinBox.value(),         #手續費最低金額
                                                                    self.ui.qtTradingFeeConstantSpinBox.value(),        #手續費固定金額
                                                                    DividendValueType.PER_SHARE,                        #股利金額種類
                                                                    0,                                                  #股票股利
                                                                    0,                                                  #現金股利
                                                                    -1,                                                 #自訂的補充保費
                                                                    0,                                                  #每股減資金額
                                                                    CapitalReductionType.CASH_RETURN,                   #減資種類
                                                                    False )                                             #是否為當沖交易
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

class StockDividendEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockDividendDialog()
        self.ui.setupUi( self )
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtDateEdit.dateChanged.connect( lambda: Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel ) )

        self.ui.qtPerShareDividendRadioButton.toggled.connect( self.update_ui )
        self.ui.qtTotalDividendRadioButton.toggled.connect( self.update_ui )
        self.ui.qtCustomExtraInsuranceCheckBox.toggled.connect( self.update_ui )

        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.dict_trading_data = {}

        self.update_ui()
        Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel )

    def update_ui( self ):
        if self.ui.qtPerShareDividendRadioButton.isChecked():
            self.ui.qtPerShareStockDividendDoubleSpinBox.setEnabled( True )
            self.ui.qtPerShareCashDividendDoubleSpinBox.setEnabled( True )
            self.ui.qtTotalStockDividendDoubleSpinBox.setEnabled( False )
            self.ui.qtTotalCashDividendDoubleSpinBox.setEnabled( False )
        else:
            self.ui.qtPerShareStockDividendDoubleSpinBox.setEnabled( False )
            self.ui.qtPerShareCashDividendDoubleSpinBox.setEnabled( False )
            self.ui.qtTotalStockDividendDoubleSpinBox.setEnabled( True )
            self.ui.qtTotalCashDividendDoubleSpinBox.setEnabled( True )

        if self.ui.qtCustomExtraInsuranceCheckBox.isChecked():
            self.ui.qtExtraInsuranceSpinBox.setEnabled( True )
        else:
            self.ui.qtExtraInsuranceSpinBox.setEnabled( False )

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_dividend_value_type( self, e_dividend_value_type ):
        if e_dividend_value_type == DividendValueType.PER_SHARE:
            self.ui.qtPerShareDividendRadioButton.setChecked( True )
        else:
            self.ui.qtTotalDividendRadioButton.setChecked( True )

        self.update_ui()

    def setup_per_share_stock_dividend( self, f_per_share_stock_dividend ):
        self.ui.qtPerShareStockDividendDoubleSpinBox.setValue( f_per_share_stock_dividend )

    def setup_per_share_cash_dividend( self, f_per_share_cash_dividend ):
        self.ui.qtPerShareCashDividendDoubleSpinBox.setValue( f_per_share_cash_dividend )

    def setup_total_stock_dividend( self, n_total_stock_dividend ):
        self.ui.qtTotalStockDividendDoubleSpinBox.setValue( n_total_stock_dividend )

    def setup_total_cash_dividend( self, n_total_cash_dividend ):
        self.ui.qtTotalCashDividendDoubleSpinBox.setValue( n_total_cash_dividend )

    def setup_custom_extra_insurance_fee( self, n_custom_extra_insurance_fee ):
        if n_custom_extra_insurance_fee == -1:
            self.ui.qtCustomExtraInsuranceCheckBox.setChecked( False )
            self.ui.qtExtraInsuranceSpinBox.setValue( 0 )
        else:
            self.ui.qtCustomExtraInsuranceCheckBox.setChecked( True )
            self.ui.qtExtraInsuranceSpinBox.setValue( n_custom_extra_insurance_fee )

        self.update_ui()

    def accept_data( self ):
        if self.ui.qtPerShareDividendRadioButton.isChecked():
            e_dividend_value_type = DividendValueType.PER_SHARE
            f_stock_dividend = self.ui.qtPerShareStockDividendDoubleSpinBox.value()
            f_cash_dividend = self.ui.qtPerShareCashDividendDoubleSpinBox.value()
        else:
            e_dividend_value_type = DividendValueType.TOTAL
            f_stock_dividend = self.ui.qtTotalStockDividendDoubleSpinBox.value()
            f_cash_dividend = self.ui.qtTotalCashDividendDoubleSpinBox.value()

        if self.ui.qtCustomExtraInsuranceCheckBox.isChecked():
            n_custom_extra_insurance_fee = self.ui.qtExtraInsuranceSpinBox.value()
        else:
            n_custom_extra_insurance_fee = -1

        if f_stock_dividend != 0 or f_cash_dividend != 0:
            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    TradingType.DIVIDEND,                               #交易種類
                                                                    TradingPriceType.PER_SHARE,                         #交易價格種類
                                                                    0,                                                  #每股交易價格
                                                                    0,                                                  #總交易價格                         
                                                                    0,                                                  #交易股數
                                                                    TradingFeeType.VARIABLE,                            #手續費種類
                                                                    1,                                                  #手續費折扣
                                                                    0,                                                  #手續費最低金額
                                                                    0,                                                  #手續費固定金額
                                                                    e_dividend_value_type,                              #股利金額種類
                                                                    f_stock_dividend,                                   #股票股利
                                                                    f_cash_dividend,                                    #現金股利
                                                                    n_custom_extra_insurance_fee,                      #自訂的補充保費
                                                                    0,                                                  #每股減資金額
                                                                    CapitalReductionType.CASH_RETURN,                   #減資種類
                                                                    False )                                             #是否為當沖交易
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

class StockCapitalIncreaseEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockCapitalIncreaseDialog()
        self.ui.setupUi( self )
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtDateEdit.dateChanged.connect( lambda: Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel ) )

        self.ui.qtOddTradeCountSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtPerSharePriceRadioButton.toggled.connect( self.update_ui_and_compute_cost )
        self.ui.qtTotalPriceRadioButton.toggled.connect( self.update_ui_and_compute_cost )
        
        self.ui.qtPerSharePriceDoubleSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtTotalPriceSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.dict_trading_data = {}

        self.update_ui_and_compute_cost()
        Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel )

    def update_ui_and_compute_cost( self ):
        self.update_ui()
        self.compute_cost()
        
    def update_ui( self ):
        with ( QSignalBlocker( self.ui.qtPerSharePriceRadioButton ),
               QSignalBlocker( self.ui.qtTotalPriceRadioButton ),
               QSignalBlocker( self.ui.qtPerSharePriceDoubleSpinBox ),
               QSignalBlocker( self.ui.qtTotalPriceSpinBox ) ):

            if self.ui.qtPerSharePriceRadioButton.isChecked():
                self.ui.qtPerSharePriceDoubleSpinBox.setEnabled( True )
                self.ui.qtTotalPriceSpinBox.setEnabled( False )
            else:
                self.ui.qtPerSharePriceDoubleSpinBox.setEnabled( False )
                self.ui.qtTotalPriceSpinBox.setEnabled( True )

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_trading_count( self, n_count ):
        with ( QSignalBlocker( self.ui.qtOddTradeCountSpinBox ) ):
            self.ui.qtOddTradeCountSpinBox.setValue( n_count )

    def setup_trading_price_type( self, e_trading_price_type ):
        with ( QSignalBlocker( self.ui.qtPerSharePriceRadioButton ), 
               QSignalBlocker( self.ui.qtTotalPriceRadioButton ) ):
            if e_trading_price_type == TradingPriceType.PER_SHARE:
                self.ui.qtPerSharePriceRadioButton.setChecked( True )
            else:
                self.ui.qtTotalPriceRadioButton.setChecked( True )

    def setup_per_share_trading_price( self, f_price ):
        with ( QSignalBlocker( self.ui.qtPerSharePriceDoubleSpinBox ) ):
            self.ui.qtPerSharePriceDoubleSpinBox.setValue( f_price )

    def setup_total_trading_price( self, n_total_price ):
        with ( QSignalBlocker( self.ui.qtTotalPriceSpinBox ) ):
            self.ui.qtTotalPriceSpinBox.setValue( n_total_price )

    def get_trading_price_type( self ):
        if self.ui.qtPerSharePriceRadioButton.isChecked():
            return TradingPriceType.PER_SHARE
        else:
            return TradingPriceType.TOTAL

    def compute_cost( self ):
        if self.ui.qtPerSharePriceRadioButton.isChecked():
            n_trading_count = int( self.ui.qtOddTradeCountSpinBox.value() ) 
            f_trading_price = Decimal( str( self.ui.qtPerSharePriceDoubleSpinBox.value() ) )
            self.ui.qtTotalCostLineEdit.setText( format( int( f_trading_price * n_trading_count ), ',' ) )
        else:
            n_total_trading_price = self.ui.qtTotalPriceSpinBox.value()
            self.ui.qtTotalCostLineEdit.setText( format( int( n_total_trading_price ), ',' ) )

    def accept_data( self ):
        if float( self.ui.qtTotalCostLineEdit.text().replace( ',', '' ) ) != 0:
            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    TradingType.CAPITAL_INCREASE,                       #交易種類
                                                                    self.get_trading_price_type(),                      #交易價格種類
                                                                    self.ui.qtPerSharePriceDoubleSpinBox.value(),       #每股交易價格
                                                                    self.ui.qtTotalPriceSpinBox.value(),                #總交易價格
                                                                    self.ui.qtOddTradeCountSpinBox.value(),             #交易股數
                                                                    TradingFeeType.VARIABLE,                            #手續費種類
                                                                    1,                                                  #手續費折扣
                                                                    0,                                                  #手續費最低金額
                                                                    0,                                                  #手續費固定金額
                                                                    DividendValueType.PER_SHARE,                        #股利金額種類
                                                                    0,                                                  #股票股利
                                                                    0,                                                  #現金股利
                                                                    -1,                                                 #自訂的補充保費
                                                                    0,                                                  #每股減資金額
                                                                    CapitalReductionType.CASH_RETURN,                   #減資種類
                                                                    False )                                             #是否為當沖交易
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

class StockCapitalReductionEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockCapitalReductionDialog()
        self.ui.setupUi( self )
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtDateEdit.dateChanged.connect( lambda: Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel ) )

        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.dict_trading_data = {}
        Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel )

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_stock_capital_reduction_value( self, f_stock_capital_reduction_per_share ):
        self.ui.qtCapitalReductionDoubleSpinBox.setValue( f_stock_capital_reduction_per_share )

    def setup_stock_capital_reduction_type( self, e_capital_reduction_type ):
        if e_capital_reduction_type == CapitalReductionType.CASH_RETURN:
            self.ui.qtCashReturnRadioButton.setChecked( True )
        else:
            self.ui.qtDeficitRadioButton.setChecked( True )

    def get_capital_reduction_type( self ):
        if self.ui.qtCashReturnRadioButton.isChecked():
            return CapitalReductionType.CASH_RETURN
        else:
            return CapitalReductionType.DEFICIT

    def accept_data( self ):
        f_stock_capital_reduction_per_share = self.ui.qtCapitalReductionDoubleSpinBox.value()
        if f_stock_capital_reduction_per_share != 0:

            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    TradingType.CAPITAL_REDUCTION,                      #交易種類
                                                                    TradingPriceType.PER_SHARE,                         #交易價格種類
                                                                    0,                                                  #每股交易價格
                                                                    0,                                                  #總交易價格                         
                                                                    0,                                                  #交易股數
                                                                    TradingFeeType.VARIABLE,                            #手續費種類
                                                                    1,                                                  #手續費折扣
                                                                    0,                                                  #手續費最低金額
                                                                    0,                                                  #手續費固定金額
                                                                    DividendValueType.PER_SHARE,                        #股利金額種類
                                                                    0,                                                  #股票股利
                                                                    0,                                                  #現金股利
                                                                    -1,                                                 #自訂的補充保費
                                                                    f_stock_capital_reduction_per_share,                #每股減資金額
                                                                    self.get_capital_reduction_type(),                  #減資種類
                                                                    False )                                             #是否為當沖交易
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

class StockSplitEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockSplitDialog()
        self.ui.setupUi( self )
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtDateEdit.dateChanged.connect( lambda: Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel ) )

        self.ui.qtSplitRadioButton.toggled.connect( self.update_ui )
        self.ui.qtMergeRadioButton.toggled.connect( self.update_ui )

        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.dict_trading_data = {}
        Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel )
        self.update_ui()

    def update_ui( self ):
        with ( QSignalBlocker( self.ui.qtSplitRadioButton ),
               QSignalBlocker( self.ui.qtMergeRadioButton ),
               QSignalBlocker( self.ui.qtStockSplitSpinBox ),
               QSignalBlocker( self.ui.qtStockMergeSpinBox ) ):

            if self.ui.qtSplitRadioButton.isChecked():
                self.ui.qtStockSplitSpinBox.setEnabled( True )
                self.ui.qtStockMergeSpinBox.setEnabled( False )
            else:
                self.ui.qtStockSplitSpinBox.setEnabled( False )
                self.ui.qtStockMergeSpinBox.setEnabled( True )

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_trading_type_and_value( self, e_trading_type, n_stock_split_merge_share ):
        with ( QSignalBlocker( self.ui.qtSplitRadioButton ),
               QSignalBlocker( self.ui.qtMergeRadioButton ) ):
            if e_trading_type == TradingType.SPLIT:
                self.ui.qtSplitRadioButton.setChecked( True )
                self.ui.qtStockSplitSpinBox.setValue( n_stock_split_merge_share )
            else:
                self.ui.qtMergeRadioButton.setChecked( True )
                self.ui.qtStockMergeSpinBox.setValue( n_stock_split_merge_share )

            self.update_ui()

    def accept_data( self ):
        if self.ui.qtSplitRadioButton.isChecked():
            n_stock_split_merge_share = self.ui.qtStockSplitSpinBox.value()
            e_trading_type = TradingType.SPLIT
        else:
            n_stock_split_merge_share = self.ui.qtStockMergeSpinBox.value()
            e_trading_type = TradingType.MERGE
        if n_stock_split_merge_share != 0:
            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    e_trading_type,                                     #交易種類
                                                                    TradingPriceType.PER_SHARE,                         #交易價格種類
                                                                    0,                                                  #每股交易價格
                                                                    0,                                                  #總交易價格                         
                                                                    0,                                                  #交易股數
                                                                    TradingFeeType.VARIABLE,                            #手續費種類
                                                                    1,                                                  #手續費折扣
                                                                    0,                                                  #手續費最低金額
                                                                    0,                                                  #手續費固定金額
                                                                    DividendValueType.PER_SHARE,                        #股利金額種類
                                                                    0,                                                  #股票股利
                                                                    0,                                                  #現金股利
                                                                    -1,                                                 #自訂的補充保費
                                                                    n_stock_split_merge_share,                          #每股減資金額
                                                                    CapitalReductionType.CASH_RETURN,                   #減資種類
                                                                    False )                                             #是否為當沖交易
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

class StockDividendPositionDateEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, obj_position_date, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockDividendPositionDateEditDialog()
        self.ui.setupUi( self )
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        self.ui.qtDateEdit.setDate( obj_position_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtDateEdit.dateChanged.connect( lambda: Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel ) )

        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel )

    def setup_position_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def accept_data( self ):
        self.str_position_date = self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" )
        self.accept()
    
    def cancel( self ):
        self.reject()

class CashTransferEditDialog( QDialog ):
    def __init__( self, str_account_name, parent = None ):
        super().__init__( parent )

        self.ui = Ui_CashTransferDialog()
        self.ui.setupUi( self )
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )

        self.ui.qtAccountNameLabel.setText( str_account_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtDateEdit.dateChanged.connect( lambda: Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel ) )

        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.dict_cash_transfer_data = {}
        Utility.update_weekly_text_by_date( self.ui.qtDateEdit, self.ui.qtWeekdayLabel )

    def setup_transfer_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_transfer_type( self, e_transfer_type ):
        if e_transfer_type == TransferType.TRANSFER_IN:
            self.ui.qtTransferInRadioButton.setChecked( True )
        else:
            self.ui.qtTransferOutRadioButton.setChecked( True )

    def setup_transfer_value( self, n_transfer_value ):
        self.ui.qtCashDividendSpinBox.setValue( n_transfer_value )

    def accept_data( self ):
        n_transfer_value = self.ui.qtCashDividendSpinBox.value()
        if n_transfer_value != 0:
            self.dict_cash_transfer_data[ TransferData.TRANSFER_DATE ] = self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" )
            self.dict_cash_transfer_data[ TransferData.TRANSFER_TYPE ] = TransferType.TRANSFER_IN if self.ui.qtTransferInRadioButton.isChecked() else TransferType.TRANSFER_OUT
            self.dict_cash_transfer_data[ TransferData.TRANSFER_VALUE ] = n_transfer_value
    
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

class AboutDialog( QDialog ):
    def __init__( self, parent = None ):
        super().__init__( parent )

        self.ui = Ui_AboutDialog()
        self.ui.setupUi( self )
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )
        self.ui.qtVersionLabel.setText( "v2.3.0" )

class AccountSettingEditDialog( QDialog ):
    def __init__( self, 
                  str_current_title : str,
                  b_extra_insurance_fee : bool, 
                  n_minimum_common_trading_fee : int, 
                  n_minimum_odd_trading_fee : int, 
                  e_decimal_round_type : DecimalRoundType, 
                  e_discount_time_type : DiscountTimeType, 
                  dict_all_company_number_to_name_and_type : dict, 
                  dict_company_number_to_transfer_fee : dict, 
                  parent = None ):
        super().__init__( parent )

        self.ui = Ui_AccountSettingEditDialog()
        self.ui.setupUi( self )
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )
        self.ui.qtDiscountTimeGroupBox.setVisible( False )
        self.ui.qtGroupNameLabel.setText( str_current_title )
        self.ui.qtStockSelectComboBox.setVisible( False )
        self.dict_all_company_number_to_name_and_type = dict_all_company_number_to_name_and_type
        self.dict_company_number_to_transfer_fee = dict_company_number_to_transfer_fee.copy()

        self.b_extra_insurance_fee = b_extra_insurance_fee
        self.n_minimum_common_trading_fee = n_minimum_common_trading_fee
        self.n_minimum_odd_trading_fee = n_minimum_odd_trading_fee
        self.e_decimal_round_type = e_decimal_round_type
        self.e_discount_time_type = e_discount_time_type
        self.dict_company_number_to_transfer_fee = dict_company_number_to_transfer_fee
        self.ui.qtExtraInsuranceCheckBox.setChecked( b_extra_insurance_fee )
        self.ui.qtMinCommonTradingFeeSpinBox.setValue( n_minimum_common_trading_fee )
        self.ui.qtMinOddTradingFeeSpinBox.setValue( n_minimum_odd_trading_fee )
        if e_decimal_round_type == DecimalRoundType.ROUND_DOWN:
            self.ui.qtRoundDownRadioButton.setChecked( True )
        elif e_decimal_round_type == DecimalRoundType.ROUND_OFF:
            self.ui.qtRoundOffRadioButton.setChecked( True )
        else:
            self.ui.qtRoundUpRadioButton.setChecked( True )

        if e_discount_time_type == DiscountTimeType.IMMEDIATE:
            self.ui.qtDiscountImmediateRadioButton.setChecked( True )
        else:
            self.ui.qtDiscountPostRadioButton.setChecked( True )

        delegate = share_ui.CenterIconDelegate()
        self.dividend_transfer_fee_model = QStandardItemModel( 0, 0 ) 
        self.ui.qtDividendTransferFeeTableView.setModel( self.dividend_transfer_fee_model )
        self.ui.qtDividendTransferFeeTableView.setItemDelegate( delegate )
        self.ui.qtDividendTransferFeeTableView.horizontalHeader().setSectionResizeMode( QHeaderView.Stretch )
        self.ui.qtDividendTransferFeeTableView.verticalHeader().setSectionResizeMode( QHeaderView.Fixed )
        self.ui.qtDividendTransferFeeTableView.clicked.connect( lambda index: self.on_transfer_fee_table_item_clicked( index, self.dividend_transfer_fee_model ) )

        self.ui.qtStockSelectComboBox.activated.connect( self.on_stock_select_combo_box_current_index_changed )
        self.ui.qtAddStockPushButton.clicked.connect( self.on_add_stock_push_button_clicked )
        self.ui.qtStockInputLineEdit.textChanged.connect( self.on_stock_input_text_changed ) 
        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.refresh_table_view()

    def on_stock_input_text_changed( self ):
        with QSignalBlocker( self.ui.qtStockInputLineEdit ), QSignalBlocker( self.ui.qtStockSelectComboBox ):
            self.ui.qtStockSelectComboBox.clear()
            str_stock_input = self.ui.qtStockInputLineEdit.text()
            str_stock_input = share_api.lowercase_english_uppercase( str_stock_input )
            if len( str_stock_input ) == 0:
                self.ui.qtStockSelectComboBox.setVisible( False )
                return
            self.ui.qtStockSelectComboBox.setVisible( True )

            for stock_number, list_stock_name_and_type in self.dict_all_company_number_to_name_and_type.items():
                str_stock_name = list_stock_name_and_type[ 0 ]
                str_stock_name_lowercase = share_api.lowercase_english_uppercase( str_stock_name )
                if str_stock_input in stock_number or str_stock_input in str_stock_name_lowercase:
                    self.ui.qtStockSelectComboBox.addItem( f"{stock_number} {str_stock_name}" )
            # self.ui.qtStockSelectComboBox.showPopup() #showPopup的話，focus會被搶走

            self.ui.qtStockInputLineEdit.setFocus()

    def on_stock_select_combo_box_current_index_changed( self, index ): 
        str_stock_input = self.ui.qtStockSelectComboBox.currentText()
        self.ui.qtStockInputLineEdit.setText( str_stock_input )
        self.ui.qtStockSelectComboBox.setVisible( False )
        self.ui.qtStockInputLineEdit.setFocus()

    def on_add_stock_push_button_clicked( self ):
        str_stock_input = self.ui.qtStockInputLineEdit.text()
        self.ui.qtStockInputLineEdit.clear()
        str_first_four_chars = str_stock_input.split(" ")[0]
        if str_first_four_chars not in self.dict_all_company_number_to_name_and_type:
            b_find = False
            for stock_number, list_stock_name_and_type in self.dict_all_company_number_to_name_and_type.items():
                str_stock_name = list_stock_name_and_type[ 0 ]
                if str_first_four_chars == str_stock_name:
                    str_first_four_chars = stock_number
                    b_find = True
                    break
            if not b_find:
                QMessageBox.warning( self, "警告", "輸入的股票代碼不存在", QMessageBox.Ok )
                return

        if str_first_four_chars not in self.dict_company_number_to_transfer_fee:
            self.dict_company_number_to_transfer_fee[ str_first_four_chars ] = 10

        self.refresh_table_view()

    def refresh_table_view( self ):
        self.dividend_transfer_fee_model.clear()
        self.dividend_transfer_fee_model.setHorizontalHeaderLabels( [ "匯費", "編輯", "刪除" ] )
        list_vertical_labels = []
        for index_row,( key_stock_number, value ) in enumerate( self.dict_company_number_to_transfer_fee.items() ):
            list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ key_stock_number ]
            str_stock_name = list_stock_name_and_type[ 0 ]
            list_vertical_labels.append( f"{key_stock_number} {str_stock_name}" )

            transfer_fee_item = QStandardItem( str( value ) )
            transfer_fee_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
            transfer_fee_item.setFlags( transfer_fee_item.flags() & ~Qt.ItemIsEditable )
            self.dividend_transfer_fee_model.setItem( index_row, 0, transfer_fee_item ) 

            edit_icon_item = QStandardItem("")
            edit_icon_item.setIcon( share_icon.get_icon( share_icon.IconType.EDIT ) )
            edit_icon_item.setFlags( edit_icon_item.flags() & ~Qt.ItemIsEditable )
            self.dividend_transfer_fee_model.setItem( index_row, 1, edit_icon_item )

            delete_icon_item = QStandardItem("")
            delete_icon_item.setIcon( share_icon.get_icon( share_icon.IconType.DELETE ) )
            delete_icon_item.setFlags( delete_icon_item.flags() & ~Qt.ItemIsEditable )
            self.dividend_transfer_fee_model.setItem( index_row, 2, delete_icon_item )
        self.dividend_transfer_fee_model.setVerticalHeaderLabels( list_vertical_labels )
        self.ui.qtDividendTransferFeeTableView.setColumnWidth( 0, 60 )
        self.ui.qtDividendTransferFeeTableView.setColumnWidth( 1, 40 )
        self.ui.qtDividendTransferFeeTableView.setColumnWidth( 2, 40 )

    def on_transfer_fee_table_item_clicked( self, index: QModelIndex, table_model ):
        item = table_model.itemFromIndex( index )
        if item is not None:
            n_column = index.column()  # 獲取列索引
            header_text = table_model.verticalHeaderItem( index.row() ).text()
            str_stock_number = header_text.split(" ")[0]
            n_current_fee = self.dict_company_number_to_transfer_fee[ str_stock_number ] 
            
            if n_column == 1:#匯出按鈕
                dialog = StockDividendTransferFeeEditSpinboxDialog( n_current_fee, self )
                if dialog.exec():
                    self.dict_company_number_to_transfer_fee[ str_stock_number ] = dialog.n_new_transfer_fee
                    self.refresh_table_view()
            elif n_column == 2:#刪除按鈕
                result = share_ui.show_warning_message_box_with_ok_cancel_button( "警告", f"確定要刪掉『{header_text}』的匯費資料嗎?", self )
                if result:
                    value = self.dict_company_number_to_transfer_fee.pop( str_stock_number, None )
                    self.refresh_table_view()
    
    def accept_data( self ):
        self.b_extra_insurance_fee = self.ui.qtExtraInsuranceCheckBox.isChecked()
        self.n_minimum_common_trading_fee = int( self.ui.qtMinCommonTradingFeeSpinBox.value() )
        self.n_minimum_odd_trading_fee = int( self.ui.qtMinOddTradingFeeSpinBox.value() )
        if self.ui.qtRoundDownRadioButton.isChecked():
            self.e_decimal_round_type = DecimalRoundType.ROUND_DOWN
        elif self.ui.qtRoundOffRadioButton.isChecked():
            self.e_decimal_round_type = DecimalRoundType.ROUND_OFF
        else:
            self.e_decimal_round_type = DecimalRoundType.ROUND_UP

        if self.ui.qtDiscountImmediateRadioButton.isChecked():
            self.e_discount_time_type = DiscountTimeType.IMMEDIATE
        else:
            self.e_discount_time_type = DiscountTimeType.DEFERRED_REFUND

        self.accept()

    def cancel( self ):
        self.reject()

class ShowItemEditDialog( QDialog ):
    def __init__( self, list_current_show_stock_info : list, list_default_show_stock_info : list, list_total_stock_info : list, parent = None ):
        super().__init__( parent )

        self.ui = Ui_ShowItemEditDialog()
        self.ui.setupUi( self )
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )

        self.ui.qtHideSelectToolButton.clicked.connect( self.move_selected_to_hidden )
        self.ui.qtShowSelectToolButton.clicked.connect( self.move_selected_to_visible )
        self.ui.qtHideAllToolButton.clicked.connect( self.move_all_to_hidden )
        self.ui.qtShowAllToolButton.clicked.connect( self.move_all_to_visible )

        self.ui.qtShowListWidget.setSelectionMode(QListWidget.SingleSelection)
        self.ui.qtShowListWidget.setDragDropMode(QListWidget.InternalMove)
        self.ui.qtShowListWidget.setDefaultDropAction(Qt.MoveAction)
        self.ui.qtShowListWidget.setDragEnabled(True)
        self.ui.qtShowListWidget.setAcceptDrops(True)
        self.ui.qtShowListWidget.setDropIndicatorShown(True)

        self.ui.qtShowListWidget.itemChanged.connect(self.on_item_changed)
        self.ui.qtHideListWidget.itemChanged.connect(self.on_item_changed)

        for e_type in list_current_show_stock_info:
            item = QListWidgetItem( f"{ g_dict_stock_info[ e_type ] }" )
            item.setFlags( item.flags() | Qt.ItemIsUserCheckable )
            item.setCheckState( Qt.Unchecked )
            item.setData( Qt.UserRole, e_type )
            self.ui.qtShowListWidget.addItem( item )

        for e_type in list_total_stock_info:
            if e_type in list_current_show_stock_info:
                continue
            item = QListWidgetItem( f"{ g_dict_stock_info[ e_type ] }" )
            item.setFlags( item.flags() | Qt.ItemIsUserCheckable )
            item.setCheckState( Qt.Unchecked )
            item.setData( Qt.UserRole, e_type )
            self.ui.qtHideListWidget.addItem( item )

        self.list_total_stock_info = list_total_stock_info
        self.list_default_show_stock_info = list_default_show_stock_info
        self.ui.qtResetToDefaultPushButton.clicked.connect( self.reset_to_default )
        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.list_final_show_stock_info = []

        self.update_ui()

    def move_selected_to_hidden( self ):
        self.move_selected_items( self.ui.qtShowListWidget, self.ui.qtHideListWidget )

    def move_selected_to_visible( self ):
        self.move_selected_items( self.ui.qtHideListWidget, self.ui.qtShowListWidget )

    def move_selected_items( self, source_widget: QListWidget, target_widget: QListWidget ):
        items_to_move = []

        # Step 1：從 source 收集被勾選的項目（順序保留）
        for i in range( source_widget.count() ):
            item = source_widget.item( i )
            if item.checkState() == Qt.Checked:
                items_to_move.append( ( i, item ) )

        # Step 2：根據原始順序移除 → 加入 target
        for index, _ in reversed( items_to_move ):  # 倒著刪除
            source_widget.takeItem( index )

        for _, item in items_to_move:  # 順著加回去
            item.setCheckState( Qt.Unchecked )
            target_widget.addItem( item )

        self.update_ui()

    def move_all_to_hidden( self ):
        self.move_all_items( self.ui.qtShowListWidget, self.ui.qtHideListWidget )

    def move_all_to_visible( self ):
        self.move_all_items( self.ui.qtHideListWidget, self.ui.qtShowListWidget )

    def move_all_items( self, source_widget: QListWidget, target_widget: QListWidget ):
        items_to_move = []

        # Step 1: 依照原順序收集 item
        for i in range( source_widget.count() ):
            item = source_widget.item( i )
            items_to_move.append( ( i, item ) )

        # Step 2: 倒著刪除（不破壞 index）
        for index, _ in reversed( items_to_move ):
            source_widget.takeItem( index )

        # Step 3: 順著加入目標
        for _, item in items_to_move:
            item.setCheckState( Qt.Unchecked )
            target_widget.addItem( item )

        self.update_ui()

    def on_item_changed( self, item: QListWidgetItem ):
        self.update_ui()

    def update_ui( self ):
        b_item_selected_in_show_list = False
        b_item_selected_in_hide_list = False
        for i in range( self.ui.qtShowListWidget.count() ):
            item = self.ui.qtShowListWidget.item(i)
            if item.checkState() == Qt.Checked:
                b_item_selected_in_show_list = True
                break
        for i in range( self.ui.qtHideListWidget.count() ):
            item = self.ui.qtHideListWidget.item(i)
            if item.checkState() == Qt.Checked:
                b_item_selected_in_hide_list = True
                break

        if b_item_selected_in_show_list:
            self.ui.qtHideSelectToolButton.setEnabled( True )
        else:
            self.ui.qtHideSelectToolButton.setEnabled( False )
        if b_item_selected_in_hide_list:
            self.ui.qtShowSelectToolButton.setEnabled( True )
        else:
            self.ui.qtShowSelectToolButton.setEnabled( False )

        if self.ui.qtHideListWidget.count() == 0:
            self.ui.qtShowAllToolButton.setEnabled( False )
        else:
            self.ui.qtShowAllToolButton.setEnabled( True )
        if self.ui.qtShowListWidget.count() == 0:
            self.ui.qtHideAllToolButton.setEnabled( False )
        else:
            self.ui.qtHideAllToolButton.setEnabled( True )

    def reset_to_default( self ):
        self.ui.qtShowListWidget.clear()
        self.ui.qtHideListWidget.clear()

        for e_type in self.list_default_show_stock_info:
            item = QListWidgetItem( f"{ g_dict_stock_info[ e_type ] }" )
            item.setFlags( item.flags() | Qt.ItemIsUserCheckable )
            item.setCheckState( Qt.Unchecked )
            item.setData( Qt.UserRole, e_type )
            self.ui.qtShowListWidget.addItem( item )

        for e_type in self.list_total_stock_info:
            if e_type in self.list_default_show_stock_info:
                continue
            item = QListWidgetItem( f"{ g_dict_stock_info[ e_type ] }" )
            item.setFlags( item.flags() | Qt.ItemIsUserCheckable )
            item.setCheckState( Qt.Unchecked )
            item.setData( Qt.UserRole, e_type )
            self.ui.qtHideListWidget.addItem( item )

        self.update_ui()

    def accept_data( self ):
        for i in range( self.ui.qtShowListWidget.count() ):
            item = self.ui.qtShowListWidget.item( i )
            e_type = item.data( Qt.UserRole )
            self.list_final_show_stock_info.append( e_type )

        self.accept()
    
    def cancel( self ):
        self.reject()

class UndoRedoManager:
    def __init__( self, max_steps = 20 ):
        self.undo_stack = []
        self.redo_stack = []
        self.max_steps = max_steps

    def save_state( self, state: dict ):
        # 存之前 deep copy 一份
        self.undo_stack.append( copy.deepcopy( state ) )
        if len( self.undo_stack ) > self.max_steps:
            self.undo_stack.pop( 0 )
        self.redo_stack.clear()

    def undo( self, current_state: dict ):
        if not self.undo_stack:
            return current_state  # 沒有可以 undo 的
        self.redo_stack.append( copy.deepcopy( current_state ) )
        return self.undo_stack.pop()

    def redo(self, current_state: dict):
        if not self.redo_stack:
            return current_state  # 沒有可以 redo 的
        self.undo_stack.append(copy.deepcopy(current_state))
        return self.redo_stack.pop()

class DiffUndoRedoManager:
    def __init__(self, max_steps=20):
        self.undo_stack = []
        self.redo_stack = []
        self.max_steps = max_steps

    def record_changes( self, old_dict: dict, new_dict: dict ):
        changes = []
        keys = set(old_dict.keys()).union(new_dict.keys())

        for key in keys:
            old_val = old_dict.get(key)
            new_val = new_dict.get(key)
            if old_val != new_val:
                changes.append({
                    'key': key,
                    'old_value': old_val,
                    'new_value': new_val
                })

        if changes:
            self.undo_stack.append(changes)
            if len(self.undo_stack) > self.max_steps:
                self.undo_stack.pop(0)
            self.redo_stack.clear()

    def undo(self, target_dict: dict):
        if not self.undo_stack:
            return

        changes = self.undo_stack.pop()
        reverse_changes = []
        for change in reversed(changes):
            key = change['key']
            reverse_changes.append({
                'key': key,
                'old_value': target_dict.get(key),
                'new_value': change['old_value']
            })
            # apply old value
            if change['old_value'] is None:
                target_dict.pop(key, None)  # key was added, so remove
            else:
                target_dict[key] = change['old_value']
        self.redo_stack.append(reverse_changes)

    def redo(self, target_dict: dict):
        if not self.redo_stack:
            return

        changes = self.redo_stack.pop()
        reverse_changes = []
        for change in reversed(changes):
            key = change['key']
            reverse_changes.append({
                'key': key,
                'old_value': target_dict.get(key),
                'new_value': change['new_value']
            })
            if change['new_value'] is None:
                target_dict.pop(key, None)
            else:
                target_dict[key] = change['new_value']
        self.undo_stack.append(reverse_changes)

class Utility():

    @staticmethod
    def round_to( u_value: Union[int, float, Decimal], e_round_type : DecimalRoundType) -> int:
        dec_value = Decimal(u_value)
        if e_round_type == DecimalRoundType.ROUND_DOWN:
            return int( dec_value )
        elif e_round_type == DecimalRoundType.ROUND_OFF:
            return int( dec_value.quantize( Decimal('1'), rounding=ROUND_HALF_UP ) )
        else:  # ROUND_UP or default
            return int( dec_value.quantize( Decimal('1'), rounding=ROUND_CEILING ) )

    @staticmethod
    def compute_cost( e_trading_type : TradingType, e_decimal_round_type : DecimalRoundType, f_trading_price, n_trading_count, f_trading_fee_discount, n_minimum_common_trading_fee, n_minimum_odd_trading_fee, b_etf, b_daying_trading, b_bond ):
        f_trading_price = Decimal( str( f_trading_price ) )#原本10.45 * 100000 = 1044999.999999999 然後取 int 就變成1044999，所以改用Decimal
        n_trading_count = Decimal( str( n_trading_count ) )
        f_trading_fee_discount = Decimal( str( f_trading_fee_discount ) )
        dict_result = {}
        if e_trading_type == TradingType.BUY or e_trading_type == TradingType.SELL:
            n_trading_value = int( f_trading_price * n_trading_count )
            n_trading_fee = Utility.round_to( n_trading_value * Decimal( '0.001425' ) * f_trading_fee_discount, e_decimal_round_type )
            
            if n_trading_value != 0:
                if n_trading_count % 1000 == 0:
                    n_trading_fee = max( n_minimum_common_trading_fee, n_trading_fee )
                else:#零股交易
                    n_trading_fee = max( n_minimum_odd_trading_fee, n_trading_fee )
            
            if e_trading_type == TradingType.SELL:
                if b_etf:
                    if b_bond:
                         n_trading_tax = 0
                    else:
                        n_trading_tax = int( n_trading_value * Decimal( '0.001' ) )
                elif b_daying_trading:
                    n_trading_tax = int( n_trading_value * Decimal( '0.0015' ) )
                else:
                    n_trading_tax = int( n_trading_value * Decimal( '0.003' ) )
            else:
                n_trading_tax = 0

            dict_result[ TradingCost.TRADING_VALUE ] = n_trading_value
            dict_result[ TradingCost.TRADING_FEE ] = n_trading_fee
            dict_result[ TradingCost.TRADING_TAX ] = n_trading_tax
            if e_trading_type == TradingType.BUY:
                dict_result[ TradingCost.TRADING_TOTAL_COST ] = n_trading_value + n_trading_fee + n_trading_tax
            else:
                dict_result[ TradingCost.TRADING_TOTAL_COST ] = n_trading_value - n_trading_fee - n_trading_tax
        else:   
            dict_result[ TradingCost.TRADING_VALUE ] = 0
            dict_result[ TradingCost.TRADING_FEE ] = 0
            dict_result[ TradingCost.TRADING_TAX ] = 0
            dict_result[ TradingCost.TRADING_TOTAL_COST ] = 0
        return dict_result

    @staticmethod
    def generate_trading_data( str_trading_date,                   #交易日期
                               e_trading_type,                     #交易種類
                               e_trading_price_type,               #交易價格種類
                               f_per_share_trading_price,          #每股交易價格
                               n_total_trading_price,              #總交易金額
                               n_trading_count,                    #交易股數
                               e_regular_buy_trading_fee_type,     #手續費種類
                               f_trading_fee_discount,             #手續費折扣
                               n_regular_buy_trading_fee_minimum,  #手續費最低金額
                               n_regular_buy_trading_fee_constant, #手續費固定金額
                               e_dividend_value_type,              #股利金額種類
                               f_stock_dividend,                   #股票股利
                               f_cash_dividend,                    #現金股利
                               n_custom_extra_insurance_fee,       #自訂的補充保費
                               f_capital_reduction_per_share,      #每股減資金額
                               e_capital_reduction_type,           #減資種類
                               b_daying_trading  ):                #是否為當沖交易
        dict_trading_data = {}
        dict_trading_data[ TradingData.TRADING_DATE ] = str_trading_date
        dict_trading_data[ TradingData.TRADING_TYPE ] = e_trading_type
        dict_trading_data[ TradingData.TRADING_PRICE_TYPE ] = e_trading_price_type
        dict_trading_data[ TradingData.PER_SHARE_TRADING_PRICE ] = f_per_share_trading_price
        dict_trading_data[ TradingData.TOTAL_TRADING_PRICE ] = n_total_trading_price
        dict_trading_data[ TradingData.TRADING_QUANTITY ] = n_trading_count
        dict_trading_data[ TradingData.REGULAR_BUY_TRADING_FEE_TYPE ] = e_regular_buy_trading_fee_type
        dict_trading_data[ TradingData.TRADING_FEE_DISCOUNT ] = f_trading_fee_discount
        dict_trading_data[ TradingData.REGULAR_BUY_TRADING_FEE_MINIMUM ] = n_regular_buy_trading_fee_minimum
        dict_trading_data[ TradingData.REGULAR_BUY_TRADING_FEE_CONSTANT ] = n_regular_buy_trading_fee_constant
        dict_trading_data[ TradingData.DIVIDEND_VALUE_TYPE ] = e_dividend_value_type
        dict_trading_data[ TradingData.STOCK_DIVIDEND ] = f_stock_dividend
        dict_trading_data[ TradingData.CASH_DIVIDEND ] = f_cash_dividend
        dict_trading_data[ TradingData.CUSTOM_EXTRA_INSURANCE_FEE ] = n_custom_extra_insurance_fee
        dict_trading_data[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] = f_capital_reduction_per_share
        dict_trading_data[ TradingData.CAPITAL_REDUCTION_TYPE ] = e_capital_reduction_type
        dict_trading_data[ TradingData.DAYING_TRADING ] = b_daying_trading
        return dict_trading_data

    @staticmethod
    def xirr(cash_flows, dates):
        """
        計算 XIRR (年化報酬率)
        :param cash_flows: list of cash flows (現金流金額，正數為收入，負數為支出)
        :param dates: list of datetime objects (對應的日期)
        :return: 年化報酬率
        """
        def xnpv(rate):
            """
            計算 XNPV
            """
            return sum(cf / ((1 + rate) ** ((d - dates[0]).days / 365.0)) for cf, d in zip(cash_flows, dates))

        # 初始猜測的 XIRR 值
        try:
            return newton(lambda r: xnpv(r), 0.1)
        except RuntimeError:
            raise ValueError("無法找到解，請檢查輸入的現金流和日期。")
           
    @staticmethod
    def update_weekly_text_by_date( qt_date_edit, qt_weekday_label ):
        obj_date = qt_date_edit.date()
        n_weekday = obj_date.dayOfWeek()
        str_weekday = share_api.get_qt_weekday_text( n_weekday )
        qt_weekday_label.setText( str_weekday )

    @staticmethod
    def get_tick_unit( f_price: float, b_etf: bool ) -> float:
        if b_etf:
            if f_price < 50:
                return 0.01
            else:
                return 0.05
        else:
            if f_price < 10:
                return 0.01
            elif f_price < 50:
                return 0.05
            elif f_price < 100:
                return 0.1
            elif f_price < 500:
                return 0.5
            elif f_price < 1000:
                return 1
            else:
                return 5
            
    @staticmethod
    def round_to_tick_unit( f_price: float, b_etf: bool ) -> float:
        f_tick = Utility.get_tick_unit( f_price, b_etf )
        return round( f_price / f_tick ) * f_tick
    
    @staticmethod
    def ceil_to_tick_unit( f_price: float, b_etf: bool ) -> float:
        f_tick = Utility.get_tick_unit( f_price, b_etf )
        return math.ceil( f_price / f_tick ) * f_tick
    
    @staticmethod
    def floor_to_tick_unit( f_price: float, b_etf: bool ) -> float:
        f_tick = Utility.get_tick_unit( f_price, b_etf )
        return math.floor( f_price / f_tick ) * f_tick

    @staticmethod
    def find_min_break_even_price_from_net( n_total_cost: int, n_total_quantity: int, b_etf: bool, b_bond: bool, n_minimum_trading_fee ) -> float:
        # 初始估算從 y 開始
        f_guess_price = float( n_total_cost / n_total_quantity )
        # 先往上對齊 tick 單位
        f_guess_price = Utility.ceil_to_tick_unit( f_guess_price, b_etf )
        max_iter = 100000  # 防止死迴圈
        count = 0

        while count < max_iter:
            n_trading_fee = max( int( f_guess_price * n_total_quantity * 0.001425 ), n_minimum_trading_fee )
            if b_etf:
                if b_bond:
                    n_trading_tax = 0
                else:
                    n_trading_tax = int( f_guess_price * n_total_quantity * 0.001 )
            else:
                n_trading_tax = int( f_guess_price * n_total_quantity * 0.003 )
            if f_guess_price * n_total_quantity - n_trading_fee - n_trading_tax >= n_total_cost:
                return f_guess_price
            # 下一個 tick 單位
            tick = Utility.get_tick_unit( f_guess_price, b_etf )
            f_guess_price += tick
            f_guess_price = round(f_guess_price, 4)  # 防止浮點誤差
            count += 1
        return 0
        # raise ValueError("找不到符合條件的 X（超過最大迭代次數）")

    @staticmethod
    def get_up_down_color( str_base, str_compare ):
        str_base = str_base.replace( ',', '' ).replace( '%', '' )
        str_compare = str_compare.replace( ',', '' ).replace( '%', '' )
        try:
            f_base = float(str_base)
            f_compare = float(str_compare)
        except ValueError:
            return QBrush('#FFFFFF')

        if f_compare > f_base:
            return QBrush('#FF0000')  # 漲紅
        elif f_compare < f_base:
            return QBrush('#00AA00')  # 跌綠
        else:
            return QBrush('#FFFFFF')  # 持平白
        
    @staticmethod
    def color_text_by_value(text ):
        try:
            # 把逗號移除再轉 float 判斷正負
            value = float(text.replace(",", "").replace("%", "").replace("％", ""))
            color = "#FF0000" if value > 0 else ("#00AA00" if value < 0 else "#FFFFFF")
        except:
            color = "#555555"  # 無法轉換就用預設
        return f'<span style="color:{color};">{text}</span>'

class Worker( QObject ):
    progress = Signal( int )  # Signal to emit progress updates
    finished = Signal( dict )  # Signal to emit the result when done

    def __init__( self, main_window ):
        super().__init__()
        self.main_window = main_window  # 傳入 MainWindow 的實例

    def run( self ):
        # Perform the time-consuming operation (e.g. loading stock data)
        self.main_window.initialize( False, self.update_progress )
        self.finished.emit({})  # Emit the result when done

    def update_progress( self, value ):
        self.progress.emit( value )  # Emit progress updates

class MainWindow( QMainWindow ):
    def __init__( self, b_unit_test = False, 
                  str_initial_data_file = 'TradingData.json', 
                  str_UI_setting_file = 'UISetting.config', 
                  str_stock_number_file = 'StockNumber.txt',
                  str_suspend_stock_number_file = 'SuspendStockNumber.txt',
                  str_stock_price_file = 'StockPrice.txt',
                  str_stock_pre_price_file = 'PreStockPrice.txt',
                  str_stock_dividend_position_date_file = 'StockDividendPositionDate.json' ):
        super( MainWindow, self ).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi( self )  # 設置 UI
        self.setWindowIcon( share_icon.get_icon( share_icon.IconType.WINDOW ) )

        if not b_unit_test:
            self.progress_bar = QProgressBar( self )
            self.progress_bar.setGeometry( 400, 350, 300, 25 )  # Adjust position and size as needed
            self.progress_bar.setMaximum( 100 )
            self.progress_bar.setVisible( False )
        size = reg_settings.value( "window_size", QSize( 1250, 893 ) )
        self.resize(size)

        self.ui.qtTabWidget.currentChanged.connect( self.on_tab_current_changed )
        self.ui.qtTabWidget.tabBarDoubleClicked.connect( self.on_tab_widget_double_clicked )
        self.ui.qtTabWidget.tabCloseRequested.connect( self.on_tab_widget_close )
        self.ui.qtTabWidget.tabBar().tabMoved.connect( self.on_tab_moved )
        self.ui.qtTabWidget.tabBar().setTabButton( 0, QTabBar.RightSide, None )  # 隱藏最後一個 tab 的 close 按鈕

        delegate = share_ui.CenterIconDelegate()

        self.per_stock_trading_data_model = QStandardItemModel( 0, 0 ) 
        self.per_stock_trading_data_model.setVerticalHeaderLabels( self.get_trading_data_header() )
        self.ui.qtTradingDataTableView.setModel( self.per_stock_trading_data_model )
        self.ui.qtTradingDataTableView.setItemDelegate( delegate )
        self.ui.qtTradingDataTableView.horizontalHeader().setSectionsMovable( True )
        self.ui.qtTradingDataTableView.horizontalHeader().sectionMoved.connect( self.on_trading_data_table_horizontal_header_section_moved )
        self.ui.qtTradingDataTableView.verticalHeader().setSectionResizeMode( QHeaderView.Fixed )
        self.ui.qtTradingDataTableView.clicked.connect( lambda index: self.on_trading_data_table_item_clicked( index, self.per_stock_trading_data_model ) )
        self.ui.qtTradingDataTableView.setContextMenuPolicy( Qt.CustomContextMenu )
        self.ui.qtTradingDataTableView.customContextMenuRequested.connect( self.show_context_menu )
        for row in range( len( self.get_trading_data_header() ) ):
            if row == 7 or row == 8:
                self.ui.qtTradingDataTableView.setRowHeight( row, 40 )
            else:
                self.ui.qtTradingDataTableView.setRowHeight( row, 25 )

        self.ui.qtHideTradingDataTableToolButton.clicked.connect( self.on_hide_trading_data_table_tool_button_clicked )

        self.ui.qtAddTradingDataPushButton.clicked.connect( self.on_add_trading_data_push_button_clicked )
        self.ui.qtAddRegularTradingDataPushButton.clicked.connect( self.on_add_regular_trading_data_push_button_clicked )
        self.ui.qtAddDividendDataPushButton.clicked.connect( self.on_add_dividend_data_push_button_clicked )
        self.ui.qtAddLimitBuyingDataPushButton.clicked.connect( self.on_add_limit_buying_data_push_button_clicked )
        self.ui.qtAddCapitalReductionDataPushButton.clicked.connect( self.on_add_capital_reduction_data_push_button_clicked )
        self.ui.qtAddStockSplitDataPushButton.clicked.connect( self.on_add_stock_split_data_push_button_clicked )
        self.ui.qtExportAllStockTradingDataPushButton.clicked.connect( self.on_export_all_to_excell_button_clicked )
        self.ui.qtExportSelectedStockTradingDataPushButton.clicked.connect( self.on_export_selected_to_excell_button_clicked )

        self.ui.qtHideTradingDataTableToolButton.setIcon( share_icon.get_icon( share_icon.IconType.MOVE_DOWN ) )

        self.ui.qtNewFileAction.setShortcut( "Ctrl+N" )
        self.ui.qtNewFileAction.triggered.connect( self.on_new_file_action_triggered )
        self.ui.qtOpenFileAction.setShortcut( "Ctrl+O" )
        self.ui.qtOpenFileAction.triggered.connect( self.on_open_file_action_triggered )
        self.ui.qtSaveAsAction.setShortcut( "Ctrl+A" )
        self.ui.qtSaveAsAction.triggered.connect( self.on_save_as_action_triggered )
        self.ui.qtSaveAction.setShortcut( "Ctrl+S" )
        self.ui.qtSaveAction.triggered.connect( self.on_save_action_triggered )
        self.ui.qtExportCurrentGroupAction.setShortcut( "Ctrl+E" )
        self.ui.qtExportCurrentGroupAction.triggered.connect( self.on_export_current_group_action_triggered )
        self.ui.qtImportFullAction.setShortcut( "Ctrl+U" )
        self.ui.qtImportFullAction.triggered.connect( self.on_import_full_action_triggered )
        self.ui.qtImportSingleStockAction.setShortcut( "Ctrl+I" )
        self.ui.qtImportSingleStockAction.triggered.connect( self.on_import_single_stock_action_triggered )
        self.ui.qtSetTemporaryFilePathAction.triggered.connect( self.on_set_temporary_file_path_action_triggered )

        self.ui.qtUndoAction.triggered.connect( self.on_undo_action_triggered )
        self.ui.qtRedoAction.triggered.connect( self.on_redo_action_triggered )

        self.ui.qtFromNewToOldAction.setChecked( True )
        self.ui.qtFromOldToNewAction.setChecked( False )
        self.ui.qtFromNewToOldAction.triggered.connect( self.on_trigger_from_new_to_old )
        self.ui.qtFromOldToNewAction.triggered.connect( self.on_trigger_from_old_to_new )

        self.ui.qtShowAllAction.setChecked( True )
        self.ui.qtShow10Action.setChecked( False )
        self.ui.qtShowAllAction.triggered.connect( self.on_trigger_show_all )
        self.ui.qtShow10Action.triggered.connect( self.on_trigger_show_10 )

        self.ui.qtUse1ShareUnitAction.setChecked( True )
        self.ui.qtUse1000ShareUnitAction.setChecked( False )
        self.ui.qtUse1ShareUnitAction.triggered.connect( self.on_trigger_use_1_share_unit )
        self.ui.qtUse1000ShareUnitAction.triggered.connect( self.on_trigger_use_1000_share_unit )

        self.ui.qtADYearAction.setChecked( True )
        self.ui.qtROCYearAction.setChecked( False )
        self.ui.qtADYearAction.triggered.connect( self.on_trigger_AD_year )
        self.ui.qtROCYearAction.triggered.connect( self.on_trigger_ROC_year )

        self.ui.qtEditShowItemAction.triggered.connect( self.on_trigger_edit_show_item )

        self.ui.qtCostWithInDividendAction.setChecked( True )
        self.ui.qtCostWithOutDividendAction.setChecked( False )
        self.ui.qtCostWithInDividendAction.triggered.connect( self.on_trigger_cost_with_in_dividend )
        self.ui.qtCostWithOutDividendAction.triggered.connect( self.on_trigger_cost_with_out_dividend )

        self.ui.qtAboutAction.triggered.connect( self.on_trigger_about )
        
        if not b_unit_test:
            self.trading_data_json_file_path = os.path.join( g_data_dir, 'StockInventory', str_initial_data_file )
            self.UISetting_file_path = os.path.join( g_data_dir, 'StockInventory', str_UI_setting_file )
            self.stock_number_file_path = os.path.join( g_data_dir, 'StockInventory', str_stock_number_file )
            self.suspend_stock_number_file_path = os.path.join( g_data_dir, 'StockInventory', str_suspend_stock_number_file )
            self.stock_price_file_path = os.path.join( g_data_dir, 'StockInventory', str_stock_price_file )
            self.stock_pre_price_file_path = os.path.join( g_data_dir, 'StockInventory', str_stock_pre_price_file )
            self.stock_dividend_position_date_file_path = os.path.join( g_data_dir, 'StockInventory', str_stock_dividend_position_date_file )
        else:
            self.trading_data_json_file_path = os.path.join( g_exe_dir, 'StockInventory', str_initial_data_file )
            self.UISetting_file_path = os.path.join( g_exe_dir, 'StockInventory', str_UI_setting_file )
            self.stock_number_file_path = os.path.join( g_exe_dir, 'StockInventory', str_stock_number_file )
            self.suspend_stock_number_file_path = os.path.join( g_exe_dir, 'StockInventory', str_suspend_stock_number_file )
            self.stock_price_file_path = os.path.join( g_exe_dir, 'StockInventory', str_stock_price_file )
            self.stock_pre_price_file_path = os.path.join( g_exe_dir, 'StockInventory', str_stock_pre_price_file )
            self.stock_dividend_position_date_file_path = os.path.join( g_exe_dir, 'StockInventory', str_stock_dividend_position_date_file )

        self.list_stock_list_table_horizontal_header = []
        self.pick_up_stock( None )
        self.header_helpers = {}
        self.dict_all_account_ui_state = {}
        self.dict_all_account_general_data = {}
        self.dict_all_account_cash_transfer_data = {}
        self.dict_all_account_all_stock_trading_data = {}
        self.dict_all_account_all_stock_trading_data_INITIAL = {}
        self.list_total_stock_info_INITIAL = [ 
                                               StockInfoType.LATEST_PRICE, #收盤價
                                               StockInfoType.QUANTITY, #庫存股數
                                               StockInfoType.LATEST_MARKET_VALUE, #現值
                                               StockInfoType.CURRENT_COST, #現股成本
                                               StockInfoType.UNREALIZED_PROFIT, #未實現損益
                                               StockInfoType.UNREALIZED_PROFIT_RATIO, #未實現報酬率
                                               StockInfoType.REALIZED_PROFIT, #已實現損益
                                               StockInfoType.BREAK_EVEN_PRICE, #損益平衡價
                                               StockInfoType.ACCUMULATED_COST, #累計成本
                                               StockInfoType.ACCUMULATED_AVERAGE_COST, #累計平均成本
                                               StockInfoType.ACCUMULATED_PROFIT, #累計損益
                                               StockInfoType.ACCUMULATED_TRADING_FEE, #累計手續費
                                               StockInfoType.ACCUMULATED_TAX, #累計交易稅
                                               StockInfoType.ACCUMULATED_DIVIDEND_INCOME, #累計股利所得
                                               StockInfoType.XIRR_VALUE, #平均年化報酬率
                                               StockInfoType.HOLDING_MARKET_RATIO, #持股淨值比
                                               ]

        self.list_show_stock_info = copy.deepcopy( self.list_total_stock_info_INITIAL )
        self.list_stock_list_column_width = [ 85 ] * ( len( list( StockInfoType ) ) + 1 )# +1 是給倒數第3欄的自動股利欄位
        self.list_stock_list_column_width.insert( 0, 140 ) #股票代碼
        self.list_stock_list_column_width.append( 40 ) #匯出欄位
        self.list_stock_list_column_width.append( 40 ) #刪除欄位
        self.n_current_tab = 0
        self.n_tab_index = 0
        self.str_current_save_file_path = None
        self.load_stylesheet( styles_css_path )
        if b_unit_test:
            self.initialize( True, None )
            self.load_initialize_data()
        else:
            if getattr( sys, 'frozen', False ):
                self.copy_required_data()
            self.start_loading_stock_data()

    def closeEvent(self, event):
        """當視窗關閉時，儲存當前大小"""
        reg_settings.setValue("window_size", self.size())
        super().closeEvent(event)

    def copy_required_data( self ):
        source_folder = os.path.join( g_exe_root_dir, 'StockInventory\\Dividend' ) 
        destination_folder = os.path.join( g_data_dir, 'StockInventory\\Dividend' ) 

        os.makedirs( destination_folder, exist_ok = True )
        file_extension = '.txt'
        # 遍歷來源資料夾的檔案
        for filename in os.listdir( source_folder ):
            source_file = os.path.join( source_folder, filename )
            destination_file = os.path.join( destination_folder, filename )
            
            # 檢查是否為檔案且符合條件
            if os.path.isfile( source_file ) and filename.endswith( file_extension ):
                shutil.copy2(source_file, destination_file )

    def download_all_required_data( self, update_progress_callback ):
        obj_today_date = datetime.datetime.today()
        obj_yesterday_date = datetime.datetime.today() - datetime.timedelta( days = 1 )
        str_today_date = obj_today_date.strftime('%Y%m%d')
        str_yesterday_date = obj_yesterday_date.strftime('%Y%m%d')
        n_current_hour = obj_today_date.hour

        self.set_progress_value( update_progress_callback, 0 )
        self.download_all_company_stock_number( str_yesterday_date )
        self.download_all_suspend_company_stock_number( str_yesterday_date )
        self.set_progress_value( update_progress_callback, 10 )
        
        list_obj_date = [ obj_yesterday_date, obj_today_date ]
        if n_current_hour < 14:
            list_file_path = [ self.stock_price_file_path, self.stock_price_file_path ]
        else:
            list_file_path = [ self.stock_pre_price_file_path, self.stock_price_file_path ]

        for obj_date, file_path in zip(list_obj_date, list_file_path):
            if obj_date == obj_today_date and n_current_hour < 14:
                break
            n_retry = 0
            while n_retry < 30:
                n_weekday = obj_date.weekday()
                if n_weekday >= 5:
                    obj_date -= datetime.timedelta( days = ( 2 if n_weekday == 6 else 1 ) )
                    continue
                str_date = obj_date.strftime( '%Y%m%d' )
                if self.download_day_stock_price( str_date, file_path ):
                    break
                obj_date -= datetime.timedelta( days = 1 )
                n_retry += 1

        self.set_progress_value( update_progress_callback, 20 )
        self.download_general_company_all_yearly_dividend_data( 2025, str_yesterday_date )
        self.download_listed_company_all_yearly_split_data( 2020, str_yesterday_date )
        self.download_listed_etf_all_yearly_split_data( 2023, str_yesterday_date )
        self.set_progress_value( update_progress_callback, 50 )
        self.download_listed_etf_all_yearly_dividend_data( 2025, str_yesterday_date )
        self.set_progress_value( update_progress_callback, 80 )
        self.download_OTC_etf_all_yearly_dividend_data( 2025, str_yesterday_date )

    def initialize( self, b_unit_test, update_progress_callback ):
        self.setEnabled( False ) # 資料下載前先Disable整個視窗
        
        if not b_unit_test:
            self.download_all_required_data( update_progress_callback )
        
        self.dict_all_company_number_to_name_and_type = self.load_all_company_stock_number()
        self.dict_all_suspend_company_number_to_name_and_type = self.load_all_suspend_company_stock_number()
        self.dict_all_company_number_to_name_and_type.update( self.dict_all_suspend_company_number_to_name_and_type )
        list_date_and_price = self.load_day_stock_price()
        self.str_latest_price_column_header = '收盤價'
        if len( list_date_and_price[ 1 ] ) != 0:
            self.dict_all_company_number_to_price_info = list_date_and_price[ 1 ]
            self.list_previous_day_data = list_date_and_price[ 2 ]
            parsed_date = datetime.datetime.strptime(list_date_and_price[ 0 ], "%Y%m%d")  # 解析成日期對象
            str_formatted_date = parsed_date.strftime("%m/%d")  # 轉換為 MM/DD 格式
            self.str_latest_price_column_header = str_formatted_date + ' 收盤價'

        self.dict_auto_stock_yearly_dividned = self.load_general_company_all_yearly_dividend_data( 2005, b_unit_test )
        self.dict_auto_stock_listed_etf_yearly_dividned = self.load_listed_etf_all_yearly_dividend_data( 2010, b_unit_test )
        self.dict_auto_stock_OTC_etf_yearly_dividned = self.load_OTC_etf_all_yearly_dividend_data( 2010, b_unit_test )

        for key_stock_number, value in self.dict_auto_stock_OTC_etf_yearly_dividned.items():
            # 00687B 這支 ETF 應該是櫃買中心的資料，但是出現在證交所(重點是資料不完整)，所以要排除
            # 為了避免有相同的問題，我們就直接把 self.dict_auto_stock_listed_etf_yearly_dividned 排除所有櫃買中心的 ETF
            if key_stock_number in self.dict_auto_stock_listed_etf_yearly_dividned:
                del self.dict_auto_stock_listed_etf_yearly_dividned[ key_stock_number ]

        # common_keys = set(self.dict_auto_stock_yearly_dividned.keys()) & set(self.dict_auto_stock_listed_etf_yearly_dividned.keys())
        self.dict_auto_stock_yearly_dividned.update( self.dict_auto_stock_listed_etf_yearly_dividned )
        for key, value in self.dict_auto_stock_OTC_etf_yearly_dividned.items():
            if key not in self.dict_auto_stock_yearly_dividned:
                self.dict_auto_stock_yearly_dividned[ key ] = value

        self.dict_all_stock_dividend_position_date = self.load_stock_dividend_position_date()

        self.set_progress_value( update_progress_callback, 100 )
        # self.load_initialize_data()
        self.setEnabled( True ) # 資料下載完後就會Enable

    def set_progress_value( self, update_progress_callback, f_progress ):
        if update_progress_callback:
            update_progress_callback( f_progress )

    def start_loading_stock_data( self ):
        # Show the progress bar
        self.progress_bar.setVisible( True )

        # Create and start a QThread for the worker
        self.thread = QThread()
        self.worker = Worker(self)

        # Move the worker to the new thread
        self.worker.moveToThread( self.thread )

        # Connect signals and slots
        self.worker.progress.connect( self.update_progress )
        self.worker.finished.connect( self.on_loading_finished )
        self.thread.started.connect( self.worker.run )
        self.worker.finished.connect( self.thread.quit )

        # Start the thread
        self.thread.start()

    def update_progress( self, value ):
        self.progress_bar.setValue( value )

    def on_loading_finished( self ):
        self.progress_bar.setVisible( False )
        self.load_initialize_data()

    def load_stylesheet( self, file_path ):
        try:
            with open(file_path, "r", encoding="utf-8") as file:  # 指定 UTF-8 編碼
                stylesheet = file.read()
                self.setStyleSheet(stylesheet)
        except FileNotFoundError:
            print(f"CSS 檔案 {file_path} 找不到")
        except Exception as e:
            print(f"讀取 CSS 檔案時發生錯誤: {e}")

    def keyReleaseEvent(self, event):
        key = event.key()
        modifiers = QApplication.keyboardModifiers()

        if key == Qt.Key_T:
            if self.ui.qtAddTradingDataPushButton.isEnabled():
                self.on_add_trading_data_push_button_clicked()
        elif key == Qt.Key_E:
            if self.ui.qtAddRegularTradingDataPushButton.isEnabled():
                self.on_add_regular_trading_data_push_button_clicked()
        elif key == Qt.Key_D:
            if self.ui.qtAddDividendDataPushButton.isEnabled():
                self.on_add_dividend_data_push_button_clicked()
        elif key == Qt.Key_A:
            if self.ui.qtAddLimitBuyingDataPushButton.isEnabled():
                self.on_add_limit_buying_data_push_button_clicked()
        elif key == Qt.Key_R:
            if self.ui.qtAddCapitalReductionDataPushButton.isEnabled():
                self.on_add_capital_reduction_data_push_button_clicked()
        elif key == Qt.Key_T:
            if self.ui.qtAddStockSplitDataPushButton.isEnabled():
                self.on_add_stock_split_data_push_button_clicked()
        elif key == Qt.Key_Enter:
            qt_line_edit = self.ui.qtTabWidget.currentWidget().findChild( QLineEdit, "StockInputLineEdit" )
            if qt_line_edit and qt_line_edit.hasFocus():
                self.on_add_stock_push_button_clicked()

    def add_new_tab_and_table( self, str_tab_title = None ): 
        str_tab_name = f"TabIndex{ self.n_tab_index }"
        increased_tab = QWidget()
        increased_tab.setObjectName( str_tab_name )
               
        uiqt_vertical_layout_main = QVBoxLayout( increased_tab )
        uiqt_vertical_layout_main.setSpacing(0)
        uiqt_vertical_layout_main.setContentsMargins(-1, 0, -1, 0)

        uiqt_stock_inventory_and_cash_transfer_tab_widget = QTabWidget( increased_tab )
        uiqt_stock_inventory_and_cash_transfer_tab_widget.setObjectName(u"qtStockInventoryAndCashTransferTabWidget")
        uiqt_stock_inventory_and_cash_transfer_tab_widget.setTabPosition( QTabWidget.West )

        stock_inventory_tab = QWidget()
        stock_inventory_tab.setObjectName( str_tab_name + "StockInventoryTab" )
        cash_transfer_tab = QWidget()
        cash_transfer_tab.setObjectName( str_tab_name + "CashTransferTab" )
        uiqt_stock_inventory_and_cash_transfer_tab_widget.insertTab( 0, stock_inventory_tab, "庫存總覽" )
        uiqt_stock_inventory_and_cash_transfer_tab_widget.insertTab( 1, cash_transfer_tab, "銀行" )
        
        uiqt_vertical_layout_main.addWidget( uiqt_stock_inventory_and_cash_transfer_tab_widget )

        # stock inventory tab
        uiqt_stock_inventory_tab_vertical_layout = QVBoxLayout( stock_inventory_tab )
        uiqt_stock_inventory_tab_vertical_layout.setSpacing( 0 )
        uiqt_stock_inventory_tab_vertical_layout.setContentsMargins( -1, 0, -1, 0 )
        uiqt_horizontal_layout_1 = QHBoxLayout()
        uiqt_stock_input_line_edit = QLineEdit( stock_inventory_tab )
        sizePolicy = QSizePolicy( QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed )
        sizePolicy.setHorizontalStretch( 0 )
        sizePolicy.setVerticalStretch( 0 )
        sizePolicy.setHeightForWidth( uiqt_stock_input_line_edit.sizePolicy().hasHeightForWidth() )

        uiqt_stock_input_line_edit.setSizePolicy( sizePolicy )
        uiqt_stock_input_line_edit.setMinimumSize( QSize( 150, 0 ) )
        uiqt_stock_input_line_edit.setMaximumSize( QSize( 150, 16777215 ) )
        uiqt_stock_input_line_edit.setObjectName( "StockInputLineEdit" )
        uiqt_horizontal_layout_1.addWidget( uiqt_stock_input_line_edit )

        uiqt_horizontal_spacer_1_0 = QSpacerItem( 10, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum )
        uiqt_horizontal_layout_1.addItem( uiqt_horizontal_spacer_1_0 )

        uiqt_add_stock_push_button = QPushButton( stock_inventory_tab )
        uiqt_add_stock_push_button.setMinimumSize( QSize( 75, 0 ) )
        uiqt_add_stock_push_button.setMaximumSize( QSize( 75, 16777215 ) )
        uiqt_add_stock_push_button.setText( "新增股票" )
        uiqt_add_stock_push_button.setObjectName( "AddStockPushButton" )
        uiqt_horizontal_layout_1.addWidget( uiqt_add_stock_push_button )

        uiqt_horizontal_spacer_1_1 = QSpacerItem( 30, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum )
        uiqt_horizontal_layout_1.addItem( uiqt_horizontal_spacer_1_1 )

        uiqt_display_type_combobox = QComboBox( stock_inventory_tab )
        uiqt_display_type_combobox.addItem( "顯示所有交易" )
        uiqt_display_type_combobox.addItem( "僅顯示目前庫存" )
        uiqt_display_type_combobox.addItem( "僅顯示已無庫存" )
        uiqt_display_type_combobox.setObjectName( "DisplayTypeComboBox" )
        uiqt_horizontal_layout_1.addWidget( uiqt_display_type_combobox )

        uiqt_horizontal_spacer_1_2 = QSpacerItem( 30, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum )
        uiqt_horizontal_layout_1.addItem( uiqt_horizontal_spacer_1_2 )

        uiqt_account_setting_push_button = QPushButton( stock_inventory_tab )
        uiqt_account_setting_push_button.setMinimumSize( QSize( 90, 0 ) )
        uiqt_account_setting_push_button.setMaximumSize( QSize( 90, 16777215 ) )
        uiqt_account_setting_push_button.setText( "各項計算設定" )
        uiqt_account_setting_push_button.setObjectName( "AccountSettingPushButton" )
        uiqt_horizontal_layout_1.addWidget( uiqt_account_setting_push_button )

        uiqt_horizontal_spacer_1_3 = QSpacerItem( 30, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum )
        uiqt_horizontal_layout_1.addItem( uiqt_horizontal_spacer_1_3 )

        uiqt_total_info_label = QLabel( stock_inventory_tab )
        uiqt_total_info_label.setObjectName( "TotalInfoLabel" )
        uiqt_horizontal_layout_1.addWidget( uiqt_total_info_label )
        uiqt_horizontal_spacer_1_4 = QSpacerItem( 30, 20, QSizePolicy.Policy.MinimumExpanding, QSizePolicy.Policy.Minimum )
        uiqt_horizontal_layout_1.addItem( uiqt_horizontal_spacer_1_4 )

        uiqt_stock_inventory_tab_vertical_layout.addLayout( uiqt_horizontal_layout_1 )

        uiqt_horizontal_layout_2 = QHBoxLayout()
        uiqt_horizontal_layout_2.setSpacing( 0 )

        uiqt_stock_select_combo_box = QComboBox( stock_inventory_tab )
        uiqt_stock_select_combo_box.setMinimumSize( QSize( 200, 0 ) )
        uiqt_stock_select_combo_box.setObjectName( "StockSelectComboBox" )
        uiqt_horizontal_layout_2.addWidget( uiqt_stock_select_combo_box )

        uiqt_horizontal_spacer_2_1 = QSpacerItem( 40, 0, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum )
        uiqt_horizontal_layout_2.addItem( uiqt_horizontal_spacer_2_1 )

        uiqt_stock_inventory_tab_vertical_layout.addLayout( uiqt_horizontal_layout_2 )

        uiqt_horizontal_layout_3 = QHBoxLayout()

        uiqt_stock_list_table_view = QTableView( stock_inventory_tab )
        uiqt_stock_list_table_view.setMinimumSize( QSize( 0, 100 ) )
        uiqt_stock_list_table_view.setObjectName( "StockListTableView" )
        uiqt_horizontal_layout_3.addWidget( uiqt_stock_list_table_view )

        uiqt_stock_inventory_tab_vertical_layout.addLayout( uiqt_horizontal_layout_3 )

        delegate = share_ui.CenterIconDelegate()
        stock_list_model = share_ui.CustomSortModel( 0, 0 )
        stock_list_model.setHorizontalHeaderLabels( self.list_stock_list_table_horizontal_header )
        uiqt_stock_list_table_view.verticalHeader().setSectionsMovable( True )
        uiqt_stock_list_table_view.verticalHeader().sectionMoved.connect( self.on_stock_list_table_vertical_header_section_moved )
        uiqt_stock_list_table_view.verticalHeader().sectionClicked.connect( self.on_stock_list_table_vertical_section_clicked )
        uiqt_stock_list_table_view.verticalHeader().setSectionResizeMode( QHeaderView.Fixed )
        self.header_helpers[ str_tab_name ] = share_ui.InteractiveVerticalHeaderHelper( uiqt_stock_list_table_view )
        uiqt_stock_list_table_view.horizontalHeader().sectionResized.connect( self.on_stock_list_table_horizontal_section_resized )
        uiqt_stock_list_table_view.setSortingEnabled( True )
        uiqt_stock_list_table_view.setModel( stock_list_model )
        uiqt_stock_list_table_view.setItemDelegate( delegate )
        uiqt_stock_list_table_view.clicked.connect( lambda index: self.on_stock_list_table_item_clicked( index, stock_list_model ) )
        # uiqt_stock_list_table_view.horizontalHeader().sortIndicatorChanged.connect( self.update_stock_list_vertical_header )
        uiqt_stock_list_table_view.horizontalHeader().sectionDoubleClicked.connect( self.refresh_stock_list_table)

        uiqt_stock_input_line_edit.textChanged.connect( self.on_stock_input_text_changed ) 

        uiqt_stock_select_combo_box.setVisible( False )
        uiqt_stock_select_combo_box.activated.connect( self.on_stock_select_combo_box_current_index_changed )
        uiqt_stock_select_combo_box.setStyleSheet( "QComboBox { combobox-popup: 0; }" )
        uiqt_stock_select_combo_box.setMaxVisibleItems( 10 )

        uiqt_add_stock_push_button.clicked.connect( self.on_add_stock_push_button_clicked )
        uiqt_account_setting_push_button.clicked.connect( self.on_account_setting_push_button_clicked )
        uiqt_display_type_combobox.activated.connect( self.on_display_type_combo_box_current_index_changed )
        
        # Cash Transfer Tab
        uiqt_cash_transfer_tab_vertical_layout = QVBoxLayout( cash_transfer_tab )
        uiqt_cash_transfer_tab_vertical_layout.setSpacing( 0 )
        uiqt_cash_transfer_tab_vertical_layout.setContentsMargins( -1, 0, -1, 0 )

        uiqt_vertical_spacer_2_0 = QSpacerItem(20, 5, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum)
        uiqt_cash_transfer_tab_vertical_layout.addItem( uiqt_vertical_spacer_2_0 )

        uiqt_horizontal_layout_4 = QHBoxLayout()
        uiqt_horizontal_layout_4.setSpacing( 0 )

        uiqt_cash_transfer_table_view = QTableView( cash_transfer_tab )
        uiqt_cash_transfer_table_view.setMinimumSize( QSize( 0, 166 ) )
        uiqt_cash_transfer_table_view.setMaximumSize( QSize( 16777215, 166 ) )
        uiqt_cash_transfer_table_view.setObjectName( "CashTransferTableView" )
        uiqt_horizontal_layout_4.addWidget( uiqt_cash_transfer_table_view )

        uiqt_cash_transfer_tab_vertical_layout.addLayout( uiqt_horizontal_layout_4 )

        uiqt_horizontal_layout_5 = QHBoxLayout()
        uiqt_horizontal_layout_5.setSpacing( 0 )

        # uiqt_horizontal_spacer_5_1 = QSpacerItem(3, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum)
        # uiqt_horizontal_layout_5.addItem( uiqt_horizontal_spacer_5_1 )

        uiqt_add_cash_transfer_push_button = QPushButton( cash_transfer_tab )
        uiqt_add_cash_transfer_push_button.setMaximumSize( QSize( 16777215, 16777215 ) )
        uiqt_add_cash_transfer_push_button.setMinimumWidth( 200 )
        uiqt_add_cash_transfer_push_button.setText( "新增入金/出金資料" )
        uiqt_add_cash_transfer_push_button.setObjectName( "AddCashTransferPushButton" )
        uiqt_add_cash_transfer_push_button.clicked.connect( self.on_add_cash_transfer_push_button_clicked )
        uiqt_horizontal_layout_5.addWidget( uiqt_add_cash_transfer_push_button )

        uiqt_horizontal_spacer_5_2 = QSpacerItem(40, 20, QSizePolicy.Policy.MinimumExpanding, QSizePolicy.Policy.Minimum)
        uiqt_horizontal_layout_5.addItem( uiqt_horizontal_spacer_5_2 )

        uiqt_vertical_spacer_2_1 = QSpacerItem(20, 5, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum)
        uiqt_cash_transfer_tab_vertical_layout.addItem( uiqt_vertical_spacer_2_1 )

        uiqt_cash_transfer_tab_vertical_layout.addLayout( uiqt_horizontal_layout_5 )
        uiqt_vertical_spacer_2_2 = QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding)
        uiqt_cash_transfer_tab_vertical_layout.addItem( uiqt_vertical_spacer_2_2 )

        cash_transfer_model = QStandardItemModel( 0, 0 )
        cash_transfer_model.setVerticalHeaderLabels( self.get_cash_transfer_header() )
        uiqt_cash_transfer_table_view.verticalHeader().setSectionResizeMode( QHeaderView.Fixed )
        uiqt_cash_transfer_table_view.horizontalHeader().setSectionResizeMode( QHeaderView.Fixed )
        uiqt_cash_transfer_table_view.setModel( cash_transfer_model )
        uiqt_cash_transfer_table_view.setItemDelegate( delegate )
        uiqt_cash_transfer_table_view.horizontalHeader().hide()
        uiqt_cash_transfer_table_view.clicked.connect( lambda index: self.on_cash_transfer_table_item_clicked( index, cash_transfer_model ) )


        if not str_tab_title:
            str_tab_title = "新帳戶"
        with QSignalBlocker( self.ui.qtTabWidget ):
            n_ori_count = self.ui.qtTabWidget.count()
            self.ui.qtTabWidget.insertTab( n_ori_count - 1, increased_tab, str_tab_title )
            self.ui.qtTabWidget.setCurrentIndex( n_ori_count - 1 )
        self.n_tab_index += 1
        return str_tab_name

    def on_tab_current_changed( self, index ): 
        n_tab_count = self.ui.qtTabWidget.count()
        if index == n_tab_count - 1:
            self.ui.qtTabWidget.setCurrentIndex( self.n_current_tab )
        else:
            self.n_current_tab = index
            self.pick_up_stock( None )
            self.refresh_stock_list_table()
            self.refresh_transfer_data_table()

            self.clear_per_stock_trading_table()
            self.update_button_enable_disable_status()

    def on_tab_widget_double_clicked( self, index ): 
        n_tab_count = self.ui.qtTabWidget.count()
        if index == n_tab_count - 1:
            str_tab_name = self.add_new_tab_and_table()
            self.dict_all_account_all_stock_trading_data[ str_tab_name ] = {}
            self.dict_all_account_ui_state[ str_tab_name ] = { "discount_checkbox": True, 
                                                               "discount_value": 0.6, 
                                                               "regular_buy_trading_price_type": TradingPriceType.PER_SHARE, 
                                                               "regular_buy_trading_fee_type": TradingFeeType.VARIABLE, 
                                                               "regular_buy_trading_fee_minimum": 1, 
                                                               "regular_buy_trading_fee_constant": 1 }
            self.dict_all_account_general_data[ str_tab_name ] = { "insurance_checkbox": False, 
                                                                   "minimum_common_trading_fee": 20, 
                                                                   "minimum_odd_trading_fee": 1, 
                                                                   "dividend_transfer_fee": {}, 
                                                                   "decimal_round_type": DecimalRoundType.ROUND_DOWN,
                                                                   "discount_time_type": DiscountTimeType.IMMEDIATE }
            self.dict_all_account_cash_transfer_data[ str_tab_name ] = []
        else:
            current_title = self.ui.qtTabWidget.tabText( index )
            dialog = EditTabTitleDialog( current_title, self )
            if dialog.exec() == QDialog.Accepted:
                new_title = dialog.get_new_title()
                self.ui.qtTabWidget.setTabText( index, new_title )
        self.auto_save_trading_data()

    def on_tab_widget_close( self, index ):

        str_tab_title = self.ui.qtTabWidget.tabText( index )
        tab_widget = self.ui.qtTabWidget.widget( index )
        result = share_ui.show_warning_message_box_with_ok_cancel_button( "警告", 
                                                                         f"確定要刪掉『{str_tab_title}』的所有資料嗎?\n建議先從「檔案」=>「匯出目前帳號資料」，匯出檔案做備份",
                                                                          self )
        if result:
            str_tab_widget_name = tab_widget.objectName()
            value = self.dict_all_account_all_stock_trading_data.pop( str_tab_widget_name, None )
            self.ui.qtTabWidget.removeTab( index )
            self.clear_per_stock_trading_table()
            self.update_button_enable_disable_status()
            self.auto_save_trading_data()

    def on_tab_moved( self, n_from, n_to ): 
        str_tab_name = self.ui.qtTabWidget.widget( self.ui.qtTabWidget.count() - 1 ).objectName()
        tab_bar = self.ui.qtTabWidget.tabBar()
        if str_tab_name != 'tab_add':
            tab_bar.moveTab( n_to, n_from )
        else:
            self.auto_save_trading_data()

    def on_stock_input_text_changed( self ): 
        qt_combo_box = self.ui.qtTabWidget.currentWidget().findChild( QComboBox, "StockSelectComboBox" )
        qt_line_edit = self.ui.qtTabWidget.currentWidget().findChild( QLineEdit, "StockInputLineEdit" )

        with QSignalBlocker( qt_combo_box ), QSignalBlocker( qt_line_edit ):
            qt_combo_box.clear()
            str_stock_input = qt_line_edit.text()
            str_stock_input = share_api.lowercase_english_uppercase( str_stock_input )
            if len( str_stock_input ) == 0:
                qt_combo_box.setVisible( False )
                return
            qt_combo_box.setVisible( True )

            for stock_number, list_stock_name_and_type in self.dict_all_company_number_to_name_and_type.items():
                str_stock_name = list_stock_name_and_type[ 0 ]
                str_stock_name_lowercase = share_api.lowercase_english_uppercase( str_stock_name )
                stock_number_lowercase = share_api.lowercase_english_uppercase( stock_number )
                if str_stock_input in stock_number_lowercase or str_stock_input in str_stock_name_lowercase:
                    qt_combo_box.addItem( f"{stock_number} {str_stock_name}" )
            # self.ui.qtStockSelectComboBox.showPopup() #showPopup的話，focus會被搶走

            qt_line_edit.setFocus()

    def on_stock_select_combo_box_current_index_changed( self, index ): 
        qt_combo_box = self.ui.qtTabWidget.currentWidget().findChild( QComboBox, "StockSelectComboBox" )
        qt_line_edit = self.ui.qtTabWidget.currentWidget().findChild( QLineEdit, "StockInputLineEdit" )

        str_stock_input = qt_combo_box.currentText()
        qt_line_edit.setText( str_stock_input )
        qt_combo_box.setVisible( False )
        qt_line_edit.setFocus()

    def on_add_stock_push_button_clicked( self ): 
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        qt_line_edit = self.ui.qtTabWidget.currentWidget().findChild( QLineEdit, "StockInputLineEdit" )
        display_type_combobox = self.ui.qtTabWidget.currentWidget().findChild( QComboBox, "DisplayTypeComboBox")

        str_stock_input = qt_line_edit.text()
        qt_line_edit.clear()
        str_first_four_chars = str_stock_input.split(" ")[0]
        if str_first_four_chars not in self.dict_all_company_number_to_name_and_type:
            b_find = False
            for str_stock_number, list_stock_name_and_type in self.dict_all_company_number_to_name_and_type.items():
                str_stock_number_lower = share_api.lowercase_english_uppercase( str_stock_number )
                str_stock_name = list_stock_name_and_type[ 0 ]
                str_stock_name_lowercase = share_api.lowercase_english_uppercase( str_stock_name )
                str_first_four_chars_lower = share_api.lowercase_english_uppercase( str_first_four_chars )
                if str_first_four_chars_lower == str_stock_name_lowercase or str_first_four_chars_lower == str_stock_number_lower:
                    str_first_four_chars = str_stock_number
                    b_find = True
                    break
            if not b_find:
                QMessageBox.warning( self, "警告", "輸入的股票代碼不存在", QMessageBox.Ok )
                return
        
        if str_first_four_chars not in dict_per_account_all_stock_trading_data:
            dict_trading_data = Utility.generate_trading_data( "0001-01-01",                     #交易日期
                                                               TradingType.TEMPLATE,             #交易種類
                                                               TradingPriceType.PER_SHARE,       #交易價格種類
                                                               0,                                #每股交易價格
                                                               0,                                #總交易價格
                                                               0,                                #交易股數
                                                               TradingFeeType.VARIABLE,          #手續費種類
                                                               1,                                #手續費折扣
                                                               0,                                #手續費最低金額
                                                               0,                                #手續費固定金額
                                                               DividendValueType.PER_SHARE,      #股利金額種類
                                                               0,                                #股票股利
                                                               0,                                #現金股利
                                                               -1,                               #自訂的補充保費
                                                               0,                                #每股減資金額
                                                               CapitalReductionType.CASH_RETURN, #減資種類
                                                               False )                           #是否為當沖交易
            dict_trading_data[ TradingData.USE_AUTO_DIVIDEND_DATA ] = AutoDividendType.AUTO
            dict_per_account_all_stock_trading_data[ str_first_four_chars ] = [ dict_trading_data ]
            sorted_list = self.process_single_trading_data( str_tab_widget_name, str_first_four_chars )
            display_type_combobox.setCurrentIndex( 0 )
            self.refresh_stock_list_table()
            self.auto_save_trading_data()
        else:
            if display_type_combobox.currentIndex() != 0:
                display_type_combobox.setCurrentIndex( 0 )
                self.refresh_stock_list_table()

        table_view = self.ui.qtTabWidget.currentWidget().findChild( QTableView, "StockListTableView" )
        if table_view:
            table_model = table_view.model()
            for row in range( table_model.rowCount() ):
                str_stock_number_and_name = table_model.item( row, 0 ).text()
                str_stock_number = str_stock_number_and_name.split(" ")[0]
                if str_stock_number == str_first_four_chars:
                    index = table_model.index( row, 0 )
                    table_view.scrollTo( index, QTableView.PositionAtTop )
                    self.on_stock_list_table_item_clicked( index, table_model )
                    break

    def on_account_setting_push_button_clicked( self ):
        n_current_index = self.ui.qtTabWidget.currentIndex()
        str_current_title = self.ui.qtTabWidget.tabText( n_current_index )
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        b_extra_insurance_fee = self.dict_all_account_general_data[ str_tab_widget_name ][ "insurance_checkbox"]
        n_minimum_common_trading_fee = self.dict_all_account_general_data[ str_tab_widget_name ][ "minimum_common_trading_fee" ]
        n_minimum_odd_trading_fee = self.dict_all_account_general_data[ str_tab_widget_name ][ "minimum_odd_trading_fee" ]
        dict_company_number_to_transfer_fee = self.dict_all_account_general_data[ str_tab_widget_name ][ "dividend_transfer_fee" ]
        e_decimal_round_type = self.dict_all_account_general_data[ str_tab_widget_name ][ "decimal_round_type" ]
        e_discount_time_type = self.dict_all_account_general_data[ str_tab_widget_name ][ "discount_time_type" ]

        dialog = AccountSettingEditDialog( str_current_title,
                                           b_extra_insurance_fee, 
                                           n_minimum_common_trading_fee, 
                                           n_minimum_odd_trading_fee, 
                                           e_decimal_round_type, 
                                           e_discount_time_type,
                                           self.dict_all_company_number_to_name_and_type,
                                           dict_company_number_to_transfer_fee,
                                           self )
        if dialog.exec():
            self.dict_all_account_general_data[ str_tab_widget_name ][ "insurance_checkbox" ] = dialog.b_extra_insurance_fee
            self.dict_all_account_general_data[ str_tab_widget_name ][ "minimum_common_trading_fee" ] = dialog.n_minimum_common_trading_fee
            self.dict_all_account_general_data[ str_tab_widget_name ][ "minimum_odd_trading_fee" ] = dialog.n_minimum_odd_trading_fee
            self.dict_all_account_general_data[ str_tab_widget_name ][ "decimal_round_type" ] = dialog.e_decimal_round_type
            self.dict_all_account_general_data[ str_tab_widget_name ][ "discount_time_type" ] = dialog.e_discount_time_type
            self.dict_all_account_general_data[ str_tab_widget_name ][ "dividend_transfer_fee" ] = dialog.dict_company_number_to_transfer_fee


            dict_per_company_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
            for key_account_name, value_dict_per_company_trading_data in self.dict_all_account_all_stock_trading_data.items():
                if key_account_name == str_tab_widget_name:
                    for key_stock_name, value_list_trading_data in value_dict_per_company_trading_data.items():
                        self.process_single_trading_data( key_account_name, key_stock_name )

            self.refresh_stock_list_table()
            if self.str_picked_stock_number != None:
                self.refresh_trading_data_table( dict_per_company_trading_data[ self.str_picked_stock_number ] )
            self.auto_save_trading_data()

    def on_display_type_combo_box_current_index_changed( self, index ):
        self.pick_up_stock( None )
        self.refresh_stock_list_table()
        self.clear_per_stock_trading_table()
        self.update_button_enable_disable_status()

    def on_stock_list_table_vertical_header_section_moved( self, n_logical_index, n_old_visual_index, n_new_visual_index ): 
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()

        #取得所有股票代碼
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ].copy()
        list_stock_number_all = []
        for index_row,( key_stock_number, value ) in enumerate( dict_per_account_all_stock_trading_data.items() ):
            list_stock_number_all.append( key_stock_number )
        
        table_view = self.ui.qtTabWidget.currentWidget().findChild( QTableView, "StockListTableView" )
        list_stock_number_visible = []
        list_visible_stock_index_of_all = []
        if table_view:
            table_model = table_view.model()

            for row in range( table_model.rowCount() ):
                str_stock_number_and_name = table_model.item( row, 0 ).text()
                str_stock_number = str_stock_number_and_name.split(" ")[0]
                if str_stock_number in list_stock_number_all:
                    list_stock_number_visible.append( str_stock_number )
                    list_visible_stock_index_of_all.append( list_stock_number_all.index( str_stock_number ) )

        element = list_stock_number_visible.pop( n_old_visual_index )
        list_stock_number_visible.insert( n_new_visual_index, element )

        for idx, val in zip( list_visible_stock_index_of_all, list_stock_number_visible ):
            list_stock_number_all[ idx ] = val

        dict_all_stock_trading_data_new = {}
        for index_row, str_stock_number in enumerate( list_stock_number_all ):
            dict_all_stock_trading_data_new[ str_stock_number ] = dict_per_account_all_stock_trading_data[ str_stock_number ]

        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ] = dict_all_stock_trading_data_new
        self.refresh_stock_list_table()
        self.auto_save_trading_data()

    def on_stock_list_table_vertical_section_clicked( self, n_logical_index ): 
        table_view = self.ui.qtTabWidget.currentWidget().findChild( QTableView, "StockListTableView" )
        if table_view:
            table_model = table_view.model()

            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
            dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
            str_stock_number_and_name = table_model.item( n_logical_index, 0 ).text()
            str_stock_number = str_stock_number_and_name.split(" ")[0]
            if str_stock_number in dict_per_account_all_stock_trading_data:
                if str_stock_number != self.str_picked_stock_number:
                    self.pick_up_stock( str_stock_number )
                    list_trading_data = dict_per_account_all_stock_trading_data[ str_stock_number ]
                    self.refresh_trading_data_table( list_trading_data )
            self.update_button_enable_disable_status()

    def on_stock_list_table_horizontal_section_resized( self, n_logical_index, n_old_size, n_new_size ): 
        if ( n_logical_index == 0 or 
             n_logical_index == len( self.list_stock_list_table_horizontal_header ) - 3 or
             n_logical_index == len( self.list_stock_list_table_horizontal_header ) - 2 or
             n_logical_index == len( self.list_stock_list_table_horizontal_header ) - 1 ):
            self.list_stock_list_column_width[ n_logical_index ] = n_new_size
        else:
            e_modify_column_type = self.list_show_stock_info[ n_logical_index - 1 ]
            self.list_stock_list_column_width[ int( e_modify_column_type.value ) ] = n_new_size

        self.save_share_UI_state()

    def on_stock_list_table_item_clicked( self, index: QModelIndex, table_model ):
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        item = table_model.itemFromIndex( index )
        if item is not None:
            n_column = index.column()  # 獲取列索引
            n_row = index.row()  # 獲取行索引
            str_stock_number_and_name = table_model.item( index.row() , 0 ).text()
            str_stock_number = str_stock_number_and_name.split(" ")[0]
            if n_column == len( self.list_stock_list_table_horizontal_header ) - 3:#自動帶入股利按鈕
                list_trading_data = dict_per_account_all_stock_trading_data[ str_stock_number ]
                if list_trading_data[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ] == AutoDividendType.AUTO:
                    list_trading_data[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ] = AutoDividendType.MANUAL
                elif list_trading_data[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ] == AutoDividendType.MANUAL:
                    list_trading_data[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ] = AutoDividendType.MANUAL_WITH_AUTO
                elif list_trading_data[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ] == AutoDividendType.MANUAL_WITH_AUTO:
                    list_trading_data[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ] = AutoDividendType.AUTO
                self.pick_up_stock( str_stock_number )
                sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
                self.refresh_stock_list_table( False )
                self.refresh_trading_data_table( sorted_list )
                self.auto_save_trading_data()
            elif n_column == len( self.list_stock_list_table_horizontal_header ) - 2:#匯出按鈕
                file_path = share_ui.open_save_file_dialog( self, "匯出交易資料", "", share_ui.FileFilter.JSON )
                if file_path:
                    list_save_tab_widget = [ self.ui.qtTabWidget.currentIndex() ]
                    self.manual_save_trading_data( list_save_tab_widget, file_path, str_stock_number )
                
                if str_stock_number != self.str_picked_stock_number:
                    self.pick_up_stock( str_stock_number )
                    list_trading_data = dict_per_account_all_stock_trading_data[ str_stock_number ]
                    self.refresh_trading_data_table( list_trading_data )
            elif n_column == len( self.list_stock_list_table_horizontal_header ) - 1:#刪除按鈕
                result = share_ui.show_warning_message_box_with_ok_cancel_button( "警告", f"確定要刪掉『{str_stock_number}』的所有資料嗎?", self )
                if result:
                    del dict_per_account_all_stock_trading_data[ str_stock_number ]
                    self.pick_up_stock( None )
                    self.refresh_stock_list_table()
                    self.clear_per_stock_trading_table()
                    self.auto_save_trading_data()
            elif str_stock_number in dict_per_account_all_stock_trading_data:
                if str_stock_number != self.str_picked_stock_number:
                    self.pick_up_stock( str_stock_number )
                    list_trading_data = dict_per_account_all_stock_trading_data[ str_stock_number ]
                    self.refresh_trading_data_table( list_trading_data )

        self.update_button_enable_disable_status()

    def update_stock_list_vertical_header( self ):
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ].copy()
        list_stock_number_all = []
        for index_row,( key_stock_number, value ) in enumerate( dict_per_account_all_stock_trading_data.items() ):
            list_stock_number_all.append( key_stock_number )

        table_view = self.ui.qtTabWidget.currentWidget().findChild( QTableView, "StockListTableView" )
        # list_stock_number_visible = []
        # list_visible_stock_index_of_all = []
        if table_view:
            table_model = table_view.model()
            dict_stock_name = {}
            for row in range( table_model.rowCount() ):
                header_text = table_model.verticalHeaderItem( row ).text()
                str_stock_number = header_text.split(" ")[0]
                dict_stock_name[ str_stock_number ] = header_text
                # if str_stock_number in list_stock_number_all:
                #     list_stock_number_visible.append( str_stock_number )


            for row in range( table_model.rowCount() ):
                index = table_model.index( row, 0 )
                hidden_data = table_model.data( index, Qt.UserRole )
                str_stock_number_in_hidden_data = str( hidden_data )
                table_model.setHeaderData( row, Qt.Vertical, dict_stock_name[ str_stock_number_in_hidden_data ] )
            #     if str_stock_number_in_hidden_data in list_stock_number_all:
            #         list_visible_stock_index_of_all.append( list_stock_number_all.index( str_stock_number_in_hidden_data ) )

            # for idx, val in zip( list_visible_stock_index_of_all, list_stock_number_visible ):
            #     list_stock_number_all[ idx ] = val


            # dict_all_stock_trading_data_new = {}
            # for index_row, str_stock_number in enumerate( list_stock_number_all ):
            #     dict_all_stock_trading_data_new[ str_stock_number ] = dict_per_account_all_stock_trading_data[ str_stock_number ]

            # self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ] = dict_all_stock_trading_data_new
            # self.auto_save_trading_data()

    def on_add_cash_transfer_push_button_clicked( self ):
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        if str_tab_widget_name == 'tab_add':
            return

        list_per_account_all_cash_trasfer_data = self.dict_all_account_cash_transfer_data[ str_tab_widget_name ]

        n_current_index = self.ui.qtTabWidget.currentIndex()
        current_title = self.ui.qtTabWidget.tabText( n_current_index )

        dialog = CashTransferEditDialog( current_title, self )

        if dialog.exec():
            dict_transfer_data = dialog.dict_cash_transfer_data
            list_per_account_all_cash_trasfer_data.append( dict_transfer_data )
            self.process_single_transfer_data( str_tab_widget_name )
            self.refresh_transfer_data_table()
            self.auto_save_trading_data()

    def on_cash_transfer_table_item_clicked( self, index: QModelIndex, table_model ):
        item = table_model.itemFromIndex( index )
        if item is not None:
            n_row = index.row()  # 獲取行索引
            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
            list_per_account_cash_transfer_data = self.dict_all_account_cash_transfer_data[ str_tab_widget_name ]

            hidden_data = table_model.data( index, Qt.UserRole )
            n_findindex = -1
            for index, dict_selected_data in enumerate( list_per_account_cash_transfer_data ):
                if dict_selected_data[ TransferData.SORTED_INDEX_NON_SAVE ] == hidden_data:
                    n_findindex = index
                    break
            if n_findindex == -1:
                return
            dict_selected_data = list_per_account_cash_transfer_data[ n_findindex ]

            if n_row == len( self.get_cash_transfer_header() ) - 2: #編輯

                n_current_index = self.ui.qtTabWidget.currentIndex()
                current_title = self.ui.qtTabWidget.tabText( n_current_index )
                dialog = CashTransferEditDialog( current_title, self )
                dialog.setup_transfer_date( dict_selected_data[ TransferData.TRANSFER_DATE ] )
                dialog.setup_transfer_type( dict_selected_data[ TransferData.TRANSFER_TYPE ] )
                dialog.setup_transfer_value( dict_selected_data[ TransferData.TRANSFER_VALUE ] )

                if dialog.exec():
                    dict_transfer_data = dialog.dict_cash_transfer_data
                    list_per_account_cash_transfer_data[ n_findindex ] = dict_transfer_data
                    self.process_single_transfer_data( str_tab_widget_name )
                    self.refresh_transfer_data_table()
                    self.auto_save_trading_data()

            elif n_row == len( self.get_cash_transfer_header() ) - 1: #刪除
                result = share_ui.show_warning_message_box_with_ok_cancel_button( "警告", f"確定要刪掉這筆匯款資料嗎?", self )
                if result:
                    del list_per_account_cash_transfer_data[ n_findindex ]
                    self.process_single_transfer_data( str_tab_widget_name )
                    self.refresh_transfer_data_table()
                    self.auto_save_trading_data()

    def on_trigger_from_new_to_old( self ):
        with ( QSignalBlocker( self.ui.qtFromNewToOldAction ),
               QSignalBlocker( self.ui.qtFromOldToNewAction ) ):
                self.ui.qtFromNewToOldAction.setChecked( True )
                self.ui.qtFromOldToNewAction.setChecked( False )
                self.on_change_display_mode()

    def on_trigger_from_old_to_new( self ):
        with ( QSignalBlocker( self.ui.qtFromNewToOldAction ),
               QSignalBlocker( self.ui.qtFromOldToNewAction ) ):
                self.ui.qtFromNewToOldAction.setChecked( False )
                self.ui.qtFromOldToNewAction.setChecked( True )
                self.on_change_display_mode()

    def on_trigger_show_all( self ):
        with ( QSignalBlocker( self.ui.qtShowAllAction ),
               QSignalBlocker( self.ui.qtShow10Action ) ):
                self.ui.qtShowAllAction.setChecked( True )
                self.ui.qtShow10Action.setChecked( False )
                self.on_change_display_mode()

    def on_trigger_show_10( self ):
        with ( QSignalBlocker( self.ui.qtShowAllAction ),
               QSignalBlocker( self.ui.qtShow10Action ) ):
                self.ui.qtShowAllAction.setChecked( False )
                self.ui.qtShow10Action.setChecked( True )
                self.on_change_display_mode()

    def on_trigger_use_1_share_unit( self ):
        with ( QSignalBlocker( self.ui.qtUse1ShareUnitAction ),
               QSignalBlocker( self.ui.qtUse1000ShareUnitAction ) ):
                self.ui.qtUse1ShareUnitAction.setChecked( True )
                self.ui.qtUse1000ShareUnitAction.setChecked( False )
                self.on_change_display_mode()
                self.refresh_stock_list_table()

    def on_trigger_use_1000_share_unit( self ):
        with ( QSignalBlocker( self.ui.qtUse1ShareUnitAction ),
               QSignalBlocker( self.ui.qtUse1000ShareUnitAction ) ):
                self.ui.qtUse1ShareUnitAction.setChecked( False )
                self.ui.qtUse1000ShareUnitAction.setChecked( True )
                self.on_change_display_mode()
                self.refresh_stock_list_table()

    def on_trigger_AD_year( self ):
        with ( QSignalBlocker( self.ui.qtADYearAction ),
               QSignalBlocker( self.ui.qtROCYearAction ) ):
                self.ui.qtADYearAction.setChecked( True )
                self.ui.qtROCYearAction.setChecked( False )
                self.on_change_display_mode()

    def on_trigger_ROC_year( self ):
        with ( QSignalBlocker( self.ui.qtADYearAction ),
               QSignalBlocker( self.ui.qtROCYearAction ) ):
                self.ui.qtADYearAction.setChecked( False )
                self.ui.qtROCYearAction.setChecked( True )
                self.on_change_display_mode()

    def on_trigger_edit_show_item( self ):
        dialog = ShowItemEditDialog( self.list_show_stock_info, self.list_total_stock_info_INITIAL, self.list_total_stock_info_INITIAL, self )

        if dialog.exec():
            self.list_show_stock_info = dialog.list_final_show_stock_info
            self.save_share_UI_state()
            self.list_stock_list_table_horizontal_header = self.get_stock_list_horizontal_header()
            self.refresh_stock_list_table()

    def on_trigger_cost_with_in_dividend( self ):
        with ( QSignalBlocker( self.ui.qtCostWithInDividendAction ),
               QSignalBlocker( self.ui.qtCostWithOutDividendAction ) ):
                self.ui.qtCostWithInDividendAction.setChecked( True )
                self.ui.qtCostWithOutDividendAction.setChecked( False )
                self.list_stock_list_table_horizontal_header = self.get_stock_list_horizontal_header()
                self.refresh_stock_list_table()
                self.on_change_display_mode()

    def on_trigger_cost_with_out_dividend( self ):
        with ( QSignalBlocker( self.ui.qtCostWithInDividendAction ),
               QSignalBlocker( self.ui.qtCostWithOutDividendAction ) ):
                self.ui.qtCostWithInDividendAction.setChecked( False )
                self.ui.qtCostWithOutDividendAction.setChecked( True )
                self.list_stock_list_table_horizontal_header = self.get_stock_list_horizontal_header()
                self.refresh_stock_list_table()
                self.on_change_display_mode()

    def on_trigger_about( self ):
        dialog = AboutDialog( self )
        if dialog.exec():
            pass

    def on_change_display_mode( self ): 
        if self.str_picked_stock_number != None:
            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
            dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
            self.refresh_trading_data_table( dict_per_account_all_stock_trading_data[ self.str_picked_stock_number ] )

        self.refresh_transfer_data_table()
        self.save_share_UI_state()

    def on_hide_trading_data_table_tool_button_clicked( self ):
        if self.ui.qtTradingDataTableView.isVisible():
            self.ui.qtHideTradingDataTableToolButton.setIcon( share_icon.get_icon( share_icon.IconType.MOVE_UP ) )
            self.ui.qtTradingDataTableView.setHidden( True )
        else:
            self.ui.qtHideTradingDataTableToolButton.setIcon( share_icon.get_icon( share_icon.IconType.MOVE_DOWN ) )
            self.ui.qtTradingDataTableView.setHidden( False )

    def on_add_trading_data_push_button_clicked( self ): 
        if self.str_picked_stock_number is None:
            return

        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        b_discount = self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"]
        f_discount_value = self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"]
        n_current_minimum_common_trading_fee = self.dict_all_account_general_data[ str_tab_widget_name ][ "minimum_common_trading_fee" ]
        n_current_minimum_odd_trading_fee = self.dict_all_account_general_data[ str_tab_widget_name ][ "minimum_odd_trading_fee" ]

        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        str_b_etf = self.dict_all_company_number_to_name_and_type[ str_stock_number ][ 1 ]
        b_etf = True if str_b_etf == "True" else False
        e_decimal_round_type = self.dict_all_account_general_data[ str_tab_widget_name ][ "decimal_round_type" ]
        dialog = StockTradingEditDialog( str_stock_number, str_stock_name, e_decimal_round_type, b_etf, b_discount, f_discount_value, n_current_minimum_common_trading_fee, n_current_minimum_odd_trading_fee, self )

        if dialog.exec():
            dict_trading_data = dialog.dict_trading_data
            if dict_trading_data[ TradingData.TRADING_FEE_DISCOUNT ] == 1:
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"] = False
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"] = 0.6
            else:
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"] = True
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"] = dict_trading_data[ TradingData.TRADING_FEE_DISCOUNT ]


            dict_per_account_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list, dict_trading_data )
            self.auto_save_trading_data()

    def on_add_regular_trading_data_push_button_clicked( self ):
        if self.str_picked_stock_number is None:
            return

        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        b_discount = self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"]
        f_discount_value = self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"]
        e_trading_price_type = self.dict_all_account_ui_state[ str_tab_widget_name ][ "regular_buy_trading_price_type"]
        e_trading_fee_type = self.dict_all_account_ui_state[ str_tab_widget_name ][ "regular_buy_trading_fee_type"]
        n_trading_fee_minimum = self.dict_all_account_ui_state[ str_tab_widget_name ][ "regular_buy_trading_fee_minimum"]
        n_trading_fee_constant = self.dict_all_account_ui_state[ str_tab_widget_name ][ "regular_buy_trading_fee_constant"]

        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        e_decimal_round_type = self.dict_all_account_general_data[ str_tab_widget_name ][ "decimal_round_type" ]
        dialog = StockRegularTradingEditDialog( str_stock_number, str_stock_name, e_decimal_round_type, e_trading_price_type, e_trading_fee_type, b_discount, f_discount_value, n_trading_fee_minimum, n_trading_fee_constant, self )

        if dialog.exec():
            dict_trading_data = dialog.dict_trading_data
            if dict_trading_data[ TradingData.TRADING_FEE_DISCOUNT ] == 1:
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"] = False
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"] = 0.6
            else:
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"] = True
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"] = dict_trading_data[ TradingData.TRADING_FEE_DISCOUNT ]
            self.dict_all_account_ui_state[ str_tab_widget_name ][ "regular_buy_trading_price_type"] = dict_trading_data[ TradingData.TRADING_PRICE_TYPE ]
            self.dict_all_account_ui_state[ str_tab_widget_name ][ "regular_buy_trading_fee_type"] = dict_trading_data[ TradingData.REGULAR_BUY_TRADING_FEE_TYPE ]
            self.dict_all_account_ui_state[ str_tab_widget_name ][ "regular_buy_trading_fee_minimum"] = dict_trading_data[ TradingData.REGULAR_BUY_TRADING_FEE_MINIMUM ]
            self.dict_all_account_ui_state[ str_tab_widget_name ][ "regular_buy_trading_fee_constant"] = dict_trading_data[ TradingData.REGULAR_BUY_TRADING_FEE_CONSTANT ]


            dict_per_account_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list, dict_trading_data )
            self.auto_save_trading_data()

    def on_add_dividend_data_push_button_clicked( self ): 
        if self.str_picked_stock_number is None:
            return

        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]

        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        dialog = StockDividendEditDialog( str_stock_number, str_stock_name, self )

        if dialog.exec():
            dict_trading_data = dialog.dict_trading_data
            dict_per_account_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list, dict_trading_data )
            self.auto_save_trading_data()

    def on_add_limit_buying_data_push_button_clicked( self ): 
        if self.str_picked_stock_number is None:
            return

        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]

        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        dialog = StockCapitalIncreaseEditDialog( str_stock_number, str_stock_name, self )

        if dialog.exec():
            dict_trading_data = dialog.dict_trading_data

            dict_per_account_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list, dict_trading_data )
            self.auto_save_trading_data()

    def on_add_capital_reduction_data_push_button_clicked( self ): 
        if self.str_picked_stock_number is None:
            return
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]

        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        dialog = StockCapitalReductionEditDialog( str_stock_number, str_stock_name, self )

        if dialog.exec():
            dict_trading_data = dialog.dict_trading_data
            dict_per_account_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list, dict_trading_data )
            self.auto_save_trading_data()

    def on_add_stock_split_data_push_button_clicked( self ): 
        if self.str_picked_stock_number is None:
            return
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]

        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        dialog = StockSplitEditDialog( str_stock_number, str_stock_name, self )

        if dialog.exec():
            dict_trading_data = dialog.dict_trading_data
            dict_per_account_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list, dict_trading_data )
            self.auto_save_trading_data()

    def on_trading_data_table_item_clicked( self, index: QModelIndex, table_model ): 
        item = table_model.itemFromIndex( index )
        if item is not None:
            n_row = index.row()  # 獲取行索引

            if ( n_row == len( self.get_trading_data_header() ) - 2 or #編輯
                n_row == len( self.get_trading_data_header() ) - 1 ): #刪除

                str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
                dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
                list_trading_data = dict_per_account_all_stock_trading_data[ self.str_picked_stock_number ]

                if self.str_picked_stock_number is None:
                    return

                hidden_data = table_model.data( index, Qt.UserRole )
                n_findindex = -1
                for index, dict_selected_data in enumerate( list_trading_data ):
                    if dict_selected_data[ TradingData.SORTED_INDEX_NON_SAVE ] == hidden_data:
                        n_findindex = index
                        break
                if n_findindex == -1:
                    return
                dict_selected_data = list_trading_data[ n_findindex ]

                str_stock_number = self.str_picked_stock_number
                list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
                str_stock_name = list_stock_name_and_type[ 0 ]
                e_decimal_round_type = self.dict_all_account_general_data[ str_tab_widget_name ][ "decimal_round_type" ]
                if n_row == len( self.get_trading_data_header() ) - 2: #編輯
                    if dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.TEMPLATE:
                        return
                    if dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.BUY or dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.SELL:
                        str_b_etf = self.dict_all_company_number_to_name_and_type[ str_stock_number ][ 1 ]
                        b_etf = True if str_b_etf == "True" else False
                        n_current_minimum_common_trading_fee = self.dict_all_account_general_data[ str_tab_widget_name ][ "minimum_common_trading_fee" ]
                        n_current_minimum_odd_trading_fee = self.dict_all_account_general_data[ str_tab_widget_name ][ "minimum_odd_trading_fee" ]
                        
                        dialog = StockTradingEditDialog( str_stock_number, str_stock_name, e_decimal_round_type, b_etf, True, 0, n_current_minimum_common_trading_fee, n_current_minimum_odd_trading_fee, self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        dialog.setup_trading_type( dict_selected_data[ TradingData.TRADING_TYPE ] )
                        dialog.setup_trading_discount( dict_selected_data[ TradingData.TRADING_FEE_DISCOUNT ] )
                        dialog.setup_trading_price( dict_selected_data[ TradingData.PER_SHARE_TRADING_PRICE ] )
                        dialog.setup_trading_count( dict_selected_data[ TradingData.TRADING_QUANTITY ] )
                        dialog.setup_daying_trading( dict_selected_data[ TradingData.DAYING_TRADING ] )
                        dialog.compute_cost()
                    elif dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.REGULAR_BUY:
                        dialog = StockRegularTradingEditDialog( str_stock_number, str_stock_name, e_decimal_round_type, TradingPriceType.PER_SHARE, TradingFeeType.VARIABLE, True, 0, 0, 0, self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        dialog.setup_trading_count( dict_selected_data[ TradingData.TRADING_QUANTITY ] )
                        dialog.setup_trading_price_type( dict_selected_data[ TradingData.TRADING_PRICE_TYPE ] )
                        dialog.setup_per_share_trading_price( dict_selected_data[ TradingData.PER_SHARE_TRADING_PRICE ] )
                        dialog.setup_total_trading_price( dict_selected_data[ TradingData.TOTAL_TRADING_PRICE ] )
                        dialog.setup_trading_fee_type( dict_selected_data[ TradingData.REGULAR_BUY_TRADING_FEE_TYPE ] )
                        dialog.setup_trading_discount( dict_selected_data[ TradingData.TRADING_FEE_DISCOUNT ] )
                        dialog.setup_trading_fee_minimum( dict_selected_data[ TradingData.REGULAR_BUY_TRADING_FEE_MINIMUM ] )
                        dialog.setup_trading_fee_constant( dict_selected_data[ TradingData.REGULAR_BUY_TRADING_FEE_CONSTANT ] )
                        dialog.compute_cost()
                    elif dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.DIVIDEND:
                        dialog = StockDividendEditDialog( str_stock_number, str_stock_name, self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        e_dividend_value_type = dict_selected_data[ TradingData.DIVIDEND_VALUE_TYPE ]
                        dialog.setup_dividend_value_type( e_dividend_value_type )
                        if e_dividend_value_type == DividendValueType.PER_SHARE:
                            dialog.setup_per_share_stock_dividend( dict_selected_data[ TradingData.STOCK_DIVIDEND ] )
                            dialog.setup_per_share_cash_dividend( dict_selected_data[ TradingData.CASH_DIVIDEND ] )
                        else:
                            dialog.setup_total_stock_dividend( dict_selected_data[ TradingData.STOCK_DIVIDEND ] )
                            dialog.setup_total_cash_dividend( dict_selected_data[ TradingData.CASH_DIVIDEND ] )

                        dialog.setup_custom_extra_insurance_fee( dict_selected_data[ TradingData.CUSTOM_EXTRA_INSURANCE_FEE ] )
                    elif dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.CAPITAL_INCREASE:
                        dialog = StockCapitalIncreaseEditDialog( str_stock_number, str_stock_name, self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        dialog.setup_trading_count( dict_selected_data[ TradingData.TRADING_QUANTITY ] )
                        dialog.setup_trading_price_type( dict_selected_data[ TradingData.TRADING_PRICE_TYPE ] )
                        dialog.setup_per_share_trading_price( dict_selected_data[ TradingData.PER_SHARE_TRADING_PRICE ] )
                        dialog.setup_total_trading_price( dict_selected_data[ TradingData.TOTAL_TRADING_PRICE ] )
                        dialog.compute_cost()
                    elif dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.CAPITAL_REDUCTION:
                        dialog = StockCapitalReductionEditDialog( str_stock_number, str_stock_name, self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        dialog.setup_stock_capital_reduction_value( dict_selected_data[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] )
                        dialog.setup_stock_capital_reduction_type( dict_selected_data[ TradingData.CAPITAL_REDUCTION_TYPE ] )
                    elif dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.SPLIT or dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.MERGE:
                        dialog = StockSplitEditDialog( str_stock_number, str_stock_name, self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        dialog.setup_trading_type_and_value( dict_selected_data[ TradingData.TRADING_TYPE ], 
                                                             dict_selected_data[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] )

                    if dialog.exec():
                        dict_trading_data = dialog.dict_trading_data
                        dict_per_account_all_stock_trading_data[ str_stock_number ][ n_findindex ] = dict_trading_data
                        sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
                        self.refresh_stock_list_table()
                        self.refresh_trading_data_table( sorted_list, dict_trading_data )
                        self.auto_save_trading_data()

                elif n_row == len( self.get_trading_data_header() ) - 1: #刪除
                    result = share_ui.show_warning_message_box_with_ok_cancel_button( "警告", f"確定要刪掉這筆交易資料嗎?", self )
                    if result:
                        del dict_per_account_all_stock_trading_data[ str_stock_number ][ n_findindex ]
                        sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
                        self.refresh_stock_list_table()
                        self.refresh_trading_data_table( sorted_list )
                        self.auto_save_trading_data()

    def on_trading_data_table_horizontal_header_section_moved( self, logicalIndex, n_old_visual_index, n_new_visual_index ):
        center_point = self.ui.qtTradingDataTableView.viewport().rect().center()
        center_index = self.ui.qtTradingDataTableView.indexAt( center_point )

        n_target_visual_index = n_new_visual_index

        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        list_trading_data = dict_per_account_all_stock_trading_data[ self.str_picked_stock_number ]

        n_old_table_model_index = self.per_stock_trading_data_model.item( 0, n_old_visual_index ).data( Qt.UserRole )
        n_old_data_index = -1
        moved_trading_data = None
        for index, dict_selected_data in enumerate( list_trading_data ):
            if dict_selected_data[ TradingData.SORTED_INDEX_NON_SAVE ] == n_old_table_model_index:
                n_old_data_index = index
                moved_trading_data = dict_selected_data
                break
        if n_old_data_index == -1:
            header = self.ui.qtTradingDataTableView.horizontalHeader()
            with QSignalBlocker( header ):
                header.moveSection( n_new_visual_index, n_old_visual_index )
            return

        per_trading_data_old = list_trading_data[ n_old_data_index ]
        e_trading_type_old = per_trading_data_old[ TradingData.TRADING_TYPE ]
        str_trading_date_old = per_trading_data_old[ TradingData.TRADING_DATE ]

        step = -1 if n_target_visual_index > n_old_visual_index else 1
        while n_target_visual_index != n_old_visual_index:
            n_new_table_model_index = self.per_stock_trading_data_model.item( 0, n_target_visual_index ).data( Qt.UserRole )
            n_new_data_index = -1
            for index, dict_selected_data in enumerate( list_trading_data ):
                if dict_selected_data[ TradingData.SORTED_INDEX_NON_SAVE ] == n_new_table_model_index:
                    n_new_data_index = index
                    break
            if n_new_data_index == -1:
                header = self.ui.qtTradingDataTableView.horizontalHeader()
                with QSignalBlocker( header ):
                    header.moveSection( n_new_visual_index, n_old_visual_index )
                return
            
            per_trading_data_new = list_trading_data[ n_new_data_index ]
            e_trading_type_new = per_trading_data_new[ TradingData.TRADING_TYPE ]
            str_trading_date_new = per_trading_data_new[ TradingData.TRADING_DATE ]
            if e_trading_type_old == e_trading_type_new and str_trading_date_old == str_trading_date_new:
                break
            n_target_visual_index += step

        if n_target_visual_index != n_old_visual_index:
            element = list_trading_data.pop( n_old_data_index )
            list_trading_data.insert( n_new_data_index, element )

            dict_per_account_all_stock_trading_data[ self.str_picked_stock_number ] = list_trading_data

            sorted_list = self.process_single_trading_data( str_tab_widget_name, self.str_picked_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list, moved_trading_data )
            self.auto_save_trading_data()
        else:
            header = self.ui.qtTradingDataTableView.horizontalHeader()
            with QSignalBlocker( header ):
                header.moveSection( n_new_visual_index, n_old_visual_index )

    def export_trading_data_to_excel( self, worksheet, str_stock_number, str_stock_name, list_trading_data ): 
        all_thin_border = Border(
        left = Side( style = 'thin', color = "000000" ),  # 左邊框：細線，黑色
        right = Side( style = 'thin', color = "000000" ), # 右邊框：細線，黑色
        top = Side( style = 'thin', color = "000000" ),   # 上邊框：細線，黑色
        bottom = Side( style = 'thin', color = "000000" ) # 下邊框：細線，黑色
        )

        top_thick_border = Border(
        left = Side( style = 'thin', color = "000000" ),  # 左邊框：細線，黑色
        right = Side( style = 'thin', color = "000000" ), # 右邊框：細線，黑色
        top = Side( style = 'thick', color = "000000" ),   # 上邊框：細線，黑色
        bottom = Side( style = 'thin', color = "000000" ) # 下邊框：細線，黑色
        )

        bottom_thick_border = Border(
        left = Side( style = 'thin', color = "000000" ),  # 左邊框：細線，黑色
        right = Side( style = 'thin', color = "000000" ), # 右邊框：細線，黑色
        top = Side( style = 'thin', color = "000000" ),   # 上邊框：細線，黑色
        bottom = Side( style = 'thick', color = "000000" ) # 下邊框：細線，黑色
        )

        right_thick_border = Border(
        left = Side( style = 'thin', color = "000000" ),  # 左邊框：細線，黑色
        right = Side( style = 'thick', color = "000000" ), # 右邊框：細線，黑色
        top = Side( style = 'thin', color = "000000" ),   # 上邊框：細線，黑色
        bottom = Side( style = 'thin', color = "000000" ) # 下邊框：細線，黑色
        )

        str_title = str_stock_number + " " + str_stock_name
        worksheet.column_dimensions['A'].width = 22

        e_auto_dividend_type = list_trading_data[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ]
        data_index = 0
        n_row_start = 0
        for dict_per_trading_data in list_trading_data:
            
            e_trading_type = dict_per_trading_data[ TradingData.TRADING_TYPE ]
            if e_trading_type == TradingType.TEMPLATE:
                continue

            if e_trading_type == TradingType.DIVIDEND:
                if e_auto_dividend_type == AutoDividendType.AUTO:#使用自動帶入股利資料，但是這筆資料不是自動股利資料，就跳過，反之亦然
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE not in dict_per_trading_data or not dict_per_trading_data[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]:
                        continue
                elif e_auto_dividend_type == AutoDividendType.MANUAL:
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE in dict_per_trading_data and dict_per_trading_data[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]:
                        continue
            if data_index % 10 == 0:
                list_data_header = self.get_trading_data_header()
                list_data_header.insert( 0, str_title )

                n_row_start = int( ( len( list_data_header ) -2 + 2 ) * int( data_index / 10 ) )
                for index_row, str_header in enumerate( list_data_header ):
                    if index_row == len( list_data_header ) - 2:
                        break
                    worksheet.cell( row = n_row_start + index_row + 1, column = 1, value = str_header ).border = all_thin_border
                    if index_row == 0:
                        worksheet.cell( row = n_row_start + index_row + 1, column = 1 ).font = Font( bold = True )
                index_column = 0

            worksheet.column_dimensions[ get_column_letter( index_column + 2 ) ].width = 12

            list_data = self.get_per_trading_data_text_list( dict_per_trading_data )
            list_data.insert( 0, "# " + str( data_index + 1 ) )
            for index_row, str_data in enumerate( list_data ):
                str_data = str_data.replace( ',', '' )
                n_cell_row = n_row_start + index_row + 1
                n_cell_column = index_column + 2
                cell = worksheet.cell( row = n_cell_row, column = n_cell_column )
                if str_data == "買進":
                    color_fill = PatternFill( start_color = "DA9694", end_color = "DA9694", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "定期定額買進":
                    color_fill = PatternFill( start_color = "FF99FF", end_color = "FF99FF", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "賣出" or str_data == "當沖賣":
                    color_fill = PatternFill( start_color = "76933C", end_color = "76933C", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "股利分配":
                    color_fill = PatternFill( start_color = "8DB4E2", end_color = "8DB4E2", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "現金減資" or str_data == "虧損減資":
                    color_fill = PatternFill( start_color = "B1A0C7", end_color = "B1A0C7", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "增資":
                    color_fill = PatternFill( start_color = "FABF8F", end_color = "FABF8F", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "股票分割":
                    color_fill = PatternFill( start_color = "732BF5", end_color = "732BF5", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "股票合併":
                    color_fill = PatternFill( start_color = "EF88BE", end_color = "EF88BE", fill_type="solid")
                    cell.fill = color_fill

                str_cell = get_column_letter( n_cell_column ) + str( n_cell_row )
                if index_row == 1:
                    worksheet.cell( row = n_cell_row, column = n_cell_column, value = str_data ).border = all_thin_border
                else:
                    try:
                        f_data= float( str_data )
                        if f_data.is_integer():
                            f_data = int( f_data )
                            worksheet[ str_cell ].number_format = "#,##0"  #顯示千位逗號
                        elif ( f_data * 10 ).is_integer():
                            worksheet[ str_cell ].number_format = "#,##0.0"
                        elif ( f_data * 100 ).is_integer():
                            worksheet[ str_cell ].number_format = "#,##0.00"
                        elif ( f_data * 1000 ).is_integer():
                            worksheet[ str_cell ].number_format = "#,##0.000"
                        worksheet.cell( row = n_cell_row, column = n_cell_column, value = f_data ).border = all_thin_border
                    except ValueError:
                        worksheet.cell( row = n_cell_row, column = n_cell_column, value = str_data ).border = all_thin_border

                worksheet[ str_cell ].alignment = Alignment( horizontal = "center", vertical = "center", shrink_to_fit = True )
            data_index += 1
            index_column += 1

    def on_export_selected_to_excell_button_clicked( self ): 
        if self.str_picked_stock_number is None:
            return
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        list_trading_data = dict_per_account_all_stock_trading_data[ str_stock_number ]
        file_path = share_ui.open_save_file_dialog( self, "輸出Excel", "", share_ui.FileFilter.EXCEL )
        if file_path:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = str_stock_number + " " + re.sub( r'[^\w\u4e00-\u9fff]', '', str_stock_name ) 
            worksheet.page_setup.orientation = "landscape"
            worksheet.page_setup.paperSize = 9
            worksheet.page_margins.left = 0.7    # 左邊界
            worksheet.page_margins.right = 0.7   # 右邊界
            worksheet.page_margins.top = 0.75    # 上邊界
            worksheet.page_margins.bottom = 0.75 # 下邊界
            worksheet.page_margins.header = 0.3  # 頁眉邊界
            worksheet.page_margins.footer = 0.3  # 頁腳邊界
            self.export_trading_data_to_excel( worksheet, str_stock_number, str_stock_name, list_trading_data )
            workbook.save( file_path )
            share_ui.show_check_message_box_with_ok_button( "匯出成功", f"已經將 { str_stock_number } { str_stock_name } 的交易資料匯出到 { file_path }", self )

    def on_export_all_to_excell_button_clicked( self ): 
        file_path = share_ui.open_save_file_dialog( self, "輸出Excel", "", share_ui.FileFilter.EXCEL )
        if file_path:
            workbook = Workbook()
            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
            n_current_index = self.ui.qtTabWidget.currentIndex()
            str_account_name = self.ui.qtTabWidget.tabText( n_current_index )
            dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]

            for index, ( key_stock_number, value_list_trading_data ) in enumerate( dict_per_account_all_stock_trading_data.items() ):
                list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ key_stock_number ]
                str_stock_name = list_stock_name_and_type[ 0 ]
                str_tab_title = key_stock_number + " " + str_stock_name
                if index == 0:
                    worksheet = workbook.active
                    worksheet.title = str_tab_title
                    worksheet.page_setup.orientation = "landscape"
                    worksheet.page_setup.paperSize = 9
                    worksheet.page_margins.left = 0.7    # 左邊界
                    worksheet.page_margins.right = 0.7   # 右邊界
                    worksheet.page_margins.top = 0.75    # 上邊界
                    worksheet.page_margins.bottom = 0.75 # 下邊界
                    worksheet.page_margins.header = 0.3  # 頁眉邊界
                    worksheet.page_margins.footer = 0.3  # 頁腳邊界
                else:
                    worksheet = workbook.create_sheet( str_tab_title, index )
                    worksheet.page_setup.orientation = "landscape"
                    worksheet.page_setup.paperSize = 9
                    worksheet.page_margins.left = 0.7    # 左邊界
                    worksheet.page_margins.right = 0.7   # 右邊界
                    worksheet.page_margins.top = 0.75    # 上邊界
                    worksheet.page_margins.bottom = 0.75 # 下邊界
                    worksheet.page_margins.header = 0.3  # 頁眉邊界
                    worksheet.page_margins.footer = 0.3  # 頁腳邊界
                self.export_trading_data_to_excel( worksheet, key_stock_number, str_stock_name, value_list_trading_data )
            workbook.save( file_path )
            share_ui.show_check_message_box_with_ok_button( "匯出成功", f"已經將{ str_account_name } 的交易資料匯出到 { file_path }", self )

    def on_new_file_action_triggered( self ): 
        b_need_save = False
        if ( ( self.dict_all_account_all_stock_trading_data_INITIAL != self.dict_all_account_all_stock_trading_data ) and
             ( self.str_current_save_file_path is None or not share_api.compare_json_files( self.str_current_save_file_path, self.trading_data_json_file_path ) ) ):
            dialog = share_ui.SaveCheckDialog( share_icon.get_icon( share_icon.IconType.WINDOW ), '開新檔案', self )
            if dialog.exec():
                if dialog.n_return == 1: #儲存
                    b_need_save = True
                elif dialog.n_return == 2: #不儲存
                    pass
                else: #取消
                    return
            else:
                return
        if b_need_save:
            if self.str_current_save_file_path:
                list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
                self.manual_save_trading_data( list_save_tab_widget, self.str_current_save_file_path )
            else:
                file_path = share_ui.open_save_file_dialog( self, "匯出交易資料", "", share_ui.FileFilter.JSON )
                if file_path:
                    list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
                    self.manual_save_trading_data( list_save_tab_widget, file_path )
                    self.str_current_save_file_path = file_path
                else:
                    return

        with QSignalBlocker( self.ui.qtTabWidget ):
            for index in range( self.ui.qtTabWidget.count() - 2, -1, -1 ):
                self.ui.qtTabWidget.removeTab( index )

        self.dict_all_account_all_stock_trading_data.clear()
        self.dict_all_account_cash_transfer_data.clear()
        self.dict_all_account_ui_state.clear()
        self.dict_all_account_general_data.clear()
        str_tab_name = self.add_new_tab_and_table()
        self.dict_all_account_all_stock_trading_data[ str_tab_name ] = {}
        self.dict_all_account_cash_transfer_data[ str_tab_name ] = []
        self.dict_all_account_ui_state[ str_tab_name ] = { "discount_checkbox": True, 
                                                           "discount_value": 0.6, 
                                                           "regular_buy_trading_price_type": TradingPriceType.PER_SHARE, 
                                                           "regular_buy_trading_fee_type": TradingFeeType.VARIABLE, 
                                                           "regular_buy_trading_fee_minimum": 1, 
                                                           "regular_buy_trading_fee_constant": 1 }
        self.dict_all_account_general_data[ str_tab_name ] = { "insurance_checkbox": False, 
                                                               "minimum_common_trading_fee": 20, 
                                                               "minimum_odd_trading_fee": 1, 
                                                               "dividend_transfer_fee":{},
                                                               "decimal_round_type": DecimalRoundType.ROUND_DOWN,
                                                               "discount_time_type": DiscountTimeType.IMMEDIATE }
        self.dict_all_account_all_stock_trading_data_INITIAL = copy.deepcopy( self.dict_all_account_all_stock_trading_data )
        self.ui.qtTabWidget.setCurrentIndex( 0 )
        self.pick_up_stock( None )
        self.refresh_stock_list_table()
        self.refresh_transfer_data_table()
        self.clear_per_stock_trading_table()
        self.update_button_enable_disable_status()
        self.auto_save_trading_data()
        self.str_current_save_file_path = None

    def on_open_file_action_triggered( self ): 
        file_path = share_ui.open_load_file_dialog( self, "匯入交易資料", "", share_ui.FileFilter.JSON )
        if file_path:
            b_need_save = False
            if ( ( self.dict_all_account_all_stock_trading_data_INITIAL != self.dict_all_account_all_stock_trading_data ) and
                 ( self.str_current_save_file_path is None or not share_api.compare_json_files( self.str_current_save_file_path, self.trading_data_json_file_path ) ) ):
                dialog = share_ui.SaveCheckDialog( share_icon.get_icon( share_icon.IconType.WINDOW ), '開啟舊檔', self )
                if dialog.exec():
                    if dialog.n_return == 1: #儲存
                        b_need_save = True
                    elif dialog.n_return == 2: #不儲存
                        pass
                    else: #取消
                        return
                else:
                    return
            if b_need_save:
                if self.str_current_save_file_path:
                    list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
                    self.manual_save_trading_data( list_save_tab_widget, self.str_current_save_file_path )
                else:
                    file_path = share_ui.open_save_file_dialog( self, "匯出交易資料", "", share_ui.FileFilter.JSON )
                    if file_path:
                        list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
                        self.manual_save_trading_data( list_save_tab_widget, file_path )
                        self.str_current_save_file_path = file_path
                    else:
                        return

            with QSignalBlocker( self.ui.qtTabWidget ):
                for index in range( self.ui.qtTabWidget.count() - 2, -1, -1 ):
                    self.ui.qtTabWidget.removeTab( index )


                self.dict_all_account_all_stock_trading_data.clear()
                self.dict_all_account_cash_transfer_data.clear()
                self.dict_all_account_ui_state.clear()
                self.dict_all_account_general_data.clear()
                self.load_trading_data_and_create_tab( file_path, 
                                                       self.dict_all_account_all_stock_trading_data, 
                                                       self.dict_all_account_ui_state, 
                                                       self.dict_all_account_cash_transfer_data,
                                                       self.dict_all_account_general_data,
                                                       True )
                if len( self.dict_all_account_all_stock_trading_data ) == 0:
                    str_tab_name = self.add_new_tab_and_table()
                    self.dict_all_account_all_stock_trading_data[ str_tab_name ] = {}
                    self.dict_all_account_cash_transfer_data[ str_tab_name ] = []
                    self.dict_all_account_ui_state[ str_tab_name ] = { "discount_checkbox": True, 
                                                                       "discount_value": 0.6, 
                                                                       "regular_buy_trading_price_type": TradingPriceType.PER_SHARE, 
                                                                       "regular_buy_trading_fee_type": TradingFeeType.VARIABLE, 
                                                                       "regular_buy_trading_fee_minimum": 1, 
                                                                       "regular_buy_trading_fee_constant": 1 }
                    self.dict_all_account_general_data[ str_tab_name ] = { "insurance_checkbox": False, 
                                                                           "minimum_common_trading_fee": 20, 
                                                                           "minimum_odd_trading_fee": 1, 
                                                                           "dividend_transfer_fee":{},
                                                                           "decimal_round_type": DecimalRoundType.ROUND_DOWN,
                                                                           "discount_time_type": DiscountTimeType.IMMEDIATE }
                self.ui.qtTabWidget.setCurrentIndex( 0 )

            self.process_all_trading_data()
            self.pick_up_stock( None )
            self.refresh_stock_list_table()
            self.process_all_transfer_data()
            self.refresh_transfer_data_table()
            self.clear_per_stock_trading_table()
            self.update_button_enable_disable_status()
            self.auto_save_trading_data()
            self.str_current_save_file_path = file_path

    def on_save_as_action_triggered( self ): 
        file_path = share_ui.open_save_file_dialog( self, "匯出交易資料", "", share_ui.FileFilter.JSON )
        if file_path:
            list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
            self.manual_save_trading_data( list_save_tab_widget, file_path )
            self.str_current_save_file_path = file_path

    def on_save_action_triggered( self ): 
        file_path = self.str_current_save_file_path or share_ui.open_save_file_dialog( self, "匯出交易資料", "", share_ui.FileFilter.JSON )
        if file_path:
            list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
            self.manual_save_trading_data( list_save_tab_widget, file_path )
            self.str_current_save_file_path = file_path

    def on_export_current_group_action_triggered( self ): 
        file_path = share_ui.open_save_file_dialog( self, "匯出交易資料", "", share_ui.FileFilter.JSON )
        if file_path:
            list_save_tab_widget = [ self.ui.qtTabWidget.currentIndex() ]
            self.manual_save_trading_data( list_save_tab_widget, file_path )

    def on_import_full_action_triggered( self ):
        file_path = share_ui.open_load_file_dialog( self, "匯入交易資料", "", share_ui.FileFilter.JSON )
        if file_path:
            dict_account_to_tab_widget_name = {}
            for index in range( self.ui.qtTabWidget.count() - 1 ):
                tab_widget = self.ui.qtTabWidget.widget( index )
                str_tab_title = self.ui.qtTabWidget.tabText( index )
                str_tab_widget_name = tab_widget.objectName()
                dict_account_to_tab_widget_name[ str_tab_title ] = str_tab_widget_name

                # self.dict_all_account_all_stock_trading_data[ self.str_picked_stock_number ] 

            dict_all_account_all_stock_trading_data_LOAD = {}
            dict_all_account_ui_state_LOAD = {}
            dict_all_account_cash_transfer_data_LOAD = {}
            dict_all_account_general_data_LOAD = {}
            self.load_trading_data_and_create_tab( file_path, 
                                                   dict_all_account_all_stock_trading_data_LOAD, 
                                                   dict_all_account_ui_state_LOAD, 
                                                   dict_all_account_cash_transfer_data_LOAD, 
                                                   dict_all_account_general_data_LOAD,
                                                   False )
            b_duplicate = False
            for str_account_name, dict_per_account_all_stock_trading_data_LOAD in dict_all_account_all_stock_trading_data_LOAD.items():
                if str_account_name in dict_account_to_tab_widget_name:
                    str_tab_widget_name = dict_account_to_tab_widget_name[ str_account_name ]
                    for key_stock_number, value in dict_per_account_all_stock_trading_data_LOAD.items():
                        if key_stock_number in self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]:
                            b_duplicate = True
                            break
                    if b_duplicate:
                        break

            if b_duplicate:
                dialog = ImportDataDuplicateOptionDialog( self )
                if dialog.exec():
                    if dialog.b_overwrite:
                        for str_account_name, dict_per_account_all_stock_trading_data_LOAD in dict_all_account_all_stock_trading_data_LOAD.items():
                            if str_account_name in dict_account_to_tab_widget_name:
                                str_tab_widget_name = dict_account_to_tab_widget_name[ str_account_name ]
                                self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ].update( dict_per_account_all_stock_trading_data_LOAD )
                            else:
                                str_tab_widget_name = self.add_new_tab_and_table( str_account_name )
                                self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ] = dict_per_account_all_stock_trading_data_LOAD
                                dict_per_account_ui_state_LOAD = dict_all_account_ui_state_LOAD[ str_account_name ]
                                self.dict_all_account_ui_state[ str_tab_widget_name ] = dict_per_account_ui_state_LOAD
                                dict_per_account_general_data_LOAD = dict_all_account_general_data_LOAD[ str_account_name ]
                                self.dict_all_account_general_data[ str_tab_widget_name ] = dict_per_account_general_data_LOAD
                    else:
                        for str_account_name, dict_per_account_all_stock_trading_data_LOAD in dict_all_account_all_stock_trading_data_LOAD.items():
                            if str_account_name in dict_account_to_tab_widget_name:
                                str_tab_widget_name = dict_account_to_tab_widget_name[ str_account_name ]
                                for key_stock_number, value in dict_per_account_all_stock_trading_data_LOAD.items():
                                    if key_stock_number not in self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]:
                                        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ key_stock_number ] = value
                                    else:
                                        value.pop( 0 ) #移除第一筆資料 因為第一筆資料是虛的
                                        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ key_stock_number ].extend( value )
                            else:
                                str_tab_widget_name = self.add_new_tab_and_table( str_account_name )
                                self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ] = dict_per_account_all_stock_trading_data_LOAD
                                dict_per_account_ui_state_LOAD = dict_all_account_ui_state_LOAD[ str_account_name ]
                                self.dict_all_account_ui_state[ str_tab_widget_name ] = dict_per_account_ui_state_LOAD
                                dict_per_account_general_data_LOAD = dict_all_account_general_data_LOAD[ str_account_name ]
                                self.dict_all_account_general_data[ str_tab_widget_name ] = dict_per_account_general_data_LOAD
            else:
                for str_account_name, dict_per_account_all_stock_trading_data_LOAD in dict_all_account_all_stock_trading_data_LOAD.items():
                    if str_account_name in dict_account_to_tab_widget_name:
                        str_tab_widget_name = dict_account_to_tab_widget_name[ str_account_name ]
                        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ].update( dict_per_account_all_stock_trading_data_LOAD )
                    else:
                        str_tab_widget_name = self.add_new_tab_and_table( str_account_name )
                        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ] = dict_per_account_all_stock_trading_data_LOAD
                        dict_per_account_ui_state_LOAD = dict_all_account_ui_state_LOAD[ str_account_name ]
                        self.dict_all_account_ui_state[ str_tab_widget_name ] = dict_per_account_ui_state_LOAD
                        dict_per_account_general_data_LOAD = dict_all_account_general_data_LOAD[ str_account_name ]
                        self.dict_all_account_general_data[ str_tab_widget_name ] = dict_per_account_general_data_LOAD
                    self.dict_all_account_cash_transfer_data[ str_tab_widget_name ] = dict_all_account_cash_transfer_data_LOAD[ str_account_name ]

            self.process_all_trading_data()
            self.pick_up_stock( None )
            self.refresh_stock_list_table()
            self.process_all_transfer_data()
            self.refresh_transfer_data_table()
            self.clear_per_stock_trading_table()
            self.update_button_enable_disable_status()
            self.auto_save_trading_data()

    def on_import_single_stock_action_triggered( self ):
        file_path = share_ui.open_load_file_dialog( self, "匯入交易資料", "", share_ui.FileFilter.JSON )
        if file_path:
            dict_all_account_all_stock_trading_data_LOAD = {}
            dict_all_account_ui_state_LOAD = {}
            dict_all_account_cash_transfer_data_LOAD = {}
            dict_all_account_general_data_LOAD = {}
            self.load_trading_data_and_create_tab( file_path, 
                                                   dict_all_account_all_stock_trading_data_LOAD, 
                                                   dict_all_account_ui_state_LOAD, 
                                                   dict_all_account_cash_transfer_data_LOAD, 
                                                   dict_all_account_general_data_LOAD,
                                                   False )
            if len( dict_all_account_all_stock_trading_data_LOAD ) == 0:
                return
            elif len( dict_all_account_all_stock_trading_data_LOAD ) > 1:
                share_ui.show_error_message_box_with_ok_button( "錯誤", "請選擇只包含單一帳戶及單一個股的檔案", self )
                return
            
            dict_per_account_all_stock_trading_data_LOAD = dict_all_account_all_stock_trading_data_LOAD.popitem()[ 1 ]
            if len( dict_per_account_all_stock_trading_data_LOAD ) == 0:
                return
            elif len( dict_per_account_all_stock_trading_data_LOAD ) > 1:
                share_ui.show_error_message_box_with_ok_button( "錯誤", "請選擇只包含單一帳戶及單一個股的檔案", self )
                return
            str_stock_number, list_trading_data = dict_per_account_all_stock_trading_data_LOAD.popitem()

            if self.ui.qtTabWidget.count() <= 1:
                share_ui.show_error_message_box_with_ok_button( "錯誤", "請先建立群組", self )
                return

            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
            dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]

            b_duplicate = str_stock_number in dict_per_account_all_stock_trading_data

            if b_duplicate:
                dialog = ImportDataDuplicateOptionDialog( self )
                if dialog.exec():
                    if dialog.b_overwrite:
                        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ str_stock_number ] = list_trading_data
                    else:
                        list_trading_data.pop( 0 ) #移除第一筆資料 因為第一筆資料是虛的
                        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ str_stock_number ].extend( list_trading_data )
            else:
                self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ str_stock_number ] = list_trading_data

            self.process_all_trading_data()
            self.pick_up_stock( None )
            self.refresh_stock_list_table()
            self.clear_per_stock_trading_table()
            self.update_button_enable_disable_status()
            self.auto_save_trading_data()

    def on_set_temporary_file_path_action_triggered( self ):
        str_default_exe_path = os.path.dirname( os.path.abspath(__file__) )
        global g_data_dir
        dialog = share_ui.SelectFolderDialog( share_icon.get_icon( share_icon.IconType.WINDOW ), reg_settings, str_default_exe_path, g_data_dir, g_user_dir, self )
        if dialog.exec():
            if getattr( sys, 'frozen', False ):
                # PyInstaller 打包後執行時
                reg_settings.setValue("TemporaryFolderPath", dialog.folder_path )
            else:
                reg_settings.setValue("TemporaryFolderPathDebug", dialog.folder_path )

            source_folder = os.path.join( g_data_dir, 'StockInventory' ) 
            destination_folder = os.path.join( dialog.folder_path, 'StockInventory' ) 

            for root, dirs, files in os.walk( source_folder ):
                # 計算對應的目的地資料夾路徑
                relative_path = os.path.relpath( root, source_folder )
                destination_path = os.path.join( destination_folder, relative_path )

                # 確保目標資料夾存在
                os.makedirs( destination_path, exist_ok = True )

                for filename in files:
                    if filename.endswith('.txt') or filename.endswith('.json') or filename.endswith('.config'):
                        source_file = os.path.join( root, filename )
                        destination_file = os.path.join( destination_path, filename )

                        # 複製檔案
                        shutil.copy2( source_file, destination_file )

            g_data_dir = dialog.folder_path
            self.trading_data_json_file_path = os.path.join( g_data_dir, 'StockInventory', 'TradingData.json' )
            self.UISetting_file_path = os.path.join( g_data_dir, 'StockInventory', 'UISetting.config' )
            self.stock_dividend_position_date_file_path = os.path.join( g_data_dir, 'StockInventory', 'StockDividendPositionDate.json' )

    def on_undo_action_triggered( self ):
        pass

    def on_redo_action_triggered( self ):
        pass

    def load_initialize_data( self ): 
        with QSignalBlocker( self.ui.qtTabWidget ):
            self.load_trading_data_and_create_tab( self.trading_data_json_file_path, 
                                                   self.dict_all_account_all_stock_trading_data, 
                                                   self.dict_all_account_ui_state, 
                                                   self.dict_all_account_cash_transfer_data, 
                                                   self.dict_all_account_general_data, 
                                                   True )
            self.load_share_UI_state()
            if len( self.dict_all_account_all_stock_trading_data ) == 0:
                str_tab_name = self.add_new_tab_and_table()
                self.dict_all_account_all_stock_trading_data[ str_tab_name ] = {}
                self.dict_all_account_ui_state[ str_tab_name ] = { "discount_checkbox": True, 
                                                                   "discount_value": 0.6, 
                                                                   "regular_buy_trading_price_type": TradingPriceType.PER_SHARE, 
                                                                   "regular_buy_trading_fee_type": TradingFeeType.VARIABLE, 
                                                                   "regular_buy_trading_fee_minimum": 1, 
                                                                   "regular_buy_trading_fee_constant": 1 }
                self.dict_all_account_cash_transfer_data[ str_tab_name ] = []
                self.dict_all_account_general_data[ str_tab_name ] = { "insurance_checkbox": False, 
                                                                       "minimum_common_trading_fee": 20, 
                                                                       "minimum_odd_trading_fee": 1, 
                                                                       "dividend_transfer_fee":{},
                                                                       "decimal_round_type": DecimalRoundType.ROUND_DOWN,
                                                                       "discount_time_type": DiscountTimeType.IMMEDIATE }
            self.dict_all_account_all_stock_trading_data_INITIAL = self.dict_all_account_all_stock_trading_data.copy()
            self.ui.qtTabWidget.setCurrentIndex( 0 )
            self.list_stock_list_table_horizontal_header = self.get_stock_list_horizontal_header()
            self.process_all_trading_data()
            self.refresh_stock_list_table()
            self.process_all_transfer_data()
            self.refresh_transfer_data_table()

    def load_trading_data_and_create_tab( self, file_path, 
                                                dict_all_account_all_stock_trading_data, 
                                                dict_all_account_ui_state, 
                                                dict_all_account_cash_transfer, 
                                                dict_all_account_general_data,
                                                b_create_tab ): 
        if not os.path.exists( file_path ):
            return

        with open( file_path, 'r', encoding='utf-8' ) as f:
            version = f.readline().strip()
            data = json.load( f )

        if version == "v1.0.0":
            for item_account in data:
                if "account_name" in item_account and \
                   "trading_data" in item_account and \
                   "discount_checkbox" in item_account and \
                   "discount_value" in item_account and \
                   "insurance_checkbox" in item_account:
                    dict_per_account_all_stock_trading_data = item_account[ "trading_data" ]
                    dict_ui_state = {}
                    dict_ui_state[ "discount_checkbox" ] = item_account[ "discount_checkbox" ]
                    dict_ui_state[ "discount_value" ] = item_account[ "discount_value" ]
                    dict_ui_state[ "regular_buy_trading_price_type" ] = TradingPriceType.PER_SHARE
                    dict_ui_state[ "regular_buy_trading_fee_type" ] = TradingFeeType.VARIABLE
                    dict_ui_state[ "regular_buy_trading_fee_minimum" ] = 1
                    dict_ui_state[ "regular_buy_trading_fee_constant" ] = 1
                    dict_general_data = {}
                    dict_general_data[ "insurance_checkbox" ] = item_account[ "insurance_checkbox" ]
                    dict_general_data[ "minimum_common_trading_fee" ] = 20
                    dict_general_data[ "minimum_odd_trading_fee" ] = 1
                    dict_general_data[ "dividend_transfer_fee" ] = {}
                    dict_general_data[ "decimal_round_type" ] = DecimalRoundType.ROUND_DOWN
                    dict_general_data[ "discount_time_type" ] = DiscountTimeType.IMMEDIATE

                    dict_per_stock_trading_data = {} 
                    for key_stock_number, value_list_trading_data  in dict_per_account_all_stock_trading_data.items():
                        list_trading_data = []
                        for item_trading_data in value_list_trading_data:
                            if ( "trading_date" in item_trading_data and
                                 "trading_type" in item_trading_data and
                                 "trading_price" in item_trading_data and
                                 "trading_count" in item_trading_data and
                                 "trading_fee_discount" in item_trading_data and
                                 "stock_dividend_per_share" in item_trading_data and
                                 "cash_dividend_per_share" in item_trading_data and
                                 "capital_reduction_per_share" in item_trading_data ):
                                
                                if item_trading_data[ "trading_type" ] == 3:
                                    e_trading_type = TradingType.CAPITAL_INCREASE
                                elif item_trading_data[ "trading_type" ] == 4:
                                    e_trading_type = TradingType.DIVIDEND
                                elif item_trading_data[ "trading_type" ] == 5:
                                    e_trading_type = TradingType.CAPITAL_REDUCTION
                                else:
                                    e_trading_type = TradingType( item_trading_data[ "trading_type" ] )
                                dict_per_trading_data = Utility.generate_trading_data( item_trading_data[ "trading_date" ],                #交易日期
                                                                                       e_trading_type,                                     #交易種類
                                                                                       TradingPriceType.PER_SHARE,                         #交易價格種類
                                                                                       item_trading_data[ "trading_price" ],               #每股交易價格
                                                                                       0,                                                  #總交易價格
                                                                                       item_trading_data[ "trading_count" ],               #交易股數
                                                                                       TradingFeeType.VARIABLE,                            #手續費種類
                                                                                       item_trading_data[ "trading_fee_discount" ],        #手續費折扣
                                                                                       0,                                                  #手續費最低金額
                                                                                       0,                                                  #手續費固定金額
                                                                                       DividendValueType.PER_SHARE,                        #股利金額種類
                                                                                       item_trading_data[ "stock_dividend_per_share" ],    #股票股利
                                                                                       item_trading_data[ "cash_dividend_per_share" ],     #現金股利
                                                                                       -1,                                                 #自訂的補充保費
                                                                                       item_trading_data[ "capital_reduction_per_share" ], #每股減資金額 
                                                                                       CapitalReductionType.CASH_RETURN,                   #減資種類
                                                                                       False )                                             #是否為當沖交易  
                                if item_trading_data[ "trading_date" ] == '0001-01-01':
                                    if item_trading_data[ "use_auto_dividend_data" ]:
                                        dict_per_trading_data[ TradingData.USE_AUTO_DIVIDEND_DATA ] = AutoDividendType.AUTO
                                    else:
                                        dict_per_trading_data[ TradingData.USE_AUTO_DIVIDEND_DATA ] = AutoDividendType.MANUAL
                                list_trading_data.append( dict_per_trading_data )
                        dict_per_stock_trading_data[ key_stock_number ] = list_trading_data
                    if b_create_tab:
                        str_tab_name = self.add_new_tab_and_table( item_account[ "account_name" ] )
                        dict_all_account_ui_state[ str_tab_name ] = dict_ui_state
                        dict_all_account_all_stock_trading_data[ str_tab_name ] = dict_per_stock_trading_data
                        dict_all_account_cash_transfer[ str_tab_name ] = []
                        dict_all_account_general_data[ str_tab_name ] = dict_general_data
                    else:
                        dict_all_account_ui_state[ item_account[ "account_name" ] ] = dict_ui_state
                        dict_all_account_all_stock_trading_data[ item_account[ "account_name" ] ] = dict_per_stock_trading_data
                        dict_all_account_cash_transfer[ item_account[ "account_name" ] ] = []
                        dict_all_account_general_data[ item_account[ "account_name" ] ] = dict_general_data
        elif version == "v1.0.1":
            for item_account in data:
                if "account_name" in item_account and \
                   "trading_data" in item_account and \
                   "discount_checkbox" in item_account and \
                   "discount_value" in item_account and \
                   "insurance_checkbox" in item_account and \
                    "trading_fee_type" in item_account and \
                    "trading_fee_minimum" in item_account and \
                    "trading_fee_constant" in item_account:
                    dict_per_account_all_stock_trading_data = item_account[ "trading_data" ]
                    dict_ui_state = {}
                    dict_ui_state[ "discount_checkbox" ] = item_account[ "discount_checkbox" ]
                    dict_ui_state[ "discount_value" ] = item_account[ "discount_value" ]
                    dict_ui_state[ "regular_buy_trading_price_type" ] = TradingPriceType.PER_SHARE
                    dict_ui_state[ "regular_buy_trading_fee_type" ] = TradingFeeType( item_account[ "trading_fee_type" ] )
                    dict_ui_state[ "regular_buy_trading_fee_minimum" ] = item_account[ "trading_fee_minimum" ]
                    dict_ui_state[ "regular_buy_trading_fee_constant" ] = item_account[ "trading_fee_constant" ]
                    dict_general_data = {}
                    dict_general_data[ "insurance_checkbox" ] = item_account[ "insurance_checkbox" ]
                    dict_general_data[ "minimum_common_trading_fee" ] = 20
                    dict_general_data[ "minimum_odd_trading_fee" ] = 1
                    dict_general_data[ "dividend_transfer_fee" ] = {}
                    dict_general_data[ "decimal_round_type" ] = DecimalRoundType.ROUND_DOWN
                    dict_general_data[ "discount_time_type" ] = DiscountTimeType.IMMEDIATE

                    dict_per_stock_trading_data = {} 
                    for key_stock_number, value_list_trading_data  in dict_per_account_all_stock_trading_data.items():
                        list_trading_data = []
                        for item_trading_data in value_list_trading_data:
                            if ( "trading_date" in item_trading_data and
                                 "trading_type" in item_trading_data and
                                 "trading_price" in item_trading_data and
                                 "trading_count" in item_trading_data and
                                 "trading_fee_discount" in item_trading_data and
                                 "stock_dividend_per_share" in item_trading_data and
                                 "cash_dividend_per_share" in item_trading_data and
                                 "capital_reduction_per_share" in item_trading_data ):
                                
                                e_trading_type = TradingType( item_trading_data[ "trading_type" ] )
                                e_trading_fee_type = TradingFeeType( item_trading_data[ "trading_fee_type" ] )
                                dict_per_trading_data = Utility.generate_trading_data( item_trading_data[ "trading_date" ],                #交易日期
                                                                                       e_trading_type,                                     #交易種類
                                                                                       TradingPriceType.PER_SHARE,                         #交易價格種類
                                                                                       item_trading_data[ "trading_price" ],               #每股交易價格
                                                                                       0,                                                  #總交易價格
                                                                                       item_trading_data[ "trading_count" ],               #交易股數
                                                                                       e_trading_fee_type,                                 #手續費種類
                                                                                       item_trading_data[ "trading_fee_discount" ],        #手續費折扣
                                                                                       item_trading_data[ "trading_fee_minimum" ],         #手續費最低金額
                                                                                       item_trading_data[ "trading_fee_constant" ],        #手續費固定金額
                                                                                       DividendValueType.PER_SHARE,                        #股利金額種類
                                                                                       item_trading_data[ "stock_dividend_per_share" ],    #股票股利
                                                                                       item_trading_data[ "cash_dividend_per_share" ],     #現金股利
                                                                                       -1,                                                 #自訂的補充保費
                                                                                       item_trading_data[ "capital_reduction_per_share" ], #每股減資金額  
                                                                                       CapitalReductionType.CASH_RETURN,                   #減資種類   
                                                                                       item_trading_data[ "daying_trading" ] )             #是否為當沖交易
                                if item_trading_data[ "trading_date" ] == '0001-01-01':
                                    if item_trading_data[ "use_auto_dividend_data" ]:
                                        dict_per_trading_data[ TradingData.USE_AUTO_DIVIDEND_DATA ] = AutoDividendType.AUTO
                                    else:
                                        dict_per_trading_data[ TradingData.USE_AUTO_DIVIDEND_DATA ] = AutoDividendType.MANUAL
                                list_trading_data.append( dict_per_trading_data )
                        dict_per_stock_trading_data[ key_stock_number ] = list_trading_data

                    list_transfer_data = item_account[ "transfer_data" ]
                    list_per_account_cash_transfer_data = []
                    for item_transfer_data in list_transfer_data:
                        if ( "transfer_date" in item_transfer_data and
                             "transfer_type" in item_transfer_data and
                             "transfer_value" in item_transfer_data ):
                            dict_per_transfer_data = {}
                            dict_per_transfer_data[ TransferData.TRANSFER_DATE ] = item_transfer_data[ "transfer_date" ]
                            dict_per_transfer_data[ TransferData.TRANSFER_TYPE ] = TransferType( item_transfer_data[ "transfer_type" ] )
                            dict_per_transfer_data[ TransferData.TRANSFER_VALUE ] = item_transfer_data[ "transfer_value" ]
                            list_per_account_cash_transfer_data.append( dict_per_transfer_data )

                    if b_create_tab:
                        str_tab_name = self.add_new_tab_and_table( item_account[ "account_name" ] )
                        dict_all_account_ui_state[ str_tab_name ] = dict_ui_state
                        dict_all_account_all_stock_trading_data[ str_tab_name ] = dict_per_stock_trading_data
                        dict_all_account_cash_transfer[ str_tab_name ] = list_per_account_cash_transfer_data
                        dict_all_account_general_data[ str_tab_name ] = dict_general_data
                    else:
                        dict_all_account_ui_state[ item_account[ "account_name" ] ] = dict_ui_state
                        dict_all_account_all_stock_trading_data[ item_account[ "account_name" ] ] = dict_per_stock_trading_data
                        dict_all_account_cash_transfer[ item_account[ "account_name" ] ] = list_per_account_cash_transfer_data
                        dict_all_account_general_data[ item_account[ "account_name" ] ] = dict_general_data
        else:
            for item_account in data:
                dict_ui_state = {}
                dict_ui_state[ "discount_checkbox" ] = item_account[ "discount_checkbox" ]
                dict_ui_state[ "discount_value" ] = item_account[ "discount_value" ]
                dict_ui_state[ "regular_buy_trading_price_type" ] = TradingPriceType( item_account[ "regular_buy_trading_price_type" ] )
                dict_ui_state[ "regular_buy_trading_fee_type" ] = TradingFeeType( item_account[ "regular_buy_trading_fee_type" ] )
                dict_ui_state[ "regular_buy_trading_fee_minimum" ] = item_account[ "regular_buy_trading_fee_minimum" ]
                dict_ui_state[ "regular_buy_trading_fee_constant" ] = item_account[ "regular_buy_trading_fee_constant" ]
                dict_general_data = {}
                dict_general_data[ "insurance_checkbox" ] = item_account[ "insurance_checkbox" ]
                dict_general_data[ "minimum_common_trading_fee" ] = item_account[ "minimum_common_trading_fee" ]
                dict_general_data[ "minimum_odd_trading_fee" ] = item_account[ "minimum_odd_trading_fee" ]
                dict_general_data[ "dividend_transfer_fee" ] = item_account[ "dividend_transfer_fee" ]
                if share_api.compare_version_order( "v2.3.0", version ):
                    dict_general_data[ "decimal_round_type" ] = DecimalRoundType( item_account[ "decimal_round_type" ] )
                    dict_general_data[ "discount_time_type" ] = DiscountTimeType( item_account[ "discount_time_type" ] )
                else:
                    dict_general_data[ "decimal_round_type" ] = DecimalRoundType.ROUND_DOWN
                    dict_general_data[ "discount_time_type" ] = DiscountTimeType.IMMEDIATE

                dict_per_stock_trading_data = {} 
                dict_per_account_all_stock_trading_data = item_account[ "trading_data" ]
                for key_stock_number, value_list_trading_data  in dict_per_account_all_stock_trading_data.items():
                    list_trading_data = []
                    for item_trading_data in value_list_trading_data:
                        e_trading_type = TradingType( item_trading_data[ "trading_type" ] )
                        e_trading_price_type = TradingPriceType( item_trading_data[ "trading_price_type" ] )
                        e_trading_fee_type = TradingFeeType( item_trading_data[ "trading_fee_type" ] )
                        e_capital_reduction_type = CapitalReductionType( item_trading_data[ "capital_reduction_type" ] )
                        e_dividend_value_type = DividendValueType( item_trading_data[ "dividend_value_type" ] )
                        dict_per_trading_data = Utility.generate_trading_data( item_trading_data[ "trading_date" ],                #交易日期
                                                                                e_trading_type,                                     #交易種類
                                                                                e_trading_price_type,                               #交易價格種類
                                                                                item_trading_data[ "per_share_trading_price" ],     #每股交易價格
                                                                                item_trading_data[ "total_trading_price" ],         #總交易價格
                                                                                item_trading_data[ "trading_count" ],               #交易股數
                                                                                e_trading_fee_type,                                 #手續費種類
                                                                                item_trading_data[ "trading_fee_discount" ],        #手續費折扣
                                                                                item_trading_data[ "trading_fee_minimum" ],         #手續費最低金額
                                                                                item_trading_data[ "trading_fee_constant" ],        #手續費固定金額
                                                                                e_dividend_value_type,                              #股利金額種類
                                                                                item_trading_data[ "stock_dividend" ],              #股票股利
                                                                                item_trading_data[ "cash_dividend" ],               #現金股利
                                                                                item_trading_data[ "custom_extra_insurance_fee" ],  #自訂的補充保費
                                                                                item_trading_data[ "capital_reduction_per_share" ], #每股減資金額
                                                                                e_capital_reduction_type,                           #減資種類     
                                                                                item_trading_data[ "daying_trading" ] )             #是否為當沖交易
                        if item_trading_data[ "trading_date" ] == '0001-01-01':
                            if share_api.compare_version_order( "v2.1.0", version ):
                                dict_per_trading_data[ TradingData.USE_AUTO_DIVIDEND_DATA ] = AutoDividendType( item_trading_data[ "use_auto_dividend_data" ] )
                            else:
                                if item_trading_data[ "use_auto_dividend_data" ]:
                                    dict_per_trading_data[ TradingData.USE_AUTO_DIVIDEND_DATA ] = AutoDividendType.AUTO
                                else:
                                    dict_per_trading_data[ TradingData.USE_AUTO_DIVIDEND_DATA ] = AutoDividendType.MANUAL
                        list_trading_data.append( dict_per_trading_data )
                    dict_per_stock_trading_data[ key_stock_number ] = list_trading_data

                list_transfer_data = item_account[ "transfer_data" ]
                list_per_account_cash_transfer_data = []
                for item_transfer_data in list_transfer_data:
                    if ( "transfer_date" in item_transfer_data and
                         "transfer_type" in item_transfer_data and
                         "transfer_value" in item_transfer_data ):
                        dict_per_transfer_data = {}
                        dict_per_transfer_data[ TransferData.TRANSFER_DATE ] = item_transfer_data[ "transfer_date" ]
                        dict_per_transfer_data[ TransferData.TRANSFER_TYPE ] = TransferType( item_transfer_data[ "transfer_type" ] )
                        dict_per_transfer_data[ TransferData.TRANSFER_VALUE ] = item_transfer_data[ "transfer_value" ]
                        list_per_account_cash_transfer_data.append( dict_per_transfer_data )

                if b_create_tab:
                    str_tab_name = self.add_new_tab_and_table( item_account[ "account_name" ] )
                    dict_all_account_ui_state[ str_tab_name ] = dict_ui_state
                    dict_all_account_all_stock_trading_data[ str_tab_name ] = dict_per_stock_trading_data
                    dict_all_account_cash_transfer[ str_tab_name ] = list_per_account_cash_transfer_data
                    dict_all_account_general_data[ str_tab_name ] = dict_general_data
                else:
                    dict_all_account_ui_state[ item_account[ "account_name" ] ] = dict_ui_state
                    dict_all_account_all_stock_trading_data[ item_account[ "account_name" ] ] = dict_per_stock_trading_data
                    dict_all_account_cash_transfer[ item_account[ "account_name" ] ] = list_per_account_cash_transfer_data
                    dict_all_account_general_data[ item_account[ "account_name" ] ] = dict_general_data

    def auto_save_trading_data( self ): 
        list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
        self.manual_save_trading_data( list_save_tab_widget, self.trading_data_json_file_path )

    def manual_save_trading_data( self, list_save_tab_indice, file_path, str_stock_number = None ): 
        export_list_all_account_all_stock_trading_data = []

        for index in range( self.ui.qtTabWidget.count() - 1 ):
            if index not in list_save_tab_indice:
                continue
            tab_widget = self.ui.qtTabWidget.widget( index )

            str_tab_title = self.ui.qtTabWidget.tabText( index )
            str_tab_widget_name = tab_widget.objectName()
            value_dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
            export_dict_per_account_all_info = {}
            export_dict_per_account_all_stock_trading_data = {}
            for key_stock, value_list_per_stock_trading_data in value_dict_per_account_all_stock_trading_data.items():
                if str_stock_number is not None and key_stock != str_stock_number:
                    continue
                export_data = []
                for item in value_list_per_stock_trading_data:
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE in item and item[ TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE ] == True:
                        continue
                    dict_per_trading_data = {}
                    dict_per_trading_data[ "trading_date" ] = item[ TradingData.TRADING_DATE ]
                    dict_per_trading_data[ "trading_type" ] = int( item[ TradingData.TRADING_TYPE ].value )
                    dict_per_trading_data[ "trading_price_type" ] = int( item[ TradingData.TRADING_PRICE_TYPE ].value )
                    if item[ TradingData.TRADING_TYPE ] == TradingType.CAPITAL_REDUCTION: #CAPITAL_REDUCTION 為了顯示，所以需要寫一些數值進去，但實際上不用存
                        dict_per_trading_data[ "per_share_trading_price" ] = 0
                        dict_per_trading_data[ "total_trading_price" ] = 0
                        dict_per_trading_data[ "trading_count" ] = 0
                    else:
                        dict_per_trading_data[ "per_share_trading_price" ] = item[ TradingData.PER_SHARE_TRADING_PRICE ]
                        dict_per_trading_data[ "total_trading_price" ] = item[ TradingData.TOTAL_TRADING_PRICE ]
                        dict_per_trading_data[ "trading_count" ] = item[ TradingData.TRADING_QUANTITY ]
                    dict_per_trading_data[ "trading_fee_type" ] = int( item[ TradingData.REGULAR_BUY_TRADING_FEE_TYPE ].value )
                    dict_per_trading_data[ "trading_fee_discount" ] = item[ TradingData.TRADING_FEE_DISCOUNT ]
                    dict_per_trading_data[ "trading_fee_minimum" ] = item[ TradingData.REGULAR_BUY_TRADING_FEE_MINIMUM ]
                    dict_per_trading_data[ "trading_fee_constant" ] = item[ TradingData.REGULAR_BUY_TRADING_FEE_CONSTANT ]
                    dict_per_trading_data[ "dividend_value_type" ] = int( item[ TradingData.DIVIDEND_VALUE_TYPE ].value )
                    dict_per_trading_data[ "stock_dividend" ] = item[ TradingData.STOCK_DIVIDEND ]
                    dict_per_trading_data[ "cash_dividend" ] = item[ TradingData.CASH_DIVIDEND ]
                    dict_per_trading_data[ "custom_extra_insurance_fee" ] = item[ TradingData.CUSTOM_EXTRA_INSURANCE_FEE ]
                    dict_per_trading_data[ "capital_reduction_per_share" ] = item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ]
                    dict_per_trading_data[ "capital_reduction_type" ] = int( item[ TradingData.CAPITAL_REDUCTION_TYPE ].value )
                    if dict_per_trading_data[ "trading_date" ] == '0001-01-01':
                        dict_per_trading_data[ "use_auto_dividend_data" ] = int( item[ TradingData.USE_AUTO_DIVIDEND_DATA ].value )
                    dict_per_trading_data[ "daying_trading" ] = item[ TradingData.DAYING_TRADING ]

                    export_data.append( dict_per_trading_data )
                export_dict_per_account_all_stock_trading_data[ key_stock ] = export_data

            list_transfer_data = []
            value_dict_per_account_all_cash_transfer_data = self.dict_all_account_cash_transfer_data[ str_tab_widget_name ]
            for dict_cash_transfer_data in value_dict_per_account_all_cash_transfer_data:
                dict_per_transfer_data = {}
                dict_per_transfer_data[ "transfer_date" ] = dict_cash_transfer_data[ TransferData.TRANSFER_DATE ]
                dict_per_transfer_data[ "transfer_type" ] = int( dict_cash_transfer_data[ TransferData.TRANSFER_TYPE ].value )
                dict_per_transfer_data[ "transfer_value" ] = dict_cash_transfer_data[ TransferData.TRANSFER_VALUE ]
                list_transfer_data.append( dict_per_transfer_data )

            export_dict_per_account_all_info[ "account_name" ] = str_tab_title
            export_dict_per_account_all_info[ "trading_data" ] = export_dict_per_account_all_stock_trading_data
            export_dict_per_account_all_info[ "transfer_data" ] = list_transfer_data
            
            export_dict_per_account_all_info[ "insurance_checkbox" ] = self.dict_all_account_general_data[ str_tab_widget_name ][ "insurance_checkbox" ]
            export_dict_per_account_all_info[ "minimum_common_trading_fee" ] = self.dict_all_account_general_data[ str_tab_widget_name ][ "minimum_common_trading_fee" ]#現股交易最低手續費
            export_dict_per_account_all_info[ "minimum_odd_trading_fee" ] = self.dict_all_account_general_data[ str_tab_widget_name ][ "minimum_odd_trading_fee" ]#零股交易最低手續費
            export_dict_per_account_all_info[ "dividend_transfer_fee" ] = self.dict_all_account_general_data[ str_tab_widget_name ][ "dividend_transfer_fee" ]
            export_dict_per_account_all_info[ "decimal_round_type" ] = int( self.dict_all_account_general_data[ str_tab_widget_name ][ "decimal_round_type" ].value )
            export_dict_per_account_all_info[ "discount_time_type" ] = int( self.dict_all_account_general_data[ str_tab_widget_name ][ "discount_time_type" ].value )
            
            export_dict_per_account_all_info[ "discount_checkbox" ] = self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"]
            export_dict_per_account_all_info[ "discount_value" ] = self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"]
            export_dict_per_account_all_info[ "regular_buy_trading_price_type" ] = int( self.dict_all_account_ui_state[ str_tab_widget_name ][ "regular_buy_trading_price_type"].value )
            export_dict_per_account_all_info[ "regular_buy_trading_fee_type" ] = int( self.dict_all_account_ui_state[ str_tab_widget_name ][ "regular_buy_trading_fee_type"].value )
            export_dict_per_account_all_info[ "regular_buy_trading_fee_minimum" ] = self.dict_all_account_ui_state[ str_tab_widget_name ][ "regular_buy_trading_fee_minimum"]
            export_dict_per_account_all_info[ "regular_buy_trading_fee_constant" ] = self.dict_all_account_ui_state[ str_tab_widget_name ][ "regular_buy_trading_fee_constant"]
            
            export_list_all_account_all_stock_trading_data.append( export_dict_per_account_all_info )

        with open( file_path, 'w', encoding='utf-8' ) as f:
            f.write( "v2.3.0" '\n' )
            json.dump( export_list_all_account_all_stock_trading_data, f, ensure_ascii=False, indent=4 )

    def save_share_UI_state( self ): 
        # 確保目錄存在，若不存在則遞歸創建
        os.makedirs( os.path.dirname( self.UISetting_file_path ), exist_ok = True )

        with open( self.UISetting_file_path, 'w', encoding='utf-8' ) as f:
            f.write( "版本," + 'v2.2.0' + '\n' )
            f.write( "顯示排序," + str( self.ui.qtFromNewToOldAction.isChecked() ) + '\n' )
            f.write( "顯示數量," + str( self.ui.qtShowAllAction.isChecked() ) + '\n' )
            f.write( "顯示單位," + str( self.ui.qtUse1ShareUnitAction.isChecked() ) + '\n' )
            f.write( "年度顯示," + str( self.ui.qtADYearAction.isChecked() ) + '\n' )
            f.write( "成本扣股利," + str( self.ui.qtCostWithInDividendAction.isChecked() ) + '\n' )
            f.write( "欄寬" )
            for i in range( len( self.list_stock_list_column_width ) ):
                f.write( f",{ self.list_stock_list_column_width[ i ] }" )
            f.write( "\n" )
            f.write( "顯示欄位" )
            for i in range( len( self.list_show_stock_info ) ):
                f.write( f",{ self.list_show_stock_info[ i ].value }" )
            f.write( "\n" )

    def load_share_UI_state( self ): 
        with ( QSignalBlocker( self.ui.qtFromNewToOldAction ),
               QSignalBlocker( self.ui.qtFromOldToNewAction ), 
               QSignalBlocker( self.ui.qtShowAllAction ), 
               QSignalBlocker( self.ui.qtShow10Action ),
               QSignalBlocker( self.ui.qtUse1ShareUnitAction ), 
               QSignalBlocker( self.ui.qtUse1000ShareUnitAction ),
               QSignalBlocker( self.ui.qtADYearAction ), 
               QSignalBlocker( self.ui.qtROCYearAction ),
               QSignalBlocker( self.ui.qtCostWithInDividendAction ), 
               QSignalBlocker( self.ui.qtCostWithOutDividendAction ) 
               ):

            if os.path.exists( self.UISetting_file_path ):
                with open( self.UISetting_file_path, 'r', encoding='utf-8' ) as f:
                    data = f.readlines()
                    for i, row in enumerate( data ):
                        row = row.strip().split( ',' )
                        if row[0] == "版本":
                            str_version = row[ 1 ]
                        elif row[0] == "顯示排序":
                            if row[ 1 ] == 'True':
                                self.ui.qtFromNewToOldAction.setChecked( True )
                                self.ui.qtFromOldToNewAction.setChecked( False )
                            else:
                                self.ui.qtFromNewToOldAction.setChecked( False )
                                self.ui.qtFromOldToNewAction.setChecked( True )
                        elif row[0] == "顯示數量":
                            if row[ 1 ] == 'True':
                                self.ui.qtShowAllAction.setChecked( True )
                                self.ui.qtShow10Action.setChecked( False )
                            else:
                                self.ui.qtShowAllAction.setChecked( False )
                                self.ui.qtShow10Action.setChecked( True )
                        elif row[0] == "顯示單位":
                            if row[ 1 ] == 'True':
                                self.ui.qtUse1ShareUnitAction.setChecked( True )
                                self.ui.qtUse1000ShareUnitAction.setChecked( False )
                            else:
                                self.ui.qtUse1ShareUnitAction.setChecked( False )
                                self.ui.qtUse1000ShareUnitAction.setChecked( True )
                        elif row[0] == "年度顯示":
                            if row[ 1 ] == 'True':
                                self.ui.qtADYearAction.setChecked( True )
                                self.ui.qtROCYearAction.setChecked( False )
                            else:
                                self.ui.qtADYearAction.setChecked( False )
                                self.ui.qtROCYearAction.setChecked( True )
                        elif row[0] == "成本扣股利":
                            if row[ 1 ] == 'True':
                                self.ui.qtCostWithInDividendAction.setChecked( True )
                                self.ui.qtCostWithOutDividendAction.setChecked( False )
                            else:
                                self.ui.qtCostWithInDividendAction.setChecked( False )
                                self.ui.qtCostWithOutDividendAction.setChecked( True )
                        elif row[0] == '欄寬':
                            self.list_stock_list_column_width = []
                            #StockInfoType 不包含"股票代碼"、"自動股利"、"匯入"、"刪除" 這四項
                            if len( row ) + 1 < len( list( StockInfoType ) ) + 4: # row[0]='欄寬' 
                                #進到這裡表示這個版本有新增欄位
                                for i in range( 1, len( row ) - 2 ):
                                    self.list_stock_list_column_width.append( int( row[ i ] ) )
                                for i in range( len( self.list_stock_list_column_width ), len( list( StockInfoType ) ) + 2 ):
                                    self.list_stock_list_column_width.append( 85 )
                                for i in range( len( row ) - 2, len( row ) ):
                                    self.list_stock_list_column_width.append( int( row[ i ] ) )
                                pass
                            else:
                                for i in range( 1, len( row ) ):
                                    self.list_stock_list_column_width.append( int( row[ i ] ) )
                        elif row[0] == '顯示欄位':
                            self.list_show_stock_info = []
                            for i in range( 1, len( row ) ):
                                self.list_show_stock_info.append( StockInfoType( int( row[ i ] ) ) )

    def save_stock_dividend_position_date( self ):
        with open( self.stock_dividend_position_date_file_path, 'w', encoding='utf-8' ) as f:
            f.write( "v2.1.0" '\n' )
            json.dump( self.dict_all_stock_dividend_position_date, f, ensure_ascii=False, indent=4 )

    def load_stock_dividend_position_date( self ):
        data = {}
        if os.path.exists( self.stock_dividend_position_date_file_path ):
            with open( self.stock_dividend_position_date_file_path, 'r', encoding='utf-8' ) as f:
                version = f.readline().strip()
                data = json.load( f )
        return data

    def process_all_trading_data( self ): 
        for key_account_name, value_dict_per_company_trading_data in self.dict_all_account_all_stock_trading_data.items():
            for key_stock_name, value_list_trading_data in value_dict_per_company_trading_data.items():
                self.process_single_trading_data( key_account_name, key_stock_name )
    
    def process_single_trading_data( self, str_tab_widget_name, str_stock_number ): 
        list_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ str_stock_number ].copy()
        #先拔掉所有的AUTO_DIVIDEND_DATA，下面有需要再重新插入，避免重複插入
        list_trading_data = [item for item in list_trading_data if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE not in item or not item[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]]

        str_stock_name = self.dict_all_company_number_to_name_and_type[ str_stock_number ][ 0 ]
        b_bond = True if '債' in str_stock_name else False
        b_KY = True if 'KY' in str_stock_name else False
        b_extra_insurance_fee = self.dict_all_account_general_data[ str_tab_widget_name ][ "insurance_checkbox"]
        dict_company_number_to_transfer_fee = self.dict_all_account_general_data[ str_tab_widget_name ][ "dividend_transfer_fee" ]
        n_dividend_transfer_fee = dict_company_number_to_transfer_fee[ str_stock_number ] if str_stock_number in dict_company_number_to_transfer_fee else 10
        n_minimum_common_trading_fee = self.dict_all_account_general_data[ str_tab_widget_name ][ "minimum_common_trading_fee" ]
        n_minimum_odd_trading_fee = self.dict_all_account_general_data[ str_tab_widget_name ][ "minimum_odd_trading_fee" ]
        e_decimal_round_type = self.dict_all_account_general_data[ str_tab_widget_name ][ "decimal_round_type" ]
        e_discount_time_type = self.dict_all_account_general_data[ str_tab_widget_name ][ "discount_time_type" ]
        if b_KY:
            b_extra_insurance_fee = False

        str_b_etf = self.dict_all_company_number_to_name_and_type[ str_stock_number ][ 1 ]
        b_etf = True if str_b_etf == "True" else False
        sorted_list = sorted( list_trading_data, key=lambda x: ( datetime.datetime.strptime( x[ TradingData.TRADING_DATE ], "%Y-%m-%d"), -x[ TradingData.TRADING_TYPE ] ) )

        str_current_date = datetime.datetime.today().strftime("%Y-%m-%d")
        e_auto_dividend_type = sorted_list[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ]
        if e_auto_dividend_type == AutoDividendType.AUTO or e_auto_dividend_type == AutoDividendType.MANUAL_WITH_AUTO:
            if str_stock_number in self.dict_auto_stock_yearly_dividned:
                auto_list_dividend = copy.deepcopy( self.dict_auto_stock_yearly_dividned[ str_stock_number ] ) 
                auto_list_dividend = sorted( auto_list_dividend, key=lambda x: ( datetime.datetime.strptime( x[ TradingData.TRADING_DATE ], "%Y-%m-%d") ) )
                if len( sorted_list ) > 1:
                    first_data = sorted_list[ 1 ]
                    for index, auto_dividend_data in enumerate( auto_list_dividend ):
                        if auto_dividend_data[ TradingData.TRADING_DATE ] > first_data[ TradingData.TRADING_DATE ]:
                            if auto_dividend_data[ TradingData.TRADING_DATE ] > str_current_date:
                                break
                            sorted_list.append( auto_dividend_data )
                    sorted_list = sorted( sorted_list, key=lambda x: ( datetime.datetime.strptime( x[ TradingData.TRADING_DATE ], "%Y-%m-%d"), -x[ TradingData.TRADING_TYPE ] ) )

        dict_per_stock_dividend_position_date = {}
        if str_stock_number in self.dict_all_stock_dividend_position_date:
            dict_per_stock_dividend_position_date = self.dict_all_stock_dividend_position_date[ str_stock_number ]
        n_accumulated_inventory = 0
        n_accumulated_cost = 0
        n_accumulated_cost_without_considering_dividend = 0
        n_accumulated_stock_dividend = 0
        n_accumulated_cash_dividend = 0
        str_last_buying_date = ''
        n_last_buying_count = 0
        queue_buying_data = deque([])
        list_calibration_data = [] #因為若是已經沒有庫存股票，那麼股利分配或是減資的資料就不會被計算
        for index, item in enumerate( sorted_list ):
            item[ TradingData.SORTED_INDEX_NON_SAVE ] = index
            e_trading_type = item[ TradingData.TRADING_TYPE ]
            str_trading_date = item[ TradingData.TRADING_DATE ]
            obj_trading_date = datetime.datetime.strptime( str_trading_date, "%Y-%m-%d")
            
            if e_trading_type == TradingType.TEMPLATE:
                list_calibration_data.append( item )
                continue
            elif e_trading_type == TradingType.BUY:
                f_trading_price = item[ TradingData.PER_SHARE_TRADING_PRICE ]
                n_trading_count = item[ TradingData.TRADING_QUANTITY ]
                f_trading_fee_discount = item[ TradingData.TRADING_FEE_DISCOUNT ]
                
                dict_result = Utility.compute_cost( e_trading_type, e_decimal_round_type, f_trading_price, n_trading_count, f_trading_fee_discount, n_minimum_common_trading_fee, n_minimum_odd_trading_fee, b_etf, False, b_bond )
                item[ TradingData.TRADING_VALUE_NON_SAVE ] = dict_result[ TradingCost.TRADING_VALUE ]
                item[ TradingData.TRADING_FEE_NON_SAVE ] = dict_result[ TradingCost.TRADING_FEE ]
                item[ TradingData.TRADING_TAX_NON_SAVE ] = dict_result[ TradingCost.TRADING_TAX ]
                item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0 
                n_per_trading_total_cost = item[ TradingData.TRADING_COST_NON_SAVE ] = dict_result[ TradingCost.TRADING_TOTAL_COST ]
                n_accumulated_inventory += n_trading_count
                n_accumulated_cost += n_per_trading_total_cost
                n_accumulated_cost_without_considering_dividend += n_per_trading_total_cost

                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
                if str_last_buying_date == str_trading_date:
                    if item[ TradingData.DAYING_TRADING ] == True:
                        n_last_buying_count += n_trading_count
                else:
                    if item[ TradingData.DAYING_TRADING ] == True:
                        str_last_buying_date = str_trading_date
                        n_last_buying_count = n_trading_count
                list_buying_data = [ n_per_trading_total_cost, n_trading_count, str_trading_date ]
                queue_buying_data.append( list_buying_data )
                queue_buying_data = deque( sorted( queue_buying_data, key = lambda x: x[ 2 ] ) )
            elif e_trading_type == TradingType.REGULAR_BUY:
                n_trading_count = item[ TradingData.TRADING_QUANTITY ]
                e_trading_price_type = item[ TradingData.TRADING_PRICE_TYPE ]
                if e_trading_price_type == TradingPriceType.PER_SHARE:
                    f_trading_price = item[ TradingData.PER_SHARE_TRADING_PRICE ]
                    n_trading_value = int( f_trading_price * n_trading_count )
                else:
                    n_trading_value = item[ TradingData.TOTAL_TRADING_PRICE ]

                e_trading_fee_type = item[ TradingData.REGULAR_BUY_TRADING_FEE_TYPE ]

                if n_trading_value == 0:
                    n_trading_fee = 0
                elif e_trading_fee_type == TradingFeeType.VARIABLE:
                    f_trading_fee_discount = Decimal( str( item[ TradingData.TRADING_FEE_DISCOUNT ] ) )
                    n_trading_fee_minimum = item[ TradingData.REGULAR_BUY_TRADING_FEE_MINIMUM ]

                    n_trading_fee = Utility.round_to( n_trading_value * Decimal( '0.001425' ) * f_trading_fee_discount, e_decimal_round_type )
                    n_trading_fee = max( n_trading_fee_minimum, n_trading_fee )
                else:
                    n_trading_fee = int( item[ TradingData.REGULAR_BUY_TRADING_FEE_CONSTANT ] )

                n_trading_total_cost = n_trading_value + n_trading_fee

                item[ TradingData.TRADING_VALUE_NON_SAVE ] = n_trading_value
                item[ TradingData.TRADING_FEE_NON_SAVE ] = n_trading_fee
                item[ TradingData.TRADING_TAX_NON_SAVE ] = 0
                item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0 
                item[ TradingData.TRADING_COST_NON_SAVE ] = n_trading_total_cost
                n_accumulated_inventory += n_trading_count
                n_accumulated_cost += n_trading_total_cost
                n_accumulated_cost_without_considering_dividend += n_trading_total_cost

                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
                list_buying_data = [ n_trading_total_cost, n_trading_count, str_trading_date ]
                queue_buying_data.append( list_buying_data )
                queue_buying_data = deque( sorted( queue_buying_data, key = lambda x: x[ 2 ] ) )
            elif e_trading_type == TradingType.SELL:
                f_trading_price = item[ TradingData.PER_SHARE_TRADING_PRICE ]
                f_trading_fee_discount = item[ TradingData.TRADING_FEE_DISCOUNT ]
                n_trading_count = item[ TradingData.TRADING_QUANTITY ]

                n_day_trading_selling_count = 0 #當沖數量
                n_general_selling_count = 0 #非當沖數量
                if ( str_trading_date == str_last_buying_date and    #賣出與買入同一天
                     item[ TradingData.DAYING_TRADING ] == True and  #有勾選是當沖交易
                     obj_trading_date >= datetime.datetime.strptime( '2017-04-28', "%Y-%m-%d" ) ): #交易日期在2017-04-28之後。因為在這之後才通過當沖交易稅減半
                    item[ TradingData.IS_REALLY_DAYING_TRADING_NON_SAVE ] = True
                    if n_trading_count <= n_last_buying_count: #賣出數量小於或等於買入數量，表示全部賣出數量都可視為當沖
                        dict_result = Utility.compute_cost( e_trading_type, e_decimal_round_type, f_trading_price, n_trading_count, f_trading_fee_discount, n_minimum_common_trading_fee, n_minimum_odd_trading_fee, b_etf, True, b_bond )
                        item[ TradingData.TRADING_VALUE_NON_SAVE ] = dict_result[ TradingCost.TRADING_VALUE ]
                        item[ TradingData.TRADING_FEE_NON_SAVE ] = dict_result[ TradingCost.TRADING_FEE ]
                        item[ TradingData.TRADING_TAX_NON_SAVE ] = dict_result[ TradingCost.TRADING_TAX ]
                        n_per_trading_total_cost = item[ TradingData.TRADING_COST_NON_SAVE ] = dict_result[ TradingCost.TRADING_TOTAL_COST ]
                        n_accumulated_cost -= n_per_trading_total_cost
                        n_accumulated_cost_without_considering_dividend -= n_per_trading_total_cost
                        n_accumulated_inventory -= n_trading_count
                        n_last_buying_count -= n_trading_count
                        n_day_trading_selling_count = n_trading_count
                    else: #賣出數量大於買入數量，表示只有部分數量都可視為當沖
                        n_trading_count_1 = n_last_buying_count
                        dict_result = Utility.compute_cost( e_trading_type, e_decimal_round_type, f_trading_price, n_trading_count_1, f_trading_fee_discount, n_minimum_common_trading_fee, n_minimum_odd_trading_fee, b_etf, True, b_bond )#這部分是當沖
                        n_trading_value_1 = dict_result[ TradingCost.TRADING_VALUE ]
                        n_trading_fee_1 = dict_result[ TradingCost.TRADING_FEE ]
                        n_trading_tax_1 = dict_result[ TradingCost.TRADING_TAX ]
                        n_trading_total_cost_1 = dict_result[ TradingCost.TRADING_TOTAL_COST ]
                        n_day_trading_selling_count = n_trading_count_1

                        n_trading_count_2 = n_trading_count - n_last_buying_count
                        dict_result = Utility.compute_cost( e_trading_type, e_decimal_round_type, f_trading_price, n_trading_count_2, f_trading_fee_discount, n_minimum_common_trading_fee, n_minimum_odd_trading_fee, b_etf, False, b_bond )#這部分不是當沖
                        n_trading_value_2 = dict_result[ TradingCost.TRADING_VALUE ]
                        n_trading_fee_2 = dict_result[ TradingCost.TRADING_FEE ]
                        n_trading_tax_2 = dict_result[ TradingCost.TRADING_TAX ]
                        n_trading_total_cost_2 = dict_result[ TradingCost.TRADING_TOTAL_COST ]
                        n_general_selling_count = n_trading_count_2

                        item[ TradingData.TRADING_VALUE_NON_SAVE ] = n_trading_value_1 + n_trading_value_2
                        item[ TradingData.TRADING_FEE_NON_SAVE ] = n_trading_fee_1 + n_trading_fee_2
                        item[ TradingData.TRADING_TAX_NON_SAVE ] = n_trading_tax_1 + n_trading_tax_2
                        n_per_trading_total_cost = item[ TradingData.TRADING_COST_NON_SAVE ] = n_trading_total_cost_1 + n_trading_total_cost_2
                        n_accumulated_cost -= n_per_trading_total_cost
                        n_accumulated_cost_without_considering_dividend -= n_per_trading_total_cost
                        n_accumulated_inventory -= n_trading_count

                        n_last_buying_count = 0
                    item[ TradingData.SELLING_PROFIT_NON_SAVE ] = 0
                else:
                    item[ TradingData.IS_REALLY_DAYING_TRADING_NON_SAVE ] = False
                    dict_result = Utility.compute_cost( e_trading_type, e_decimal_round_type, f_trading_price, n_trading_count, f_trading_fee_discount, n_minimum_common_trading_fee, n_minimum_odd_trading_fee, b_etf, False, b_bond )
                    item[ TradingData.TRADING_VALUE_NON_SAVE ] = dict_result[ TradingCost.TRADING_VALUE ]
                    item[ TradingData.TRADING_FEE_NON_SAVE ] = dict_result[ TradingCost.TRADING_FEE ]
                    item[ TradingData.TRADING_TAX_NON_SAVE ] = dict_result[ TradingCost.TRADING_TAX ]
                    n_per_trading_total_cost = item[ TradingData.TRADING_COST_NON_SAVE ] = dict_result[ TradingCost.TRADING_TOTAL_COST ]
                    n_accumulated_cost -= n_per_trading_total_cost
                    n_accumulated_cost_without_considering_dividend -= n_per_trading_total_cost
                    n_accumulated_inventory -= n_trading_count
                    n_general_selling_count = n_trading_count

                n_accumulate_cost_for_this_selling = 0
                #一般現股交易是先買先賣
                if len( queue_buying_data ) != 0 and n_general_selling_count > 0:
                    list_buying_data = queue_buying_data[ 0 ]
                    n_buying_cost = list_buying_data[ 0 ]
                    n_buying_count = list_buying_data[ 1 ]
                    str_date = list_buying_data[ 2 ]
                    if n_general_selling_count < n_buying_count:
                        n_used_buying_cost = int( n_buying_cost * n_general_selling_count / n_buying_count )
                        n_rest_buying_cost = n_buying_cost - n_used_buying_cost
                        n_rest_buying_count = n_buying_count - n_general_selling_count
                        queue_buying_data[ 0 ] = [ n_rest_buying_cost, n_rest_buying_count, str_date ]
                        n_accumulate_cost_for_this_selling += n_used_buying_cost
                    else:
                        while n_general_selling_count >= n_buying_count:
                            queue_buying_data.popleft()
                            n_general_selling_count -= n_buying_count
                            n_accumulate_cost_for_this_selling += n_buying_cost
                            if n_general_selling_count == 0:
                                break
                            if len( queue_buying_data ) == 0:
                                break
                            list_buying_data = queue_buying_data[ 0 ]
                            n_buying_cost = list_buying_data[ 0 ]
                            n_buying_count = list_buying_data[ 1 ]
                            if n_general_selling_count < n_buying_count:
                                n_used_buying_cost = int( n_buying_cost * n_general_selling_count / n_buying_count )
                                n_rest_buying_cost = n_buying_cost - n_used_buying_cost
                                n_rest_buying_count = n_buying_count - n_general_selling_count
                                queue_buying_data[ 0 ] = [ n_rest_buying_cost, n_rest_buying_count, str_date ]
                                n_accumulate_cost_for_this_selling += n_used_buying_cost
                
                #當沖交易則需要從最新的購買交易去取
                if len( queue_buying_data ) != 0 and n_day_trading_selling_count > 0:
                    list_buying_data = queue_buying_data[ -1 ]
                    n_buying_cost = list_buying_data[ 0 ]
                    n_buying_count = list_buying_data[ 1 ]
                    str_date = list_buying_data[ 2 ]
                    if n_day_trading_selling_count < n_buying_count:
                        n_used_buying_cost = int( n_buying_cost * n_day_trading_selling_count / n_buying_count )
                        n_rest_buying_cost = n_buying_cost - n_used_buying_cost
                        n_rest_buying_count = n_buying_count - n_day_trading_selling_count
                        queue_buying_data[ -1 ] = [ n_rest_buying_cost, n_rest_buying_count, str_date ]
                        n_accumulate_cost_for_this_selling += n_used_buying_cost
                    else:
                        while n_day_trading_selling_count >= n_buying_count:
                            queue_buying_data.pop()
                            n_day_trading_selling_count -= n_buying_count
                            n_accumulate_cost_for_this_selling += n_buying_cost
                            if n_day_trading_selling_count == 0:
                                break
                            if len( queue_buying_data ) == 0:
                                break
                            list_buying_data = queue_buying_data[ -1 ]
                            n_buying_cost = list_buying_data[ 0 ]
                            n_buying_count = list_buying_data[ 1 ]
                            if n_day_trading_selling_count < n_buying_count:
                                n_used_buying_cost = int( n_buying_cost * n_day_trading_selling_count / n_buying_count )
                                n_rest_buying_cost = n_buying_cost - n_used_buying_cost
                                n_rest_buying_count = n_buying_count - n_day_trading_selling_count
                                queue_buying_data[ -1 ] = [ n_rest_buying_cost, n_rest_buying_count, str_date ]
                                n_accumulate_cost_for_this_selling += n_used_buying_cost                

                item[ TradingData.SELLING_PROFIT_NON_SAVE ] = n_per_trading_total_cost - n_accumulate_cost_for_this_selling
                item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0
                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
            elif e_trading_type == TradingType.CAPITAL_INCREASE:
                n_trading_count = item[ TradingData.TRADING_QUANTITY ]
                e_trading_price_type = item[ TradingData.TRADING_PRICE_TYPE ]
                if e_trading_price_type == TradingPriceType.PER_SHARE:
                    f_trading_price = item[ TradingData.PER_SHARE_TRADING_PRICE ]
                    n_trading_value = int( f_trading_price * n_trading_count )
                else:
                    n_trading_value = item[ TradingData.TOTAL_TRADING_PRICE ]

                n_per_trading_total_cost = item[ TradingData.TRADING_COST_NON_SAVE ] = n_trading_value
                item[ TradingData.TRADING_VALUE_NON_SAVE ] = n_per_trading_total_cost
                item[ TradingData.TRADING_FEE_NON_SAVE ] = 0
                item[ TradingData.TRADING_TAX_NON_SAVE ] = 0
                item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0 
                n_accumulated_inventory += n_trading_count
                n_accumulated_cost += n_per_trading_total_cost
                n_accumulated_cost_without_considering_dividend += n_per_trading_total_cost

                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
                list_buying_data = [ 0, n_trading_count, str_trading_date ]
                queue_buying_data.append( list_buying_data )
                queue_buying_data = deque( sorted( queue_buying_data, key = lambda x: x[ 2 ] ) )
            elif e_trading_type == TradingType.DIVIDEND:
                if n_accumulated_inventory <= 0: #沒有庫存就不用算股利了
                    continue
                e_dividend_value_type = item[ TradingData.DIVIDEND_VALUE_TYPE ]
                if e_dividend_value_type == DividendValueType.PER_SHARE:
                    n_stock_dividend_value_gain = int( Decimal( str( item[ TradingData.STOCK_DIVIDEND ] ) ) * Decimal( str( n_accumulated_inventory ) ) ) # n_stock_dividend_value_gain單位為元
                    n_stock_dividend_share_gain = int( Decimal( str( item[ TradingData.STOCK_DIVIDEND ] ) ) * Decimal( str( n_accumulated_inventory ) ) / Decimal( '10' ) ) # n_stock_dividend_share_gain單位為股 除以10是因為票面額10元
                    n_cash_dividend_gain = int( Decimal( str(item[ TradingData.CASH_DIVIDEND ] ) ) * Decimal( str( n_accumulated_inventory ) ) ) # n_cash_dividend_gain單位為元
                else:
                    n_stock_dividend_value_gain = int( Decimal( str( item[ TradingData.STOCK_DIVIDEND ] ) ) * Decimal( '10' ) ) # n_stock_dividend_value_gain單位為元
                    n_stock_dividend_share_gain = int( Decimal( str( item[ TradingData.STOCK_DIVIDEND ] ) ) ) # n_stock_dividend_share_gain單位為股 除以10是因為票面額10元
                    n_cash_dividend_gain = int( Decimal( str(item[ TradingData.CASH_DIVIDEND ] ) ) ) # n_cash_dividend_gain單位為元

                if e_auto_dividend_type == AutoDividendType.AUTO:
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE not in item or not item[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]:#表示這筆股利資料是手動輸入
                        n_stock_dividend_share_gain = 0
                        n_stock_dividend_value_gain = 0
                        n_cash_dividend_gain = 0
                elif e_auto_dividend_type == AutoDividendType.MANUAL:
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE in item and item[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]:
                        n_stock_dividend_share_gain = 0
                        n_stock_dividend_value_gain = 0
                        n_cash_dividend_gain = 0

                item[ TradingData.TRADING_VALUE_NON_SAVE ] = 0
                item[ TradingData.TRADING_TAX_NON_SAVE ] = 0
                item[ TradingData.TRADING_COST_NON_SAVE ] = 0

                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = n_stock_dividend_share_gain
                if str_trading_date in dict_per_stock_dividend_position_date:
                    item[ TradingData.STOCK_DIVIDEND_POSITION_DATE_NON_SAVE ] = dict_per_stock_dividend_position_date[ str_trading_date ]
                else:
                    item[ TradingData.STOCK_DIVIDEND_POSITION_DATE_NON_SAVE ] = str_trading_date
                n_accumulated_inventory += n_stock_dividend_share_gain
                
                n_extra_insurance_fee_total = 0
                if b_extra_insurance_fee:
                    if int( Decimal( str(item[ TradingData.CUSTOM_EXTRA_INSURANCE_FEE ] ) ) ) != -1:
                        n_extra_insurance_fee_total = int( Decimal( str(item[ TradingData.CUSTOM_EXTRA_INSURANCE_FEE ] ) ) )
                    else:
                        if n_cash_dividend_gain + n_stock_dividend_value_gain > 20000:
                            n_extra_insurance_fee_for_cash_dividend = 0
                            n_extra_insurance_fee_for_stock_dividend = 0
                            if obj_trading_date.year >= 2013 and obj_trading_date.year < 2021:
                                n_extra_insurance_fee_for_cash_dividend = int( math.ceil( Decimal( str( n_cash_dividend_gain ) ) * Decimal( str( '0.0191' ) ) ) )
                                n_extra_insurance_fee_for_stock_dividend = int( math.ceil( Decimal( str( n_stock_dividend_value_gain ) ) * Decimal( str( '0.0191' ) ) ) )
                            elif obj_trading_date.year >= 2021:
                                n_extra_insurance_fee_for_cash_dividend = int( math.ceil( Decimal( str( n_cash_dividend_gain ) ) * Decimal( str( '0.0211' ) ) ) )
                                n_extra_insurance_fee_for_stock_dividend = int( math.ceil( Decimal( str( n_stock_dividend_value_gain ) ) * Decimal( str( '0.0211' ) ) ) )
                            n_extra_insurance_fee_total = n_extra_insurance_fee_for_cash_dividend + n_extra_insurance_fee_for_stock_dividend
                if n_cash_dividend_gain > n_dividend_transfer_fee:
                    item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = n_cash_dividend_gain
                    item[ TradingData.TRADING_FEE_NON_SAVE ] = n_dividend_transfer_fee
                    item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = n_extra_insurance_fee_total
                    n_accumulated_cost = n_accumulated_cost - n_cash_dividend_gain + n_dividend_transfer_fee + n_extra_insurance_fee_total
                    n_accumulated_cash_dividend = n_accumulated_cash_dividend + n_cash_dividend_gain - n_dividend_transfer_fee - n_extra_insurance_fee_total #要想一下要不要扣掉手續費
                else:
                    item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
                    item[ TradingData.TRADING_FEE_NON_SAVE ] = 0
                    item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0 
                n_accumulated_stock_dividend += n_stock_dividend_share_gain
                list_buying_data = [ 0, n_stock_dividend_share_gain, item[ TradingData.STOCK_DIVIDEND_POSITION_DATE_NON_SAVE ] ]
                queue_buying_data.append( list_buying_data )
                queue_buying_data = deque( sorted( queue_buying_data, key = lambda x: x[ 2 ] ) )
            elif e_trading_type == TradingType.CAPITAL_REDUCTION:
                if n_accumulated_inventory == 0: #沒有庫存就不用算減資了
                    continue
                item[ TradingData.PER_SHARE_TRADING_PRICE ] = -item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ]
                item[ TradingData.TRADING_QUANTITY ] = n_accumulated_inventory
                item[ TradingData.TRADING_VALUE_NON_SAVE ] = -int( n_accumulated_inventory * item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] )
                item[ TradingData.TRADING_FEE_NON_SAVE ] = 0
                item[ TradingData.TRADING_TAX_NON_SAVE ] = 0
                item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0 
                item[ TradingData.TRADING_COST_NON_SAVE ] = 0
                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
                if item[ TradingData.CAPITAL_REDUCTION_TYPE ] == CapitalReductionType.CASH_RETURN:
                    n_accumulated_cost = n_accumulated_cost - int( Decimal( str( n_accumulated_inventory ) ) * Decimal( str( item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] ) ) )
                    n_accumulated_cost_without_considering_dividend = n_accumulated_cost_without_considering_dividend - int( Decimal( str( n_accumulated_inventory ) ) * Decimal( str( item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] ) ) )
                    
                n_accumulated_inventory = int( Decimal( str( n_accumulated_inventory ) ) * ( Decimal( str( '10' ) ) - Decimal( str( item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] ) ) ) / Decimal( str( '10' ) ) )
                n_total_rest_buying_count = 0
                for list_buying_data in queue_buying_data:
                    list_buying_data[ 1 ]= int( Decimal( str( list_buying_data[ 1 ] ) ) * ( Decimal( str( '10' ) ) - Decimal( str( item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] ) ) ) / Decimal( str( '10' ) ) )
                    n_total_rest_buying_count += list_buying_data[ 1 ]
                if n_accumulated_inventory != n_total_rest_buying_count:
                    list_buying_data = queue_buying_data[ -1 ]
                    list_buying_data[ 1 ] += ( n_accumulated_inventory - n_total_rest_buying_count )
                    queue_buying_data[ -1 ]= list_buying_data           
            elif e_trading_type == TradingType.SPLIT:
                if n_accumulated_inventory == 0: #沒有庫存就不用算分割了
                    continue
                item[ TradingData.PER_SHARE_TRADING_PRICE ] = 0
                item[ TradingData.TRADING_QUANTITY ] = 0
                item[ TradingData.TRADING_VALUE_NON_SAVE ] = 0
                item[ TradingData.TRADING_FEE_NON_SAVE ] = 0
                item[ TradingData.TRADING_TAX_NON_SAVE ] = 0
                item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0 
                item[ TradingData.TRADING_COST_NON_SAVE ] = 0
                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
                n_accumulated_inventory = int( Decimal( str( n_accumulated_inventory ) ) * Decimal( str( item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] ) ) )
                for i in range ( len( queue_buying_data ) ):
                    queue_buying_data[ i ][ 1 ] = int( Decimal( str( queue_buying_data[ i ][ 1 ] ) ) * Decimal( str( item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] ) ) )
            elif e_trading_type == TradingType.MERGE:
                if n_accumulated_inventory == 0: #沒有庫存就不用算合併了
                    continue
                item[ TradingData.PER_SHARE_TRADING_PRICE ] = 0
                item[ TradingData.TRADING_QUANTITY ] = 0
                item[ TradingData.TRADING_VALUE_NON_SAVE ] = 0
                item[ TradingData.TRADING_FEE_NON_SAVE ] = 0
                item[ TradingData.TRADING_TAX_NON_SAVE ] = 0
                item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0 
                item[ TradingData.TRADING_COST_NON_SAVE ] = 0
                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
                n_accumulated_inventory = int( Decimal( str( n_accumulated_inventory ) ) / Decimal( str( item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] ) ) )
                for i in range ( len( queue_buying_data ) ):
                    queue_buying_data[ i ][ 1 ] = int( Decimal( str( queue_buying_data[ i ][ 1 ] ) ) / Decimal( str( item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] ) ) )
            item[ TradingData.ACCUMULATED_COST_NON_SAVE ] = n_accumulated_cost
            item[ TradingData.ACCUMULATED_COST_WITHOUT_CONSIDERING_DIVIDEND_NON_SAVE ] = n_accumulated_cost_without_considering_dividend
            item[ TradingData.ACCUMULATED_QUANTITY_NON_SAVE ] = n_accumulated_inventory
            item[ TradingData.ACCUMULATED_AVERAGE_COST_NON_SAVE ] = n_accumulated_cost / n_accumulated_inventory if n_accumulated_inventory != 0 else 0
            item[ TradingData.ACCUMULATED_AVERAGE_COST_WITHOUT_CONSIDERING_DIVIDEND_NON_SAVE ] = n_accumulated_cost_without_considering_dividend / n_accumulated_inventory if n_accumulated_inventory != 0 else 0
            item[ TradingData.ACCUMULATED_STOCK_DIVIDEND_GAIN_NON_SAVE ] = n_accumulated_stock_dividend
            item[ TradingData.ACCUMULATED_CASH_DIVIDEND_GAIN_NON_SAVE ] = n_accumulated_cash_dividend
            if index == len( sorted_list ) - 1:
                n_total_rest_buying_cost = 0
                n_total_rest_buying_count = 0
                for list_buying_data in queue_buying_data:
                    n_total_rest_buying_cost += list_buying_data[ 0 ]
                    n_total_rest_buying_count += list_buying_data[ 1 ]
                if n_total_rest_buying_count != 0:
                    item[ TradingData.CURRENT_BUYING_COST_NON_SAVE ] = n_total_rest_buying_cost
                else:
                    item[ TradingData.CURRENT_BUYING_COST_NON_SAVE ] = 0


            list_calibration_data.append( item )

        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ str_stock_number ] = list_calibration_data.copy()
        return list_calibration_data
    
    def process_all_transfer_data( self ): 
        for key_account_name, value_dict_per_company_transfer_data in self.dict_all_account_cash_transfer_data.items():
            self.process_single_transfer_data( key_account_name )

    def process_single_transfer_data( self, str_tab_widget_name ):
        list_transfer_data = self.dict_all_account_cash_transfer_data[ str_tab_widget_name ]
        if len( list_transfer_data ) == 0:
            return
        sorted_list = sorted( list_transfer_data, key=lambda x: ( datetime.datetime.strptime( x[ TransferData.TRANSFER_DATE ], "%Y-%m-%d") ) )

        n_total_value = 0
        for index, item in enumerate( sorted_list ):
            item[ TransferData.SORTED_INDEX_NON_SAVE ] = index
            e_transfer_type = item[ TransferData.TRANSFER_TYPE ]
            if e_transfer_type == TransferType.TRANSFER_IN:
                n_total_value += item[ TransferData.TRANSFER_VALUE ]
            elif e_transfer_type == TransferType.TRANSFER_OUT:
                n_total_value -= item[ TransferData.TRANSFER_VALUE ]
            
            item[ TransferData.TOTAL_VALUE_NON_SAVE ] = n_total_value

        self.dict_all_account_cash_transfer_data[ str_tab_widget_name ] = sorted_list

    def refresh_stock_list_table( self, clear_table = True ): 
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        qt_table_view = self.ui.qtTabWidget.currentWidget().findChild( QTableView, "StockListTableView" )
        qt_display_type_combobox = self.ui.qtTabWidget.currentWidget().findChild( QComboBox, "DisplayTypeComboBox")
        n_minimum_common_trading_fee = self.dict_all_account_general_data[ str_tab_widget_name ][ "minimum_common_trading_fee" ]
        e_decimal_round_type = self.dict_all_account_general_data[ str_tab_widget_name ][ "decimal_round_type" ]
        n_all_stock_total_cost = 0
        n_all_stock_current_total_market_value = 0 #目前總市值
        n_all_stock_current_total_market_value_after_fee_and_tax = 0 #目前總市值(扣掉手續費和稅金)
        n_all_stock_current_buying_cost = 0 #目前總成本
        n_all_stock_unrealized_profit = 0 #未實現獲利
        n_all_stock_accumulated_profit = 0 #累計獲利

        list_all_stock_trading_date = []
        list_all_stock_trading_flows = []
        list_all_stock_market_value = []
        if qt_table_view:
            qt_table_model = qt_table_view.model()
            if clear_table:
                qt_table_model.clear()

            qt_table_model.setHorizontalHeaderLabels( self.list_stock_list_table_horizontal_header )
            
            with QSignalBlocker( qt_table_view.horizontalHeader() ):
                list_vertical_labels = []
                obj_current_date = datetime.datetime.today()
                n_index_row = 0
                for key_stock_number, value_list_stock_trading_data in dict_per_account_all_stock_trading_data.items():
                    dict_trading_data_first = value_list_stock_trading_data[ 0 ] #取第一筆交易資料，因為第一筆交易資料有存是否使用自動帶入股利
                    dict_trading_data_last = value_list_stock_trading_data[ len( value_list_stock_trading_data ) - 1 ] #取最後一筆交易資料，因為最後一筆交易資料的庫存等內容才是所有累計的結果

                    n_accumulated_quantity = dict_trading_data_last.get( TradingData.ACCUMULATED_QUANTITY_NON_SAVE, 0 )
                    if qt_display_type_combobox.currentIndex() == 1 and n_accumulated_quantity == 0:#僅顯示目前有庫存的股票
                        continue
                    elif qt_display_type_combobox.currentIndex() == 2 and n_accumulated_quantity != 0:#僅顯示目前沒有庫存的股票
                        continue

                    list_vertical_labels.append( "   " )#Header用空白的字串，只是為了可以上下拖動
                    list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ key_stock_number ]
                    str_stock_name, str_b_etf = list_stock_name_and_type
                    b_bond = True if '債' in str_stock_name else False
                    b_etf = True if str_b_etf == "True" else False
                    b_suspend_company = key_stock_number in self.dict_all_suspend_company_number_to_name_and_type
                    if b_suspend_company:
                        str_stock_number_and_name = f"{key_stock_number} {str_stock_name} (已下市)"
                    else:
                        str_stock_number_and_name = f"{key_stock_number} {str_stock_name}"

                    #現股總成本
                    n_current_buying_cost = dict_trading_data_last.get( TradingData.CURRENT_BUYING_COST_NON_SAVE, 0 )
                    n_all_stock_current_buying_cost += n_current_buying_cost
                    if n_accumulated_quantity == 0:
                        f_current_average_cost = 0#現股成本
                        str_break_even_price = '-'
                    else:
                        f_current_average_cost = round( n_current_buying_cost / n_accumulated_quantity, 2 )#現股成本
                        f_break_even_price = Utility.find_min_break_even_price_from_net( n_current_buying_cost, n_accumulated_quantity, b_etf, b_bond, n_minimum_common_trading_fee )
                        str_break_even_price = format( f_break_even_price, ",.2f" )

                    #目前庫存
                    if self.ui.qtUse1ShareUnitAction.isChecked():
                        str_accumulated_quantity = format( n_accumulated_quantity, "," )
                    else:
                        f_accumulated_quantity_1000_share = n_accumulated_quantity / 1000
                        if f_accumulated_quantity_1000_share.is_integer():
                            str_accumulated_quantity = format( int( f_accumulated_quantity_1000_share ), "," )
                        else:
                            str_accumulated_quantity = format( f_accumulated_quantity_1000_share, "," )

                    #累計成本、累計平均成本
                    if self.ui.qtCostWithInDividendAction.isChecked():
                        n_per_stock_accumulated_cost = dict_trading_data_last.get( TradingData.ACCUMULATED_COST_NON_SAVE, 0 )
                        f_per_stock_accumulated_average_cost = round( dict_trading_data_last.get( TradingData.ACCUMULATED_AVERAGE_COST_NON_SAVE, 0 ), 3 )
                    else:
                        n_per_stock_accumulated_cost = dict_trading_data_last.get( TradingData.ACCUMULATED_COST_WITHOUT_CONSIDERING_DIVIDEND_NON_SAVE, 0 )
                        f_per_stock_accumulated_average_cost = round( dict_trading_data_last.get( TradingData.ACCUMULATED_AVERAGE_COST_WITHOUT_CONSIDERING_DIVIDEND_NON_SAVE, 0 ), 3 )
                    n_all_stock_total_cost += n_per_stock_accumulated_cost

                    n_per_stock_accumulated_stock_dividend = dict_trading_data_last.get( TradingData.ACCUMULATED_STOCK_DIVIDEND_GAIN_NON_SAVE, 0 )
                    n_per_stock_accumulated_cash_dividend = dict_trading_data_last.get( TradingData.ACCUMULATED_CASH_DIVIDEND_GAIN_NON_SAVE, 0 )

                    list_per_stock_trading_date = []
                    list_per_stock_trading_flows = []
                    n_per_stock_accumulated_trading_fee = 0
                    n_per_stock_accumulated_trading_tax = 0
                    n_per_stock_realized_profit = 0
                    b_invalid_data = False
                    for trading_data in value_list_stock_trading_data:
                        e_trading_type = trading_data[ TradingData.TRADING_TYPE ]
                        if e_trading_type == TradingType.TEMPLATE:
                            continue
                        if TradingData.ACCUMULATED_QUANTITY_NON_SAVE in trading_data and trading_data[ TradingData.ACCUMULATED_QUANTITY_NON_SAVE ] < 0:
                            b_invalid_data = True
                        n_per_stock_accumulated_trading_fee += trading_data[ TradingData.TRADING_FEE_NON_SAVE ]
                        n_per_stock_accumulated_trading_tax += trading_data[ TradingData.TRADING_TAX_NON_SAVE ]
                        n_per_stock_realized_profit += trading_data.get( TradingData.SELLING_PROFIT_NON_SAVE, 0 )

                        if key_stock_number in self.dict_all_company_number_to_price_info:
                            if ( e_trading_type == TradingType.BUY or
                                e_trading_type == TradingType.REGULAR_BUY or 
                                e_trading_type == TradingType.CAPITAL_INCREASE ):
                                n_cost = -trading_data[ TradingData.TRADING_COST_NON_SAVE ]
                            elif e_trading_type == TradingType.SELL:
                                n_cost = trading_data[ TradingData.TRADING_COST_NON_SAVE ]
                            elif e_trading_type == TradingType.DIVIDEND:
                                n_cost = trading_data[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] - trading_data[ TradingData.TRADING_FEE_NON_SAVE ] - trading_data[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] 
                            elif e_trading_type == TradingType.CAPITAL_REDUCTION:
                                #trading_data[ TradingData.TRADING_VALUE_NON_SAVE ]本身是負值，但因為是退款，所以在算XIRR時應該要用正數，因此取負號
                                n_cost = -trading_data[ TradingData.TRADING_VALUE_NON_SAVE ] 
                            
                            list_all_stock_trading_flows.append( n_cost )
                            list_per_stock_trading_flows.append( n_cost )

                            str_trading_date = trading_data[ TradingData.TRADING_DATE ]
                            obj_trading_date = datetime.datetime.strptime( str_trading_date, "%Y-%m-%d" )
                            list_all_stock_trading_date.append( obj_trading_date )
                            list_per_stock_trading_date.append( obj_trading_date )

                    n_per_stock_accumulated_dividend_profit = 0

                    b_valid_price = False
                    if key_stock_number in self.dict_all_company_number_to_price_info:
                        try:
                            f_stock_price = float( self.dict_all_company_number_to_price_info[ key_stock_number ] )
                            b_valid_price = True
                        except ValueError:
                            b_valid_price = False

                    if b_valid_price:
                        if key_stock_number in self.list_previous_day_data:
                            qt_stock_price_color = QBrush( '#777777' )
                        else:
                            qt_stock_price_color = QBrush( '#FFFFFF' )
                        str_stock_price = format( f_stock_price, "," )
                        n_per_stock_market_value = int( round( n_accumulated_quantity * f_stock_price, 0 ) )
                        str_per_stock_market_value = format( n_per_stock_market_value, "," )
                        list_all_stock_market_value.append( n_per_stock_market_value )

                        n_expect_trading_fee = 0
                        if n_accumulated_quantity > 0:
                            n_trading_fee = Utility.round_to( n_per_stock_market_value * Decimal( '0.001425' ), e_decimal_round_type )
                            n_expect_trading_fee = max( n_trading_fee, n_minimum_common_trading_fee )

                        if b_etf:
                            if b_bond:
                                n_expect_trading_tax = 0
                            else:
                                n_expect_trading_tax = int( n_per_stock_market_value * Decimal( '0.001' ) )
                        else:
                            n_expect_trading_tax = int( n_per_stock_market_value * Decimal( '0.003' ) )

                        n_per_stock_unrealized_profit = n_per_stock_market_value - n_expect_trading_fee - n_expect_trading_tax - n_current_buying_cost
                        n_all_stock_unrealized_profit += n_per_stock_unrealized_profit

                        str_per_stock_unrealized_profit = format( n_per_stock_unrealized_profit, "," )
                        if n_per_stock_market_value != 0:
                            if n_current_buying_cost == 0:
                                str_per_stock_unrealized_profit_ratio = "9999.99%"
                            else:
                                f_per_stock_current_profit_ratio = ( n_per_stock_unrealized_profit / n_current_buying_cost ) * 100
                                str_per_stock_unrealized_profit_ratio = format( f_per_stock_current_profit_ratio, ".2f" ) + "%"
                        else:
                            str_per_stock_unrealized_profit_ratio = "0%"

                        n_per_stock_accumulated_profit = n_per_stock_market_value - n_expect_trading_fee - n_expect_trading_tax - n_per_stock_accumulated_cost
                        n_all_stock_accumulated_profit += n_per_stock_accumulated_profit
                        n_all_stock_current_total_market_value += n_per_stock_market_value
                        n_all_stock_current_total_market_value_after_fee_and_tax += ( n_per_stock_market_value - n_expect_trading_fee - n_expect_trading_tax )
                        n_per_stock_accumulated_dividend_profit = int( n_per_stock_accumulated_stock_dividend * f_stock_price ) + n_per_stock_accumulated_cash_dividend
                        str_per_stock_accumulated_profit = format( n_per_stock_accumulated_profit, "," )
                        if n_accumulated_quantity > 0:
                            list_per_stock_trading_flows.append( n_per_stock_market_value - n_expect_trading_fee - n_expect_trading_tax )    
                            list_per_stock_trading_date.append( obj_current_date )

                        str_per_stock_xirr = "-"
                        if len( list_per_stock_trading_flows ) > 1:
                            try:
                                per_stock_xirr_result = Utility.xirr( list_per_stock_trading_flows, list_per_stock_trading_date )
                                str_per_stock_xirr = f"{per_stock_xirr_result:.3%}"
                            except ValueError as e:
                                str_per_stock_xirr = "-"
                    else:
                        list_all_stock_market_value.append( 0 )
                        str_stock_price = "N/A"
                        str_per_stock_market_value = "N/A"
                        str_per_stock_accumulated_profit = "N/A"
                        str_per_stock_unrealized_profit = "N/A"
                        qt_stock_price_color = QBrush( '#FFFFFF' )
                        str_per_stock_xirr = "-"
                        str_per_stock_unrealized_profit_ratio = "-"

                    qt_standard_item = QStandardItem( str_stock_number_and_name )
                    qt_standard_item.setTextAlignment( Qt.AlignLeft | Qt.AlignVCenter )
                    qt_standard_item.setFlags( qt_standard_item.flags() & ~Qt.ItemIsEditable )
                    qt_standard_item.setData( key_stock_number, Qt.UserRole )
                    if b_invalid_data:
                        qt_standard_item.setBackground( QBrush( '#FF0000' ) )
                    qt_table_model.setItem( n_index_row, 0, qt_standard_item ) 

                    for n_column, e_type in enumerate( self.list_show_stock_info ):
                        str_data = ""
                        qt_color = QBrush( '#FFFFFF' )
                        if e_type == StockInfoType.LATEST_PRICE:#收盤價
                            str_data = str_stock_price
                            qt_color = qt_stock_price_color
                        elif e_type == StockInfoType.QUANTITY:#庫存股數
                            str_data = str_accumulated_quantity
                        elif e_type == StockInfoType.LATEST_MARKET_VALUE:#現值
                            str_data = str_per_stock_market_value
                        elif e_type == StockInfoType.CURRENT_COST:#現股成本
                            str_data = format( f_current_average_cost, "," )
                        elif e_type == StockInfoType.UNREALIZED_PROFIT:#未實現損益
                            str_data = str_per_stock_unrealized_profit
                            qt_color = Utility.get_up_down_color( "0", str_per_stock_unrealized_profit )
                        elif e_type == StockInfoType.UNREALIZED_PROFIT_RATIO:#未實現報酬率
                            str_data = str_per_stock_unrealized_profit_ratio
                            qt_color = Utility.get_up_down_color( "0", str_per_stock_unrealized_profit_ratio )
                        elif e_type == StockInfoType.REALIZED_PROFIT:#已實現損益
                            str_data = format( n_per_stock_realized_profit, "," )
                            qt_color = Utility.get_up_down_color( "0", str_data )
                        elif e_type == StockInfoType.BREAK_EVEN_PRICE:#損益平衡價
                            str_data = str_break_even_price
                        elif e_type == StockInfoType.ACCUMULATED_COST:#累計成本
                            str_data = format( n_per_stock_accumulated_cost, "," )
                        elif e_type == StockInfoType.ACCUMULATED_AVERAGE_COST:#累計平均成本
                            str_data = format( f_per_stock_accumulated_average_cost, "," )
                        elif e_type == StockInfoType.ACCUMULATED_PROFIT:#累計損益
                            str_data = str_per_stock_accumulated_profit
                            qt_color = Utility.get_up_down_color( "0", str_per_stock_accumulated_profit )
                        elif e_type == StockInfoType.ACCUMULATED_TRADING_FEE:#累計手續費
                            str_data = format( n_per_stock_accumulated_trading_fee, "," )
                        elif e_type == StockInfoType.ACCUMULATED_TAX:#累計交易稅
                            str_data = format( n_per_stock_accumulated_trading_tax, "," )
                        elif e_type == StockInfoType.ACCUMULATED_DIVIDEND_INCOME:#累計股利所得
                            str_data = format( n_per_stock_accumulated_dividend_profit, "," )
                        elif e_type == StockInfoType.XIRR_VALUE:#平均年化報酬率
                            str_data = str_per_stock_xirr
                            qt_color = Utility.get_up_down_color( "0", str_per_stock_xirr )
                        elif e_type == StockInfoType.HOLDING_MARKET_RATIO:#持股淨值比
                            pass

                        qt_standard_item = QStandardItem( str_data )
                        qt_standard_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
                        qt_standard_item.setFlags( qt_standard_item.flags() & ~Qt.ItemIsEditable )
                        qt_standard_item.setData( key_stock_number, Qt.UserRole )
                        qt_standard_item.setForeground( qt_color )
                        qt_table_model.setItem( n_index_row, n_column + 1, qt_standard_item ) 

                    qt_use_auto_dividend_item = QStandardItem()
                    if dict_trading_data_first[ TradingData.USE_AUTO_DIVIDEND_DATA ] == AutoDividendType.AUTO:
                        qt_use_auto_dividend_item.setIcon( share_icon.get_icon( share_icon.IconType.STAR_SOLID ) )
                        qt_use_auto_dividend_item.setText( "自動" )
                    elif dict_trading_data_first[ TradingData.USE_AUTO_DIVIDEND_DATA ] == AutoDividendType.MANUAL:
                        qt_use_auto_dividend_item.setIcon( share_icon.get_icon( share_icon.IconType.STAR_DASHED ) )
                        qt_use_auto_dividend_item.setText( "手動" )
                    elif dict_trading_data_first[ TradingData.USE_AUTO_DIVIDEND_DATA ] == AutoDividendType.MANUAL_WITH_AUTO:
                        qt_use_auto_dividend_item.setIcon( share_icon.get_icon( share_icon.IconType.STAR_HALF_DASHED ) )
                        qt_use_auto_dividend_item.setText( "混和" )

                    qt_use_auto_dividend_item.setFlags( qt_use_auto_dividend_item.flags() & ~Qt.ItemIsEditable )
                    qt_table_model.setItem( n_index_row, len( self.list_stock_list_table_horizontal_header ) - 3, qt_use_auto_dividend_item)
                        
                    qt_export_icon_item = QStandardItem("")
                    qt_export_icon_item.setIcon( share_icon.get_icon( share_icon.IconType.EXPORT ) )
                    qt_export_icon_item.setFlags( qt_export_icon_item.flags() & ~Qt.ItemIsEditable )
                    qt_table_model.setItem( n_index_row, len( self.list_stock_list_table_horizontal_header ) - 2, qt_export_icon_item )
                    qt_delete_icon_item = QStandardItem("")
                    qt_delete_icon_item.setIcon( share_icon.get_icon( share_icon.IconType.DELETE ) )
                    qt_delete_icon_item.setFlags( qt_delete_icon_item.flags() & ~Qt.ItemIsEditable )
                    qt_table_model.setItem( n_index_row, len( self.list_stock_list_table_horizontal_header ) - 1, qt_delete_icon_item )
                    n_index_row += 1

                #設定動態顯示欄位的欄寬
                for n_column, e_type in enumerate( self.list_show_stock_info ):
                    n_index = int( e_type.value )
                    if n_index < len( self.list_stock_list_column_width ):
                        qt_table_view.setColumnWidth( n_column + 1, self.list_stock_list_column_width[ n_index ] )
                    else:
                        assert False, f"不該執行到這裡！"
                
                #設定固定顯示欄位的欄寬
                for n_column in range( len( self.list_stock_list_table_horizontal_header ) ):
                    if ( n_column == 0 or #股票代碼
                         n_column == len( self.list_stock_list_table_horizontal_header ) - 3 or # 自動帶入股利
                         n_column == len( self.list_stock_list_table_horizontal_header ) - 2 or # 匯出
                         n_column == len( self.list_stock_list_table_horizontal_header ) - 1 ): # 刪除
                        qt_table_view.setColumnWidth( n_column, self.list_stock_list_column_width[ n_column ] )

                for n_index_row in range( len( dict_per_account_all_stock_trading_data ) ):
                    qt_table_view.setRowHeight( n_index_row, 25 )
                qt_table_model.setVerticalHeaderLabels( list_vertical_labels )
        for n_index_row, n_per_market_value in enumerate( list_all_stock_market_value ):
            f_holding_ratio = n_per_market_value / n_all_stock_current_total_market_value if n_all_stock_current_total_market_value != 0 else 0
            f_holding_ratio = f_holding_ratio * 100
            str_holding_ratio = format( f_holding_ratio, ".1f" ) + "%"
            for n_column, e_type in enumerate( self.list_show_stock_info ):
                str_data = ""
                qt_color = QBrush( '#FFFFFF' )
                if e_type == StockInfoType.HOLDING_MARKET_RATIO:#持股淨值比
                    qt_standard_item = QStandardItem( str_holding_ratio )
                    qt_standard_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
                    qt_standard_item.setFlags( qt_standard_item.flags() & ~Qt.ItemIsEditable )
                    qt_standard_item.setData( key_stock_number, Qt.UserRole )
                    qt_standard_item.setForeground( qt_color )
                    qt_table_model.setItem( n_index_row, n_column + 1, qt_standard_item ) 
        f_unrealized_profit_ratio = n_all_stock_unrealized_profit / n_all_stock_current_buying_cost if n_all_stock_current_buying_cost != 0 else 0
        f_unrealized_profit_ratio = f_unrealized_profit_ratio * 100

        str_total_quantity_value = format( n_all_stock_current_total_market_value, "," )
        str_current_cost_value = format( n_all_stock_current_buying_cost, "," )
        str_unrealized_profit_value = format( n_all_stock_unrealized_profit, "," )
        str_unrealized_profit_ratio_value = format( f_unrealized_profit_ratio, ".2f" ) + "%"
        str_accumulated_cost_value = format( n_all_stock_total_cost, "," )
        str_accumulated_profit_value = format( n_all_stock_accumulated_profit, "," )
        str_xirr_value = "-"

        list_all_stock_trading_flows.append( n_all_stock_current_total_market_value_after_fee_and_tax )
        list_all_stock_trading_date.append( obj_current_date )
        if len( list_all_stock_trading_flows ) > 1:
            try:
                per_stock_xirr_result = Utility.xirr( list_all_stock_trading_flows, list_all_stock_trading_date )
                str_xirr_value = f"{per_stock_xirr_result:.3%}"
            except ValueError as e:
                pass
        
        if self.ui.qtCostWithInDividendAction.isChecked():
            str_extra_info = '(含息)'
        else:
            str_extra_info = '(不含息)'
        str_total_info = (
            f'<span style="color:#878787;">市值:</span> <span style="color:white;">{str_total_quantity_value}</span> &nbsp;&nbsp; '
            f'<span style="color:#878787;">現股成本:</span> <span style="color:white;">{str_current_cost_value}</span> &nbsp;&nbsp; '
            f'<span style="color:#878787;">未實現損益:</span> {Utility.color_text_by_value(str_unrealized_profit_value)} &nbsp;&nbsp; '
            f'<span style="color:#878787;">未實現報酬率:</span> {Utility.color_text_by_value(str_unrealized_profit_ratio_value)}<br>'
            f'<span style="color:#878787;">累計成本{str_extra_info}:</span> <span style="color:white;">{str_accumulated_cost_value}</span> &nbsp;&nbsp; '
            f'<span style="color:#878787;">累計損益{str_extra_info}:</span> {Utility.color_text_by_value(str_accumulated_profit_value)} &nbsp;&nbsp; '
            f'<span style="color:#878787;">年化報酬率:</span> {Utility.color_text_by_value(str_xirr_value)}'
        )

        qt_total_info_value_label = self.ui.qtTabWidget.currentWidget().findChild(QLabel, "TotalInfoLabel")
        qt_total_info_value_label.setText( str_total_info )
        qt_total_info_value_label.setTextFormat( Qt.RichText )

    def get_stock_list_horizontal_header( self ):
        list_header = []
        list_header.append( '股票代碼及名稱' )
        for e_type in self.list_show_stock_info:
            if e_type == StockInfoType.QUANTITY:
                if self.ui.qtUse1ShareUnitAction.isChecked():
                    list_header.append( "庫存股數" )
                else:
                    list_header.append( "庫存張數" )
            elif e_type == StockInfoType.LATEST_PRICE:
                list_header.append( self.str_latest_price_column_header )
            elif e_type == StockInfoType.ACCUMULATED_COST:
                if self.ui.qtCostWithInDividendAction.isChecked():
                    list_header.append( "累計成本\n(含息)" )
                else:
                    list_header.append( "累計成本\n(不含息)" )
            elif e_type == StockInfoType.ACCUMULATED_AVERAGE_COST:
                if self.ui.qtCostWithInDividendAction.isChecked():
                    list_header.append( "累計平均成本\n(含息)" )
                else:
                    list_header.append( "累計平均成本\n(不含息)" )
            elif e_type == StockInfoType.ACCUMULATED_PROFIT:
                if self.ui.qtCostWithInDividendAction.isChecked():
                    list_header.append( "累計損益\n(含息)" )
                else:
                    list_header.append( "累計損益\n(不含息)" )
            else:
                list_header.append( g_dict_stock_info[ e_type ] )

        list_header.append( '自動帶入股利' )
        list_header.append( '匯出' )
        list_header.append( '刪除' )
        return list_header

    def refresh_transfer_data_table( self ):
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        list_per_account_all_cash_transfer_data = self.dict_all_account_cash_transfer_data[ str_tab_widget_name ]

        table_view = self.ui.qtTabWidget.currentWidget().findChild( QTableView, "CashTransferTableView" )
        if table_view:
            table_model = table_view.model()
            table_model.clear()
            table_model.setVerticalHeaderLabels( self.get_cash_transfer_header() )

            if self.ui.qtFromNewToOldAction.isChecked():
                loop_list = list_per_account_all_cash_transfer_data[::-1]
            else:
                loop_list = list_per_account_all_cash_transfer_data

            for column, item in enumerate( loop_list ):
                str_date = item[ TransferData.TRANSFER_DATE ]
                str_year = str_date.split( '-' )[ 0 ]
                if self.ui.qtROCYearAction.isChecked():
                    str_year = str( int( str_year ) - 1911 )

                str_month_date = str_date[ 5: ].replace( '-', '/' )
                str_date = f"{str_year}/{str_month_date}"


                str_total_value = format( item[ TransferData.TOTAL_VALUE_NON_SAVE ], "," )

                e_transfer_type = item[ TransferData.TRANSFER_TYPE ]
                if e_transfer_type == TransferType.TRANSFER_IN:
                    str_color = QBrush( '#FF0000' )
                    str_transfer_value = format( item[ TransferData.TRANSFER_VALUE ], "," )
                else:
                    str_color = QBrush( '#00AA00' )
                    str_transfer_value = format( -item[ TransferData.TRANSFER_VALUE ], "," )

                date_standard_item = QStandardItem( str_date )
                date_standard_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
                date_standard_item.setFlags( date_standard_item.flags() & ~Qt.ItemIsEditable )
                table_model.setItem( 0, column, date_standard_item )

                transfer_value_standard_item = QStandardItem( str_transfer_value )
                transfer_value_standard_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
                transfer_value_standard_item.setFlags( transfer_value_standard_item.flags() & ~Qt.ItemIsEditable )
                transfer_value_standard_item.setForeground( QBrush( str_color ) )
                table_model.setItem( 1, column, transfer_value_standard_item )

                total_value_standard_item = QStandardItem( str_total_value )
                total_value_standard_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
                total_value_standard_item.setFlags( total_value_standard_item.flags() & ~Qt.ItemIsEditable )
                table_model.setItem( 2, column, total_value_standard_item )

                edit_icon_item = QStandardItem("")
                edit_icon_item.setIcon( share_icon.get_icon( share_icon.IconType.EDIT ) )
                edit_icon_item.setFlags( edit_icon_item.flags() & ~Qt.ItemIsEditable )
                edit_icon_item.setData( item[ TransferData.SORTED_INDEX_NON_SAVE ], Qt.UserRole )
                delete_icon_item = QStandardItem("")
                delete_icon_item.setIcon( share_icon.get_icon( share_icon.IconType.DELETE ) )
                delete_icon_item.setFlags( delete_icon_item.flags() & ~Qt.ItemIsEditable )
                delete_icon_item.setData( item[ TransferData.SORTED_INDEX_NON_SAVE ], Qt.UserRole )

                table_model.setItem( 3, column, edit_icon_item )
                table_model.setItem( 4, column, delete_icon_item )

                if self.ui.qtShow10Action.isChecked():
                    if column == 9:
                        break

    def clear_per_account_transfer_table( self ):
        self.per_stock_trading_data_model.clear()
        self.per_stock_trading_data_model.setVerticalHeaderLabels( self.get_trading_data_header() )

    def refresh_trading_data_table( self, sorted_list, target_trading_data = None ):
        self.clear_per_stock_trading_table()
        self.per_stock_trading_data_model.setVerticalHeaderLabels( self.get_trading_data_header() )
        n_invalid_data_index = -1
        for index, dict_per_trading_data in enumerate( sorted_list ):
            if TradingData.ACCUMULATED_QUANTITY_NON_SAVE in dict_per_trading_data and dict_per_trading_data[ TradingData.ACCUMULATED_QUANTITY_NON_SAVE ] < 0:
                n_invalid_data_index = index
                break
        if self.ui.qtFromNewToOldAction.isChecked():
            loop_list = sorted_list[::-1]
        else:
            loop_list = sorted_list

        e_auto_dividend_type = sorted_list[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ]
        column = 0
        n_scroll_column = -1

        for index, dict_per_trading_data in enumerate( loop_list ):
            e_trading_type = dict_per_trading_data[ TradingData.TRADING_TYPE ]
            if e_trading_type == TradingType.TEMPLATE:
                continue

            if e_trading_type == TradingType.DIVIDEND:
                if e_auto_dividend_type == AutoDividendType.AUTO:#使用自動帶入股利資料，但是這筆資料不是自動股利資料，就跳過，反之亦然
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE not in dict_per_trading_data or not dict_per_trading_data[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]:
                        continue
                elif e_auto_dividend_type == AutoDividendType.MANUAL:
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE in dict_per_trading_data and dict_per_trading_data[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]:
                        continue

            list_data = self.get_per_trading_data_text_list( dict_per_trading_data )

            if TradingData.IS_REALLY_DAYING_TRADING_NON_SAVE in dict_per_trading_data:
                if dict_per_trading_data[ TradingData.IS_REALLY_DAYING_TRADING_NON_SAVE ]:
                    list_data[ 1 ] = "當沖賣"

            for row, data in enumerate( list_data ):
                standard_item = QStandardItem( data )
                if data == "買進":
                    standard_item.setBackground( QBrush( '#DA9694' ) )
                elif data == "定期定額買進":
                    standard_item.setBackground( QBrush( '#FF99FF' ) )
                elif data == "賣出" or data == "當沖賣":
                    standard_item.setBackground( QBrush( '#76933C' ) )
                elif data == "股利分配":
                    standard_item.setBackground( QBrush( '#8DB4E2' ) )
                elif data == "現金減資" or data == "虧損減資":
                    standard_item.setBackground( QBrush( '#B1A0C7' ) )
                elif data == "增資":
                    standard_item.setBackground( QBrush( '#FABF8F' ) )
                elif data == "股票分割":
                    standard_item.setBackground( QBrush( '#732BF5' ) )
                elif data == "股票合併":
                    standard_item.setBackground( QBrush( '#EF88BE' ) )
                else:
                    if n_invalid_data_index != -1:
                        if self.ui.qtFromNewToOldAction.isChecked():
                            if index < len( loop_list ) - n_invalid_data_index:
                                standard_item.setBackground( QBrush( '#FF0000' ) )
                        else:
                            if index >= n_invalid_data_index:
                                standard_item.setBackground( QBrush( '#FF0000' ) )

                if e_trading_type == TradingType.DIVIDEND and dict_per_trading_data[ TradingData.STOCK_DIVIDEND ] != 0:
                    str_dividend_tooltip = f"股票股利入帳日：{ dict_per_trading_data[ TradingData.STOCK_DIVIDEND_POSITION_DATE_NON_SAVE ] }\n可使用滑鼠右鍵設定"
                    standard_item.setData( str_dividend_tooltip, Qt.ToolTipRole )
                
                standard_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
                standard_item.setFlags( standard_item.flags() & ~Qt.ItemIsEditable )
                standard_item.setData( dict_per_trading_data[ TradingData.SORTED_INDEX_NON_SAVE ], Qt.UserRole )
                standard_item.setData( e_trading_type, Qt.UserRole + 1 )
                self.per_stock_trading_data_model.setItem( row, column, standard_item ) 

            if ( e_trading_type != TradingType.DIVIDEND or
                 ( TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE not in dict_per_trading_data or not dict_per_trading_data[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE] ) ): 
                edit_icon_item = QStandardItem("")
                edit_icon_item.setIcon( share_icon.get_icon( share_icon.IconType.EDIT ) )
                edit_icon_item.setFlags( edit_icon_item.flags() & ~Qt.ItemIsEditable )
                edit_icon_item.setData( dict_per_trading_data[ TradingData.SORTED_INDEX_NON_SAVE ], Qt.UserRole )
                delete_icon_item = QStandardItem("")
                delete_icon_item.setIcon( share_icon.get_icon( share_icon.IconType.DELETE ) )
                delete_icon_item.setFlags( delete_icon_item.flags() & ~Qt.ItemIsEditable )
                delete_icon_item.setData( dict_per_trading_data[ TradingData.SORTED_INDEX_NON_SAVE ], Qt.UserRole )

                self.per_stock_trading_data_model.setItem( len( list_data ), column, edit_icon_item )
                self.per_stock_trading_data_model.setItem( len( list_data ) + 1, column, delete_icon_item )



            if target_trading_data:
                if ( dict_per_trading_data[ TradingData.TRADING_DATE ] == target_trading_data[ TradingData.TRADING_DATE ] and 
                     dict_per_trading_data[ TradingData.TRADING_TYPE ] == target_trading_data[ TradingData.TRADING_TYPE ] and 
                     dict_per_trading_data[ TradingData.TOTAL_TRADING_PRICE ] == target_trading_data[ TradingData.TOTAL_TRADING_PRICE ] and 
                     dict_per_trading_data[ TradingData.STOCK_DIVIDEND ] == target_trading_data[ TradingData.STOCK_DIVIDEND ] and 
                     dict_per_trading_data[ TradingData.CASH_DIVIDEND ] == target_trading_data[ TradingData.CASH_DIVIDEND ] and 
                     dict_per_trading_data[ TradingData.TRADING_QUANTITY ] == target_trading_data[ TradingData.TRADING_QUANTITY ] and 
                     dict_per_trading_data[ TradingData.PER_SHARE_TRADING_PRICE ] == target_trading_data[ TradingData.PER_SHARE_TRADING_PRICE ] ):
                    n_scroll_column = column
            column += 1
            if self.ui.qtShow10Action.isChecked():
                if column == 10:
                    break

        list_horizontal_header = [ "" ] * column
        self.per_stock_trading_data_model.setHorizontalHeaderLabels( list_horizontal_header )
        
        for row in range( len( self.get_trading_data_header() ) ):
            if row == 7 or row == 8:
                self.ui.qtTradingDataTableView.setRowHeight( row, 40 )
            else:
                self.ui.qtTradingDataTableView.setRowHeight( row, 25 )

        if n_scroll_column != -1:
            # scrollbar = self.ui.qtTradingDataTableView.horizontalScrollBar()
            # column_x_position = self.ui.qtTradingDataTableView.columnViewportPosition( n_scroll_column - 1 )
            # scrollbar.setValue( column_x_position )
            self.ui.qtTradingDataTableView.scrollTo( self.ui.qtTradingDataTableView.model().index( 0, n_scroll_column ) )
            selection_model = self.ui.qtTradingDataTableView.selectionModel()
            selection_model.clearSelection()
            index1 = self.per_stock_trading_data_model.index( 0, n_scroll_column )
            selection = QItemSelection()
            selection.select( index1, index1 )
            selection_model.select( selection, QItemSelectionModel.Select )
            self.ui.qtTradingDataTableView.setFocus()

    def show_context_menu( self, position: QPoint ):
        index = self.ui.qtTradingDataTableView.indexAt( position )
        if not index.isValid():
            return

        e_trading_type = index.data( Qt.UserRole + 1 )
        if e_trading_type == TradingType.DIVIDEND:
            if self.str_picked_stock_number is None:
                return
            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
            dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
            list_trading_data = dict_per_account_all_stock_trading_data[ self.str_picked_stock_number ]

            # hidden_data = self.per_stock_trading_data_model.data( index, Qt.UserRole )
            hidden_data = index.data( Qt.UserRole )
            n_findindex = -1
            for index, dict_selected_data in enumerate( list_trading_data ):
                if dict_selected_data[ TradingData.SORTED_INDEX_NON_SAVE ] == hidden_data:
                    n_findindex = index
                    break
            if n_findindex == -1:
                return
            dict_selected_data = list_trading_data[ n_findindex ]
            f_stock_dividend = Decimal( str( dict_selected_data[ TradingData.STOCK_DIVIDEND ] ) )
            if f_stock_dividend == 0:
                return

            qt_menu = QMenu()
            qt_edit_position_date_action = qt_menu.addAction( "編輯股票股利入帳日" )
            exe_action = qt_menu.exec( self.ui.qtTradingDataTableView.viewport().mapToGlobal( position ) )
            if exe_action == qt_edit_position_date_action:

                str_stock_dividend_position_date = dict_selected_data[ TradingData.STOCK_DIVIDEND_POSITION_DATE_NON_SAVE ] #股票股利入帳日
                obj_stock_dividend_position_date = datetime.datetime.strptime( str_stock_dividend_position_date, "%Y-%m-%d" )

                str_stock_number = self.str_picked_stock_number
                list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
                str_stock_name = list_stock_name_and_type[ 0 ]

                dialog = StockDividendPositionDateEditDialog( str_stock_number, str_stock_name, obj_stock_dividend_position_date, self )
                if dialog.exec():
                    str_stock_dividend_trading_date = dict_selected_data[ TradingData.TRADING_DATE ] #股利除權息日
                    if str_stock_number not in self.dict_all_stock_dividend_position_date:
                        self.dict_all_stock_dividend_position_date[ str_stock_number ] = {}
                    self.dict_all_stock_dividend_position_date[ str_stock_number ][ str_stock_dividend_trading_date ] = dialog.str_position_date
                    self.save_stock_dividend_position_date()
                    sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
                    self.refresh_stock_list_table()
                    self.refresh_trading_data_table( sorted_list, dict_selected_data )

    def clear_per_stock_trading_table( self ):
        self.per_stock_trading_data_model.clear()
        self.per_stock_trading_data_model.setVerticalHeaderLabels( self.get_trading_data_header() )
        for row in range( len( self.get_trading_data_header() ) ):
            if row == 7 or row == 8:
                self.ui.qtTradingDataTableView.setRowHeight( row, 40 )
            else:
                self.ui.qtTradingDataTableView.setRowHeight( row, 25 )

    def get_cash_transfer_header( self ):
        return [ '日期', '入金/出金', '餘額', '編輯', '刪除' ]

    def get_trading_data_header( self ):
        list_vertical_header = ['日期', # 0
                                '交易種類', # 1 
                                '交易價格', # 2
                                '交易股數', # 3
                                '交易金額', # 4
                                '手續費 / 交易稅', # 5
                                '應收付', # 6
                                '全部股票股利(股) /\n每股股票股利(元)', # 7
                                '全部現金股利(元) /\n每股現金股利(元)', # 8
                                '補充保費', # 9
                                '累計成本', # 10
                                '庫存股數', # 11
                                '累計平均成本', # 12
                                '單筆損益', # 13
                                '編輯', # 14
                                '刪除' ]# 15
        if self.ui.qtUse1ShareUnitAction.isChecked():
            list_vertical_header[ 3 ] = '交易股數'
            list_vertical_header[ 7 ] = '全部股票股利(股) /\n每股股票股利(元)'
            list_vertical_header[ 11 ] = '庫存股數'
        else:
            list_vertical_header[ 3 ] = '交易張數'
            list_vertical_header[ 7 ] = '全部股票股利(張) /\n每股股票股利(元)'
            list_vertical_header[ 11 ] = '庫存張數'
        if self.ui.qtCostWithInDividendAction.isChecked():
            list_vertical_header[ 10 ] = '累計成本(含息)'
            list_vertical_header[ 12 ] = '累計平均成本(含息)'
        else:
            list_vertical_header[ 10 ] = '累計成本(不含息)'
            list_vertical_header[ 12 ] = '累計平均成本(不含息)'

        return list_vertical_header

    def get_per_trading_data_text_list( self, dict_per_trading_data ):
        e_trading_type = dict_per_trading_data[ TradingData.TRADING_TYPE ]
        if e_trading_type == TradingType.TEMPLATE:
            return []
        str_date = dict_per_trading_data[ TradingData.TRADING_DATE ]
        str_year = str_date.split( '-' )[ 0 ]
        if self.ui.qtROCYearAction.isChecked():
            str_year = str( int( str_year ) - 1911 )

        str_month_date = str_date[ 5: ]#.replace( '-', '/' )
        obj_date = datetime.datetime.strptime( str_date, "%Y-%m-%d" )
        n_weekday = obj_date.weekday()
        str_weekday = share_api.get_obj_datetime_weekday_text( n_weekday )

        f_trading_price = dict_per_trading_data[ TradingData.PER_SHARE_TRADING_PRICE ]
        n_trading_quantity = dict_per_trading_data[ TradingData.TRADING_QUANTITY ]
        n_trading_value = dict_per_trading_data[ TradingData.TRADING_VALUE_NON_SAVE ]
        n_trading_fee = dict_per_trading_data[ TradingData.TRADING_FEE_NON_SAVE ]
        n_trading_tax = dict_per_trading_data[ TradingData.TRADING_TAX_NON_SAVE ]
        n_extra_insurance_fee = dict_per_trading_data[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ]
        n_per_trading_total_cost = dict_per_trading_data[ TradingData.TRADING_COST_NON_SAVE ]
        e_dividend_value_type = dict_per_trading_data[ TradingData.DIVIDEND_VALUE_TYPE ]
        n_accumulated_quantity = dict_per_trading_data[ TradingData.ACCUMULATED_QUANTITY_NON_SAVE ]
        if e_dividend_value_type == DividendValueType.PER_SHARE:
            f_stock_dividend_per_share = dict_per_trading_data[ TradingData.STOCK_DIVIDEND ]
            f_cash_dividend_per_share = dict_per_trading_data[ TradingData.CASH_DIVIDEND ]
        else:
            n_stock_dividend_total = dict_per_trading_data[ TradingData.STOCK_DIVIDEND ]
            n_cash_dividend_total = dict_per_trading_data[ TradingData.CASH_DIVIDEND ]
            n_ori_accumulated_quantity = n_accumulated_quantity - n_stock_dividend_total

            f_stock_dividend_per_share = float( n_stock_dividend_total / n_ori_accumulated_quantity * 10 ) 
            f_cash_dividend_per_share = float( n_cash_dividend_total / n_ori_accumulated_quantity )
        n_stock_dividend_gain = dict_per_trading_data[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ]
        n_cash_dividend_gain = dict_per_trading_data[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ]
        if self.ui.qtCostWithInDividendAction.isChecked():
            n_accumulated_cost = dict_per_trading_data[ TradingData.ACCUMULATED_COST_NON_SAVE ]
            f_accumulated_average_cost = round( dict_per_trading_data[ TradingData.ACCUMULATED_AVERAGE_COST_NON_SAVE ], 3 )
        else:
            n_accumulated_cost = dict_per_trading_data[ TradingData.ACCUMULATED_COST_WITHOUT_CONSIDERING_DIVIDEND_NON_SAVE ]
            f_accumulated_average_cost = round( dict_per_trading_data[ TradingData.ACCUMULATED_AVERAGE_COST_WITHOUT_CONSIDERING_DIVIDEND_NON_SAVE ], 3 )
        if self.ui.qtUse1ShareUnitAction.isChecked():
            str_trading_quantity = format( n_trading_quantity, "," )
            str_stock_dividend_gain = format( n_stock_dividend_gain, "," )
            str_accumulated_quantity = format( n_accumulated_quantity, "," )
        else:
            f_trading_count = n_trading_quantity / 1000
            f_stock_dividend_gain = n_stock_dividend_gain / 1000
            f_accumulated_inventory = n_accumulated_quantity / 1000
            if f_trading_count.is_integer():
                str_trading_quantity = format( int( f_trading_count ), "," )
            else:
                str_trading_quantity = format( f_trading_count, "," )
            if f_stock_dividend_gain.is_integer():
                str_stock_dividend_gain = format( int( f_stock_dividend_gain ), "," )
            else:
                str_stock_dividend_gain = format( f_stock_dividend_gain, "," )
            if f_accumulated_inventory.is_integer():
                str_accumulated_quantity = format( int( f_accumulated_inventory ), "," )
            else:
                str_accumulated_quantity = format( f_accumulated_inventory, "," )

        str_trading_price = format( f_trading_price, "," )
        str_trading_value = format( n_trading_value, "," )
        str_trading_fee = format( n_trading_fee, "," )
        str_trading_tax = format( n_trading_tax, "," )
        str_extra_insurance_fee = format( n_extra_insurance_fee, "," )
        str_per_trading_total_cost = format( n_per_trading_total_cost, "," )
        str_stock_dividend = str_stock_dividend_gain + ' / ' + str( f_stock_dividend_per_share )
        str_cash_dividend = format( n_cash_dividend_gain, "," ) + ' / ' + str( f_cash_dividend_per_share )
        str_accumulated_cost = format( n_accumulated_cost, "," )
        str_accumulated_average_cost = format( f_accumulated_average_cost, "," )

        str_selling_profit = "N/A"
        if e_trading_type == TradingType.BUY:
            n_per_trading_total_cost = -n_per_trading_total_cost
            str_per_trading_total_cost = format( n_per_trading_total_cost, "," )
            str_trading_type = "買進"
            str_trading_tax = "0"
            str_extra_insurance_fee = "N/A"
            str_stock_dividend = "N/A"
            str_cash_dividend = "N/A"
        elif e_trading_type == TradingType.REGULAR_BUY:
            n_per_trading_total_cost = -n_per_trading_total_cost
            str_per_trading_total_cost = format( n_per_trading_total_cost, "," )
            str_trading_type = "定期定額買進"
            str_trading_tax = "0"
            str_extra_insurance_fee = "N/A"
            str_stock_dividend = "N/A"
            str_cash_dividend = "N/A"
        elif e_trading_type == TradingType.SELL:
            str_trading_type = "賣出"
            str_extra_insurance_fee = "N/A"
            str_stock_dividend = "N/A"
            str_cash_dividend = "N/A"
            str_selling_profit = format( dict_per_trading_data[ TradingData.SELLING_PROFIT_NON_SAVE ], "," )
        elif e_trading_type == TradingType.CAPITAL_INCREASE:
            n_per_trading_total_cost = -n_per_trading_total_cost
            str_per_trading_total_cost = format( n_per_trading_total_cost, "," )
            str_trading_type = "增資"
            str_trading_fee = "0"
            str_trading_tax = "0"
            str_extra_insurance_fee = "N/A"
            str_stock_dividend = "N/A"
            str_cash_dividend = "N/A"
        elif e_trading_type == TradingType.DIVIDEND:
            str_trading_type = "股利分配"
            str_trading_price = "N/A"
            str_trading_quantity = "N/A"
            str_trading_value = "N/A"
            str_trading_tax = "0"
            str_per_trading_total_cost = "N/A"
        elif e_trading_type == TradingType.CAPITAL_REDUCTION:
            if dict_per_trading_data[ TradingData.CAPITAL_REDUCTION_TYPE ] == CapitalReductionType.CASH_RETURN:
                str_trading_type = "現金減資"
            else:
                str_trading_type = "虧損減資"
            str_trading_fee = "0"
            str_trading_tax = "0"
            str_extra_insurance_fee = "N/A"
            str_per_trading_total_cost = "N/A"
            str_stock_dividend = "N/A"
            str_cash_dividend = "N/A"
        elif e_trading_type == TradingType.SPLIT:
            str_trading_type = "股票分割"
            str_trading_price = "N/A"
            n_split_value = dict_per_trading_data[ TradingData.CAPITAL_REDUCTION_PER_SHARE ]
            str_trading_quantity = "1分" + str( n_split_value ) #
            str_trading_value = "N/A"
            str_trading_fee = "0"
            str_trading_tax = "0"
            str_extra_insurance_fee = "N/A"
            str_per_trading_total_cost = "N/A"
            str_stock_dividend = "N/A"
            str_cash_dividend = "N/A"
        elif e_trading_type == TradingType.MERGE:
            str_trading_type = "股票合併"
            str_trading_price = "N/A"
            n_split_value = dict_per_trading_data[ TradingData.CAPITAL_REDUCTION_PER_SHARE ]
            str_trading_quantity = str( n_split_value ) + "合1" #
            str_trading_value = "N/A"
            str_trading_fee = "0"
            str_trading_tax = "0"
            str_extra_insurance_fee = "N/A"
            str_per_trading_total_cost = "N/A"
            str_stock_dividend = "N/A"
            str_cash_dividend = "N/A"

        list_data = [ str_year + "-" + str_month_date + " " + str_weekday, #交易日期
                      str_trading_type,             #交易種類
                      str_trading_price,            #交易價格
                      str_trading_quantity,         #交易股數
                      str_trading_value,            #交易金額
                      str_trading_fee + " / " + str_trading_tax, #手續費 / 交易稅
                      str_per_trading_total_cost,   #應收付
                      str_stock_dividend,           #全部股票股利 / 每股股票股利
                      str_cash_dividend,            #全部現金股利 / 每股現金股利
                      str_extra_insurance_fee,      #補充保費
                      str_accumulated_cost,         #累計成本
                      str_accumulated_quantity,     #庫存股數
                      str_accumulated_average_cost, #累計平均成本
                      str_selling_profit ]          #單筆損益  
        return list_data

    def update_button_enable_disable_status( self ): 
        if self.str_picked_stock_number is None or self.ui.qtTabWidget.currentWidget().objectName() not in self.dict_all_account_all_stock_trading_data:
            self.ui.qtAddTradingDataPushButton.setEnabled( False )
            self.ui.qtAddRegularTradingDataPushButton.setEnabled( False )
            self.ui.qtAddDividendDataPushButton.setEnabled( False )
            self.ui.qtAddLimitBuyingDataPushButton.setEnabled( False )
            self.ui.qtAddCapitalReductionDataPushButton.setEnabled( False )
            self.ui.qtAddStockSplitDataPushButton.setEnabled( False )
            self.ui.qtExportSelectedStockTradingDataPushButton.setEnabled( False )
        else:
            self.ui.qtAddTradingDataPushButton.setEnabled( True )
            self.ui.qtAddRegularTradingDataPushButton.setEnabled( True )
            self.ui.qtAddLimitBuyingDataPushButton.setEnabled( True )
            self.ui.qtAddCapitalReductionDataPushButton.setEnabled( True )
            self.ui.qtAddStockSplitDataPushButton.setEnabled( True )
            self.ui.qtExportSelectedStockTradingDataPushButton.setEnabled( True )
            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()

            dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
            if self.str_picked_stock_number in dict_per_account_all_stock_trading_data:
                list_trading_data = dict_per_account_all_stock_trading_data[ self.str_picked_stock_number ]
                e_auto_dividend_type = list_trading_data[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ]
                self.ui.qtAddDividendDataPushButton.setEnabled( e_auto_dividend_type == AutoDividendType.MANUAL or e_auto_dividend_type == AutoDividendType.MANUAL_WITH_AUTO )
            else:
                self.ui.qtAddDividendDataPushButton.setEnabled( False )

    def pick_up_stock( self, str_stock_number ):
        self.str_picked_stock_number = str_stock_number
        if str_stock_number is not None:
            str_stock_name = self.dict_all_company_number_to_name_and_type[ str_stock_number ][ 0 ]
            self.ui.qtCurrentSelectCompanyLabel.setText( f"目前選擇公司：{ str_stock_number } { str_stock_name }" )
            self.ui.qtCurrentSelectCompanyLabel.setStyleSheet("color: yellow; ")
        else:
            self.ui.qtCurrentSelectCompanyLabel.setText( "" )

    #region 從網上下載資料
    def download_all_company_stock_number( self, str_date ): 
        dict_company_number_to_name = {}
        b_need_to_download = False
        if os.path.exists( self.stock_number_file_path ):
            with open( self.stock_number_file_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                for i, row in enumerate( data ):
                    if i == 0:
                        if row.strip() != str_date:
                            if share_api.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                                b_need_to_download = True
                        else:
                            break
                    else:
                        ele = row.strip().split( ',' )
                        if len( ele ) == 2:
                            b_need_to_download = True
                            break

                        dict_company_number_to_name[ ele[ 0 ] ] = [ ele[ 1 ], ele[ 2 ] ]
        else:
            b_need_to_download = True

        if b_need_to_download:
            tds = Download.download_company_stock_number()
            if len( tds ) == 0:
                return
            
            # 確保目錄存在，若不存在則遞歸創建
            os.makedirs( os.path.dirname( self.stock_number_file_path ), exist_ok = True )
            with open( self.stock_number_file_path, 'w', encoding='utf-8' ) as f:
                f.write( str_date + '\n' )
                for row in tds:
                    f.write( str( row[ 0 ] ) + ',' + str( row[ 1 ] ) + ',' + str( row[ 2 ] ) + '\n' )
                    dict_company_number_to_name.pop( row[ 0 ], None )
    
                for key_number, value_name_and_etf in dict_company_number_to_name.items(): #保留已下市公司的資料
                    f.write( str( key_number ) + ',' + str( value_name_and_etf[ 0 ] ) + ',' + str( value_name_and_etf[ 1 ] ) + '\n' )

    def load_all_company_stock_number( self ):
        dict_company_number_to_name = {}
        if os.path.exists( self.stock_number_file_path ):
            with open( self.stock_number_file_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                for i, row in enumerate( data ):
                    if i == 0:
                        continue
                    else:
                        ele = row.strip().split( ',' )
                        dict_company_number_to_name[ ele[ 0 ] ] = [ ele[ 1 ], ele[ 2 ] ]
        return dict_company_number_to_name
        
    def download_all_suspend_company_stock_number( self, str_date ):
        dict_suspend_company_number_to_name = {}
        b_need_to_download = False
        if os.path.exists( self.suspend_stock_number_file_path ):
            with open( self.suspend_stock_number_file_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                for i, row in enumerate( data ):
                    if i == 0:
                        if row.strip() != str_date:
                            if share_api.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                                b_need_to_download = True
                        else:
                            break
                    else:
                        ele = row.strip().split( ',' )
                        dict_suspend_company_number_to_name[ ele[ 0 ] ] = [ ele[ 1 ], ele[ 2 ], ele[ 3 ] ]
        else:
            b_need_to_download = True

        if b_need_to_download:
            tds = Download.download_suspend_company_stock_number()
            if len( tds ) == 0:
                return
            
            # 確保目錄存在，若不存在則遞歸創建
            os.makedirs( os.path.dirname( self.suspend_stock_number_file_path ), exist_ok = True )
            with open( self.suspend_stock_number_file_path, 'w', encoding='utf-8' ) as f:
                f.write( str_date + '\n' )
                for row in tds:
                    f.write( str( row[ 0 ] ) + ',' + str( row[ 1 ] ) + ',' + str( row[ 2 ] ) + ',' + str( row[ 3 ] ) + '\n' )
                    dict_suspend_company_number_to_name.pop( row[ 0 ], None )
    
                for key_number, value_name_and_etf in dict_suspend_company_number_to_name.items():
                    f.write( str( key_number ) + ',' + str( value_name_and_etf[ 0 ] ) + ',' + str( value_name_and_etf[ 1 ] ) + ',' + str( value_name_and_etf[ 2 ] ) + '\n' )

    def load_all_suspend_company_stock_number( self ):
        dict_suspend_company_number_to_name = {}
        if os.path.exists( self.suspend_stock_number_file_path ):
            with open( self.suspend_stock_number_file_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                for i, row in enumerate( data ):
                    if i == 0:
                        continue
                    else:
                        ele = row.strip().split( ',' )
                        dict_suspend_company_number_to_name[ ele[ 0 ] ] = [ ele[ 1 ], ele[ 2 ] ]
        return dict_suspend_company_number_to_name

    def download_day_stock_price( self, str_date, stock_price_file_path ):
        b_need_to_download = False
        if os.path.exists( stock_price_file_path ):
            with open( stock_price_file_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                if data[ 0 ].strip() != str_date or data[ 1 ].strip() != 'O':
                    if share_api.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                        b_need_to_download = True
        else:
            b_need_to_download = True

        if b_need_to_download:
            all_stock_price = []
            b_get_listed_company_stock_price = False
            listed_company_stock_price = Download.download_listed_stock_price_by_date( str_date )
            if len( listed_company_stock_price ) != 0:
                b_get_listed_company_stock_price = True
                all_stock_price.extend( listed_company_stock_price )
            
            b_get_OTC_company_stock_price = False
            OTC_company_stock_price = Download.download_OTC_stock_price_by_date( str_date )
            if len( OTC_company_stock_price ) != 0:
                b_get_OTC_company_stock_price = True
                all_stock_price.extend( OTC_company_stock_price )

            b_get_ROTC_company_stock_price = False
            ROTC_company_stock_price = Download.download_ROTC_stock_price_by_date( str_date )
            if len( ROTC_company_stock_price ) != 0:
                b_get_ROTC_company_stock_price = True
                all_stock_price.extend( ROTC_company_stock_price )
            

            if len( all_stock_price ) == 0:
                print( "no data" )
                return False 
            
            # 確保目錄存在，若不存在則遞歸創建
            os.makedirs( os.path.dirname( stock_price_file_path ), exist_ok = True )
            with open( stock_price_file_path, 'w', encoding='utf-8' ) as f:
                f.write( str_date + '\n' )
                if b_get_listed_company_stock_price and b_get_OTC_company_stock_price and b_get_ROTC_company_stock_price:
                    f.write( 'O\n' )
                else:
                    f.write( 'X\n' )
                for row in all_stock_price:
                    f.write( str( row[ 0 ] ) + ',' + str( row[ 1 ] ) + ',' + str( row[ 2 ] ) + '\n' )
        return True
    
    def load_day_stock_price( self ):
        dict_company_number_to_price_info = {}
        pre_data = []
        if os.path.exists( self.stock_pre_price_file_path ):#先下載前一天的股價
            with open( self.stock_pre_price_file_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                for i, row in enumerate( data ):
                    if i == 0:
                        str_date = row.strip()
                    elif i > 1:
                        ele = row.strip().split( ',' )
                        dict_company_number_to_price_info[ ele[ 0 ] ] = ele[ 2 ]
                        pre_data.append( ele[ 0 ] )
        today_data = []
        obj_today_date = datetime.datetime.today()
        str_today_date = obj_today_date.strftime( "%Y%m%d" )
        b_display_old_data = False
        if os.path.exists( self.stock_price_file_path ):#若當天股價已經出來的話就覆蓋過去
            with open( self.stock_price_file_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                for i, row in enumerate( data ):
                    if i == 0:
                        str_date = row.strip()
                        b_display_old_data = str_date == str_today_date
                    elif i > 1:
                        ele = row.strip().split( ',' )
                        dict_company_number_to_price_info[ ele[ 0 ] ] = ele[ 2 ]
                        today_data.append( ele[ 0 ] )
        list_within_pre_data_but_without_today_data = []
        if b_display_old_data:
            list_within_pre_data_but_without_today_data = list( set( pre_data ) - set( today_data ) )


        return [ str_date, dict_company_number_to_price_info, list_within_pre_data_but_without_today_data ]

    def process_output_file_path( self, str_output_path, list_file_exist, str_folder_name, str_file_name, n_year, n_season, b_overwrite, b_unit_test ):
        str_season = ''
        if n_season == 1 or n_season == 2 or n_season == 3 or n_season == 4:
            str_season = '_Q' + str( n_season )
        if str_output_path == None:
            if b_unit_test:
                str_output_path = os.path.join( g_exe_dir, 'StockInventory', 'UnitTestData', str_file_name + str( n_year ) + str_season + '.txt' )
            else:
                str_output_path = os.path.join( g_data_dir, 'StockInventory', str_folder_name, str_file_name + str( n_year ) + str_season + '.txt' )
        # 確保目錄存在，若不存在則遞歸創建
        os.makedirs( os.path.dirname( str_output_path ), exist_ok = True )
        if b_overwrite or not os.path.exists( str_output_path ):
            list_file_exist[0] = False
        
        return str_output_path

    def download_general_company_all_yearly_dividend_data( self, n_dividend_data_start_year, str_date ):
        print( "\033[32m>>>>>>>>>>>>>>> Start to download all yearly dividend data.\033[0m" )
        current_date = datetime.datetime.today()
        n_current_year = current_date.year

        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911

            b_overwrite = False
            file_exist = [ True ]
            str_output_path = self.process_output_file_path( None, file_exist, 'Dividend', 'GeneralCompanyDividend_', n_year, 0, b_overwrite, False )
            if not file_exist[0] or n_year == n_current_year:
                self.download_general_company_yearly_dividend_data( n_year, str_date, str_output_path, True )
                print(f"Finish {n_year} yearly dividend " )

        print( "\033[32m<<<<<<<<<<<<<<< Finish downloading all yearly dividend data.\033[0m" )

    def download_general_company_yearly_dividend_data( self, n_year, str_date, str_output_path, b_overwrite ):
        # 假如是西元，轉成民國
        if n_year > 1990:
            n_year -= 1911

        file_exist = [ True ]
        str_output_path = self.process_output_file_path( str_output_path, file_exist, 'Dividend', 'GeneralCompanyDividend_', n_year, 0, b_overwrite, False )
        if file_exist[0]:
            print("dividend file exists")
            return

        b_need_to_download = False
        if os.path.exists( str_output_path ):
            with open( str_output_path, 'r', encoding='utf-8' ) as f:
                date = f.readline().strip()
                if date != str_date:
                    if share_api.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                        b_need_to_download = True
        else:
            b_need_to_download = True

        if b_need_to_download:
            all_company_dividend = Download.download_general_company_dividend_by_year( n_year )
            
            if len( all_company_dividend ) == 0:
                print( "no data" )
                return
            with open( str_output_path, 'w', encoding = 'utf-8' ) as f:
                f.write( str_date + '\n' )
                f.write( str( n_year ) + '\n' )
                f.write( '[0]公司代號,[1]公司名稱,[2]股利所屬期間,[3]權利分派基準日,[4]股票股利_盈餘轉增資配股(元/股),[5]股票股利_法定盈餘公積、資本公積轉增資配股(元/股),\
                          [6]股票股利_除權交易日,[7]現金股利_盈餘分配之股東現金股利(元/股),[8]現金股利_法定盈餘公積、資本公積發放之現金(元/股),[9]現金股利_特別股配發現金股利(元/股),\
                          [10]現金股利_除息交易日,[11]現金股利_現金股利發放日,[12]現金增資總股數(股),[13]現金增資認股比率(%),[14]現金增資認購價(元/股),\
                          [15]公告日期,[16]公告時間,[17]普通股每股面額\n' )
                for row in all_company_dividend:
                    b_first = True
                    for ele in row:
                        if b_first:
                            f.write( str( ele ) )
                            b_first = False
                        else:
                            f.write( ',' + str( ele ) )
                    f.write( '\n')  # 用逗號分隔每個元素，並換行

    def load_general_company_all_yearly_dividend_data( self, n_dividend_data_start_year, b_unit_test ):
        current_date = datetime.datetime.today()
        n_current_year = current_date.year
        dict_stock_yearly_dividned = {}
        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911
            list_yearly_dividend = self.read_general_company_yearly_dividend_raw_data( n_year, b_unit_test )
            if list_yearly_dividend != None:
                for index, item in enumerate( list_yearly_dividend ):
                    f_stock_dividend_per_share = self.get_value_from_string( item[4] ) + self.get_value_from_string( item[5] )
                    str_stock_dividend_date = item[6]
                    f_cash_dividend_per_share = self.get_value_from_string( item[7] ) + self.get_value_from_string( item[8] ) + self.get_value_from_string( item[9] )
                    if ( self.get_value_from_string( item[7] ) + self.get_value_from_string( item[8] ) ) != 0 and self.get_value_from_string( item[9] ) != 0 :
                        assert False, f"不該執行到這裡！"
                    str_cash_dividend_date = item[10]
                    str_cash_dividend_distribute_date = item[11]
                    str_year_month_date = ''

                    if ( f_stock_dividend_per_share != 0 and str_stock_dividend_date != '' ) and \
                       ( f_cash_dividend_per_share != 0 and str_cash_dividend_date != '' ):
                        #同時有現金股利和股票股利
                        
                        if str_stock_dividend_date == str_cash_dividend_date:
                            list_year_month_date = str_stock_dividend_date.split( '/' )
                            str_year = str( int( list_year_month_date[0] ) + 1911 )
                            str_year_month_date = str_year + '-' + list_year_month_date[1] + '-' + list_year_month_date[2]
                        else:
                            assert False, f"不該執行到這裡！"
                    elif ( f_stock_dividend_per_share != 0  and str_stock_dividend_date != '' ):
                        #只有股票股利
                        list_year_month_date = str_stock_dividend_date.split( '/' )
                        str_year = str( int( list_year_month_date[0] ) + 1911 )
                        str_year_month_date = str_year + '-' + list_year_month_date[1] + '-' + list_year_month_date[2]
                        pass
                    elif ( f_cash_dividend_per_share != 0 and str_cash_dividend_date != '' ):
                        #只有現金股利
                        list_year_month_date = str_cash_dividend_date.split( '/' )
                        str_year = str( int( list_year_month_date[0] ) + 1911 )
                        str_year_month_date = str_year + '-' + list_year_month_date[1] + '-' + list_year_month_date[2]
                        pass
                    
                    if str_year_month_date != '':
                        dict_dividend_data = Utility.generate_trading_data( str_year_month_date,              #交易日期
                                                                            TradingType.DIVIDEND,             #交易種類
                                                                            TradingPriceType.PER_SHARE,       #交易價格種類
                                                                            0,                                #每股交易價格
                                                                            0,                                #總交易價格
                                                                            0,                                #交易股數
                                                                            TradingFeeType.VARIABLE,          #手續費種類
                                                                            1,                                #手續費折扣
                                                                            0,                                #手續費最低金額
                                                                            0,                                #手續費固定金額
                                                                            DividendValueType.PER_SHARE,      #股利金額種類
                                                                            f_stock_dividend_per_share,       #股票股利
                                                                            f_cash_dividend_per_share,        #現金股利
                                                                            -1,                               #自訂的補充保費
                                                                            0,                                #每股減資金額
                                                                            CapitalReductionType.CASH_RETURN, #減資種類
                                                                            False )                           #是否為當沖交易
                        dict_dividend_data[ TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE ] = True

                        if item[0] in dict_stock_yearly_dividned:
                            dict_stock_yearly_dividned[ item[0] ].append( dict_dividend_data )
                        else:
                            dict_stock_yearly_dividned[ item[0] ] = [ dict_dividend_data ]

        return dict_stock_yearly_dividned

    def read_general_company_yearly_dividend_raw_data( self, n_year, b_unit_test ):
        if n_year > 1990:
            n_year -= 1911
        file_exist = [ True ]
        file_path = self.process_output_file_path( None, file_exist, 'Dividend', 'GeneralCompanyDividend_', n_year, 0, False, b_unit_test )
        if not os.path.exists( file_path ):
            return None
        
        list_all_company_dividend = []
        with open( file_path, 'r', encoding = 'utf-8' ) as f:
            data = f.readlines()
            for i, row in enumerate( data ):
                if i == 0 or i == 1 or i == 2:#i=0 檔案下載日期, i=1 資料年度, i=2 欄位名稱
                    continue
                else:
                    row = row.strip().split( ',' )
                    list_all_company_dividend.append( row )
            print( "Read " + 'StockDividend_Y' + str( n_year ) )

        return list_all_company_dividend

    def get_value_from_string( self, str_value ):
        if str_value == '' or str_value == '--':
            return 0
        return Decimal( str_value )

    def download_listed_etf_all_yearly_dividend_data( self, n_dividend_data_start_year, str_date ):
        print( "\033[32m>>>>>>>>>>>>>>> Start to download listed etf all yearly dividend data.\033[0m" )
        current_date = datetime.datetime.today()
        n_current_year = current_date.year

        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911

            b_overwrite = False
            file_exist = [ True ]
            str_output_path = self.process_output_file_path( None, file_exist, 'Dividend', 'ListedETFDividend_', n_year, 0, b_overwrite, False )
            if not file_exist[0] or n_year == n_current_year:
                self.download_listed_etf_yearly_dividend_data( n_year, str_date, str_output_path, True )
                print(f"Finish Listed etf {n_year} yearly dividend " )

        print( "\033[32m<<<<<<<<<<<<<<< Finish downloading listed etf all yearly dividend data.\033[0m" )

    def download_listed_etf_yearly_dividend_data( self, n_year, str_date, str_output_path, b_overwrite ):
        # 假如是西元，轉成民國
        if n_year > 1990:
            n_year -= 1911

        file_exist = [ True ]
        str_output_path = self.process_output_file_path( str_output_path, file_exist, 'Dividend', 'ListedETFDividend_', n_year, 0, b_overwrite, False )
        if file_exist[0]:
            print("dividend file exists")
            return

        b_need_to_download = False
        if os.path.exists( str_output_path ):
            with open( str_output_path, 'r', encoding='utf-8' ) as f:
                date = f.readline().strip()
                if date != str_date:
                    if share_api.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                        b_need_to_download = True
        else:
            b_need_to_download = True

        if n_year < 1990:
            n_year += 1911

        if b_need_to_download:
            try:
                json_value = Download.download_listed_etf_dividend_by_year( n_year )
                with open( str_output_path, 'w', encoding='utf-8' ) as f:
                    f.write( str_date + '\n' )
                    # 證券代號	
                    # 證券簡稱	
                    # 除息交易日	
                    # 收益分配基準日	
                    # 收益分配發放日	
                    # 收益分配金額 (每1受益權益單位)	
                    # 收益分配標準 (102年度起啟用)	
                    # 公告年度
                    json.dump( json_value[ 'data' ], f, ensure_ascii=False, indent=4 )
            except Exception as e:
                print(f"Final error: {e}")
            pass

    def load_listed_etf_all_yearly_dividend_data( self, n_dividend_data_start_year, b_unit_test ):
        current_date = datetime.datetime.today()
        n_current_year = current_date.year
        dict_stock_yearly_dividned = {}
        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911
            list_yearly_dividend = self.read_listed_etf_yearly_dividend_raw_data( n_year, b_unit_test )
            if list_yearly_dividend != None:
                for index, item in enumerate( list_yearly_dividend ):
                    if item[5] != None:
                        try:
                            value = float( item[5].strip() )
                            if value > 0:
                                # 字串是數值且大於 0
                                str_stock_number = item[ 0 ]
                                taiwan_date_str  = item[ 2 ]
                                taiwan_year = int(taiwan_date_str .split('年')[0]) + 1911
                                str_year_month_date = f"{taiwan_year}-{taiwan_date_str.split('年')[1].replace('月', '-').replace('日', '')}"

                                f_cash_dividend_per_share = Decimal( item[ 5 ] )
                                dict_dividend_data = Utility.generate_trading_data( str_year_month_date,              #交易日期
                                                                                    TradingType.DIVIDEND,             #交易種類
                                                                                    TradingPriceType.PER_SHARE,       #交易價格種類
                                                                                    0,                                #每股交易價格
                                                                                    0,                                #總交易價格 
                                                                                    0,                                #交易股數
                                                                                    TradingFeeType.VARIABLE,          #手續費種類
                                                                                    1,                                #手續費折扣
                                                                                    0,                                #手續費最低金額
                                                                                    0,                                #手續費固定金額
                                                                                    DividendValueType.PER_SHARE,      #股利金額種類
                                                                                    0,                                #股票股利
                                                                                    f_cash_dividend_per_share,        #現金股利
                                                                                    -1,                               #自訂的補充保費
                                                                                    0,                                #每股減資金額
                                                                                    CapitalReductionType.CASH_RETURN, #減資種類
                                                                                    False )                           #是否為當沖交易
                                
                                dict_dividend_data[ TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE ] = True
                                if str_stock_number in dict_stock_yearly_dividned:
                                    dict_stock_yearly_dividned[ str_stock_number ].append( dict_dividend_data )
                                else:
                                    dict_stock_yearly_dividned[ str_stock_number ] = [ dict_dividend_data ]
                                pass
                        except ValueError:
                            # 不是有效的數值
                            pass

        return dict_stock_yearly_dividned

    def read_listed_etf_yearly_dividend_raw_data( self, n_year, b_unit_test ):
        if n_year > 1990:
            n_year -= 1911
        file_exist = [ True ]
        file_path = self.process_output_file_path( None, file_exist, 'Dividend', 'ListedETFDividend_', n_year, 0, False, b_unit_test )
        if not os.path.exists( file_path ):
            return None
        
        with open( file_path, 'r', encoding = 'utf-8' ) as f:
            date = f.readline()
            list_data = json.load( f )

            return list_data
        return []

    def download_OTC_etf_all_yearly_dividend_data( self, n_dividend_data_start_year, str_date ):
        print( "\033[32m>>>>>>>>>>>>>>> Start to download OTC etf all yearly dividend data.\033[0m" )
        current_date = datetime.datetime.today()
        n_current_year = current_date.year

        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911

            b_overwrite = False
            file_exist = [ True ]
            str_output_path = self.process_output_file_path( None, file_exist, 'Dividend', 'OTCETFDividend_', n_year, 0, b_overwrite, False )
            if not file_exist[0] or n_year == n_current_year:
                self.download_OTC_etf_yearly_dividend_data( n_year, str_date, str_output_path, True )
                print(f"Finish OTC etf {n_year} yearly dividend " )

        print( "\033[32m<<<<<<<<<<<<<<< Finish downloading OTC etf all yearly dividend data.\033[0m" )

    def download_OTC_etf_yearly_dividend_data( self, n_year, str_date, str_output_path, b_overwrite ):
        # 假如是西元，轉成民國
        if n_year > 1990:
            n_year -= 1911

        file_exist = [ True ]
        str_output_path = self.process_output_file_path( str_output_path, file_exist, 'Dividend', 'OTCETFDividend_', n_year, 0, b_overwrite, False )
        if file_exist[0]:
            print("dividend file exists")
            return

        b_need_to_download = False
        if os.path.exists( str_output_path ):
            with open( str_output_path, 'r', encoding='utf-8' ) as f:
                date = f.readline().strip()
                if date != str_date:
                    if share_api.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                        b_need_to_download = True
        else:
            b_need_to_download = True

        if b_need_to_download:
            try:
                json_value = Download.download_OTC_etf_dividend_by_year( n_year )
                with open( str_output_path, 'w', encoding='utf-8' ) as f:
                    f.write( str_date + '\n' )
                    # 民國104年之前(包含104年)   
                    # 除權息日期                 
                    # 代號	                    
                    # 名稱	                    
                    # 除權息前收盤價             	
                    # 除權息參考價               	
                    # 權值                      	
                    # 息值                      	
                    # 權值＋息值	             
                    # 權/息                   
                    # 漲停價	
                    # 跌停價	
                    # 開始交易基準價	
                    # 減除股利參考價	
                    # 現金股利	
                    # 每仟股無償配股	
                    # 員工紅利轉增資	
                    # 現金增資股數	
                    # 現金增資認購價	
                    # 公開承銷股數	
                    # 員工認購股數	
                    # 原股東認購股數	
                    # 按持股比例仟股認購

                    # 民國105年之後
                    # 除權息日期                 
                    # 代號	                    
                    # 名稱	                    
                    # 除權息前收盤價             	
                    # 除權息參考價               	
                    # 權值                      	
                    # 息值                      	
                    # 權值＋息值	             
                    # 權/息                   
                    # 漲停價	
                    # 跌停價	
                    # 開始交易基準價	
                    # 減除股利參考價	
                    # 現金股利	
                    # 每仟股無償配股	
                    # 缺---------------員工紅利轉增資	
                    # 現金增資股數	
                    # 現金增資認購價	
                    # 公開承銷股數	
                    # 員工認購股數	
                    # 原股東認購股數	
                    # 按持股比例仟股認購
                    json.dump( json_value[ 'tables' ][0]['data'], f, ensure_ascii=False, indent=4 )

            except Exception as e:
                print(f"Final error: {e}")
        pass

    def load_OTC_etf_all_yearly_dividend_data( self, n_dividend_data_start_year, b_unit_test ):
        current_date = datetime.datetime.today()
        n_current_year = current_date.year
        dict_stock_yearly_dividned = {}
        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911
            list_yearly_dividend = self.read_OTC_etf_yearly_dividend_raw_data( n_year, b_unit_test )
            if list_yearly_dividend != None:
                for index, item in enumerate( list_yearly_dividend ):
                    if item[13] != None:
                        try:
                            value = float( item[ 13 ].strip() )
                            if value > 0:
                                # 字串是數值且大於 0
                                str_stock_number = item[ 1 ]
                                taiwan_date_str  = item[ 0 ]
                                # taiwan_year = int(taiwan_date_str .split('年')[0]) + 1911
                                taiwan_year = int(taiwan_date_str .split('/')[0]) + 1911
                                str_year_month_date = f"{taiwan_year}-{taiwan_date_str.split('/')[1]}-{taiwan_date_str.split('/')[2]}"

                                str_cash_dividend = item[ 13 ].strip()
                                f_cash_dividend_per_share = Decimal( str_cash_dividend )
                                dict_dividend_data = Utility.generate_trading_data( str_year_month_date,              #交易日期
                                                                                    TradingType.DIVIDEND,             #交易種類
                                                                                    TradingPriceType.PER_SHARE,       #交易價格種類
                                                                                    0,                                #每股交易價格
                                                                                    0,                                #總交易價格
                                                                                    0,                                #交易股數
                                                                                    TradingFeeType.VARIABLE,          #手續費種類
                                                                                    1,                                #手續費折扣
                                                                                    0,                                #手續費最低金額
                                                                                    0,                                #手續費固定金額
                                                                                    DividendValueType.PER_SHARE,      #股利金額種類
                                                                                    0,                                #股票股利
                                                                                    f_cash_dividend_per_share,        #現金股利
                                                                                    -1,                               #自訂的補充保費
                                                                                    0,                                #每股減資金額
                                                                                    CapitalReductionType.CASH_RETURN, #減資種類
                                                                                    False )                           #是否為當沖交易
                                
                                dict_dividend_data[ TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE ] = True
                                if str_stock_number in dict_stock_yearly_dividned:
                                    dict_stock_yearly_dividned[ str_stock_number ].append( dict_dividend_data )
                                else:
                                    dict_stock_yearly_dividned[ str_stock_number ] = [ dict_dividend_data ]
                                pass
                        except ValueError:
                            # 不是有效的數值
                            pass

        return dict_stock_yearly_dividned
    
    def read_OTC_etf_yearly_dividend_raw_data( self, n_year, b_unit_test ):
        if n_year > 1990:
            n_year -= 1911
        file_exist = [ True ]
        file_path = self.process_output_file_path( None, file_exist, 'Dividend', 'OTCETFDividend_', n_year, 0, False, b_unit_test )
        if not os.path.exists( file_path ):
            return None
        
        with open( file_path, 'r', encoding = 'utf-8' ) as f:
            date = f.readline()
            list_data = json.load( f )

            return list_data
        return []
    
    def download_listed_company_all_yearly_split_data( self, n_split_data_start_year, str_date ):
        print( "\033[32m>>>>>>>>>>>>>>> Start to download all yearly split data.\033[0m" )
        current_date = datetime.datetime.today()
        n_current_year = current_date.year

        for n_year in range( n_split_data_start_year, n_current_year + 1 ):
            b_overwrite = False
            file_exist = [ True ]
            str_output_path = self.process_output_file_path( None, file_exist, 'SplitMerge', 'ListedCompanySplit_', n_year, 0, b_overwrite, False )
            if not file_exist[0] or n_year == n_current_year:
                self.download_listed_company_yearly_split_data( n_year, str_date, str_output_path, True )
                print(f"Finish {n_year} yearly split " )

        print( "\033[32m<<<<<<<<<<<<<<< Finish downloading all yearly split data.\033[0m" )
    
    def download_listed_company_yearly_split_data( self, n_year, str_date, str_output_path, b_overwrite ):
        file_exist = [ True ]
        str_output_path = self.process_output_file_path( str_output_path, file_exist, 'SplitMerge', 'ListedCompanySplit_', n_year, 0, b_overwrite, False )
        if file_exist[0]:
            print("split file exists")
            return

        b_need_to_download = False
        if os.path.exists( str_output_path ):
            with open( str_output_path, 'r', encoding='utf-8' ) as f:
                date = f.readline().strip()
                if date != str_date:
                    if share_api.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                        b_need_to_download = True
        else:
            b_need_to_download = True

        if b_need_to_download:
            try:
                json_value = Download.download_listed_stock_split_by_year( n_year, n_year )
                if json_value != None:
                    with open( str_output_path, 'w', encoding='utf-8' ) as f:
                        f.write( str_date + '\n' )
                        json.dump( json_value['data'], f, ensure_ascii=False, indent=4 )

            except Exception as e:
                print(f"Final error: {e}")
        pass

    def download_listed_etf_all_yearly_split_data( self, n_split_data_start_year, str_date ):
        print( "\033[32m>>>>>>>>>>>>>>> Start to download all yearly split data.\033[0m" )
        current_date = datetime.datetime.today()
        n_current_year = current_date.year

        for n_year in range( n_split_data_start_year, n_current_year + 1 ):
            b_overwrite = False
            file_exist = [ True ]
            str_output_path = self.process_output_file_path( None, file_exist, 'SplitMerge', 'ListedEtfSplit_', n_year, 0, b_overwrite, False )
            if not file_exist[0] or n_year == n_current_year:
                self.download_listed_etf_yearly_split_data( n_year, str_date, str_output_path, True )
                print(f"Finish {n_year} yearly split " )

        print( "\033[32m<<<<<<<<<<<<<<< Finish downloading all yearly split data.\033[0m" )
    
    def download_listed_etf_yearly_split_data( self, n_year, str_date, str_output_path, b_overwrite ):
        file_exist = [ True ]
        str_output_path = self.process_output_file_path( str_output_path, file_exist, 'SplitMerge', 'ListedEtfSplit_', n_year, 0, b_overwrite, False )
        if file_exist[0]:
            print("split file exists")
            return

        b_need_to_download = False
        if os.path.exists( str_output_path ):
            with open( str_output_path, 'r', encoding='utf-8' ) as f:
                date = f.readline().strip()
                if date != str_date:
                    if share_api.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                        b_need_to_download = True
        else:
            b_need_to_download = True

        if b_need_to_download:
            try:
                json_value = Download.download_listed_etf_split_merge_by_year( n_year, n_year )
                if json_value != None:
                    with open( str_output_path, 'w', encoding='utf-8' ) as f:
                        f.write( str_date + '\n' )
                        json.dump( json_value['data'], f, ensure_ascii=False, indent=4 )

            except Exception as e:
                print(f"Final error: {e}")
        pass
    #endregion

def run_app():
    app = QApplication( sys.argv )
    app.setStyle('Fusion')
    window = MainWindow( False, 
                        'TradingData.json',
                        'UISetting.config',
                        'StockNumber.txt',
                        'SuspendStockNumber.txt',
                        'StockPrice.txt',
                        'PreStockPrice.txt',
                        'StockDividendPositionDate.json' )

    window.show()
    return app.exec()

if __name__ == "__main__":
    sys.exit( run_app() )

#region 工具備忘

#打包指令
# cd D:\_2.code\PythonStockPrice   
# pyinstaller --hidden-import "babel.numbers" --add-data "resources;./resources" --add-data "StockInventory/Dividend;./StockInventory/Dividend" --add-data "../FoxInfoShareUtility/foxinfo_share_utility/icons;foxinfo_share_utility/icons" --noconsole StockPriceMainWindow.py
# pyinstaller --hidden-import "babel.numbers" --add-data "resources;./resources" --add-data "StockInventory/Dividend;./StockInventory/Dividend" --add-data "../FoxInfoShareUtility/foxinfo_share_utility/icons;foxinfo_share_utility/icons" --console StockPriceMainWindow.py

# 要把.ui檔變成.py
# cd D:\_2.code\PythonStockPrice
# pyside6-uic QtStockPriceMainWindowTemplate.ui -o QtStockPriceMainWindowTemplate.py
# pyside6-uic QtStockPriceMainWindow.ui -o QtStockPriceMainWindow.py
# pyside6-uic QtStockTradingEditDialog.ui -o QtStockTradingEditDialog.py
# pyside6-uic QtStockRegularTradingEditDialog.ui -o QtStockRegularTradingEditDialog.py
# pyside6-uic QtStockDividendEditDialog.ui -o QtStockDividendEditDialog.py
# pyside6-uic QtStockCapitalReductionEditDialog.ui -o QtStockCapitalReductionEditDialog.py
# pyside6-uic QtStockCapitalIncreaseEditDialog.ui -o QtStockCapitalIncreaseEditDialog.py
# pyside6-uic QtStockSplitEditDialog.ui -o QtStockSplitEditDialog.py
# pyside6-uic QtDuplicateOptionDialog.ui -o QtDuplicateOptionDialog.py
# pyside6-uic QtCashTransferEditDialog.ui -o QtCashTransferEditDialog.py
# pyside6-uic QtStockDividendTransferFeeEditSpinboxDialog.ui -o QtStockDividendTransferFeeEditSpinboxDialog.py
# pyside6-uic QtAboutDialog.ui -o QtAboutDialog.py
# pyside6-uic QtStockDividendPositionDateEditDialog.ui -o QtStockDividendPositionDateEditDialog.py
# pyside6-uic QtShowItemEditDialog.ui -o QtShowItemEditDialog.py
# pyside6-uic QtAccountSettingEditDialog.ui -o QtAccountSettingEditDialog.py


# 下載上市櫃公司股利資料
# https://mopsov.twse.com.tw/mops/web/t108sb27

# 以下兩個網站都可以下載"上市"ETF的股利
# https://www.twse.com.tw/zh/products/securities/etf/products/div.html
# https://www.twse.com.tw/zh/ETFortune/dividendList

# 以下這個網站可以下載"上櫃"ETF的股利
# https://www.tpex.org.tw/zh-tw/announce/market/ex/cal.html

# 下市/下櫃/撤銷登錄興櫃/不繼續公開發行彙總表
# https://mopsov.twse.com.tw/mops/web/t51sb01_1

# 採彈性面額(每股面額非新台幣10元)公司專區
# https://mopsov.twse.com.tw/mops/web/t51sb09



# 證交所  上市公司變更股票面額恢復買賣參考價格( 分割 )
# https://www.twse.com.tw/zh/announcement/change/twtb8u.html

# 櫃買中心  上櫃公司變更股票面額恢復買賣參考價格( 分割 )
# https://www.tpex.org.tw/zh-tw/announce/market/change/reference.html
# https://www.tpex.org.tw/zh-tw/mainboard/listed/flexible-face-value.html 這個好像更完整

# 證交所  上市ETF "分割" 及 "反分割" 恢復買賣參考價格
# https://www.twse.com.tw/zh/announcement/split/twtcau.html

# 櫃買中心  上櫃ETF "分割" 恢復買賣參考價格 *目前尚無發生過
# https://www.tpex.org.tw/zh-tw/announce/market/etf-split/reference.html

# 櫃買中心  上櫃ETF "反分割" 恢復買賣參考價格 *目前尚無發生過
# https://www.tpex.org.tw/zh-tw/announce/market/etf-rev-split/reference.html


# 證交所  上市公司股票減資 恢復買賣參考價格
# https://www.twse.com.tw/zh/announcement/reduction/twtauu.html
# 櫃買中心  上市公司股票減資 恢復買賣參考價格
# https://www.tpex.org.tw/zh-tw/announce/market/reduction/reference.html


# 靜態掃描
# pylint --disable=all --enable=E1120,E1121 StockPriceMainWindow.py 只顯示參數數量錯誤
# pylint -E StockPriceMainWindow.py 只顯示錯誤級別的資訊
#endregion
